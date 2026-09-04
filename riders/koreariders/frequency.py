# -*- coding: utf-8 -*-
"""선구별 열차종별 운행횟수 — how many trains a day run over each section.

`6. 운전` sheets 2(5)–2(7) break the network into 117 sections and give
trains/day on each, by train type. That is the one published thing that speaks
directly to what happens at a junction, because the count *changes* there and
the change is the service stepping on or off:

    경부선  서울-금천구청   새마을 26 + 무궁화 40 = 66
            금천구청-의왕   26 + 40 = 66
            의왕-천안      26 + 40 = 66
            천안-조치원     20 + 31 = 51      <- 15 trains leave at 천안
    장항선  천안-신창       6 + 9   = 15      <- and here they are

Flat from 서울 to 천안 means nothing joins or leaves in between, which is how
we know a junction step at 용산 is wrong however well it fits the 승하차.

    python frequency.py            # print every line's section profile
"""

import re
import sys

import openpyxl

import lines as LN

MEMBER = "1.지역간철도/6. 운전(1~6)_완.xlsx"
SHEETS = ("2(5)", "2(6)", "2(7)")

# The sheets name their lines a little differently from the passenger tables,
# and a few of ours are too new to appear at all (강릉선, 중부내륙선) -- those
# simply get no constraint.
ALIAS = {
    "수서평택선": "수서고속선",
    "경의1선": "경의선",
}

# Sheet column heading -> the yearbook passenger types it covers. 전동차 is
# 광역전철 and has no 승하차 row of its own, so it is left out; ITX-청춘 is a
# different service from ITX-새마을 and only runs on 경춘선.
COLUMNS = {
    "KTX": ["KTX"],
    "SRT": ["SRT"],
    "고속열차": ["KTX", "SRT"],
    "새마을": ["새마을", "ITX-새마을"],
    "ITX-청춘": ["ITX-새마을"],
    "무궁화": ["무궁화"],
    "통근": ["통근"],
}


def _flat(v):
    return re.sub(r"\s+", "", str(v)) if v is not None else ""


def sections():
    """[(line, from, to, {train type: trains per day}), ...] in sheet order."""
    wb = openpyxl.load_workbook(LN._open(MEMBER), data_only=True)
    out = []
    for name in SHEETS:
        ws = wb[name]
        # 2(5) puts 선로용량 one column left of the other two, so find the
        # headings rather than assuming where they sit.
        head = {}
        for c in range(1, 20):
            for r in (4, 5):
                h = _flat(ws.cell(row=r, column=c).value)
                if h in COLUMNS:
                    head[c] = h
        line = ""
        for r in range(6, ws.max_row + 1):
            a = _flat(ws.cell(row=r, column=1).value)
            if a:
                line = ALIAS.get(a, a)
            sec = _flat(ws.cell(row=r, column=3).value)
            if not sec or "-" not in sec:
                continue
            runs = {}
            for c, h in head.items():
                v = ws.cell(row=r, column=c).value
                if isinstance(v, (int, float)) and v:
                    for k in COLUMNS[h]:
                        runs[k] = runs.get(k, 0) + int(v)
            frm, to = sec.split("-", 1)
            out.append((line, frm, to, runs))
    return out


_BY_LINE = None


def by_line():
    """{line: [(from, to, runs), ...]}, sections kept in their published order.

    Cached -- opening the workbook takes long enough that doing it once per
    line turns a report into a coffee break.
    """
    global _BY_LINE
    if _BY_LINE is None:
        out = {}
        for line, frm, to, runs in sections():
            out.setdefault(line, []).append((frm, to, runs))
        _BY_LINE = out
    return _BY_LINE


def changes(canon, kinds):
    """{station: (before, after)} trains a day either side of a boundary.

    Only interior boundaries — where one section ends and the next begins — say
    anything about a junction. A station that is not a boundary has no step at
    all, which is the useful half of this: it forbids one.
    """
    secs = by_line().get(canon)
    if not secs:
        return None
    want = set(kinds)
    tot = [sum(n for k, n in runs.items() if k in want) for _, _, runs in secs]
    out = {}
    for i in range(len(secs) - 1):
        if secs[i][1] == secs[i + 1][0]:        # contiguous, so a real boundary
            out[secs[i][1]] = (tot[i], tot[i + 1])
    return out


def match(nm, ch):
    """Look a chain station up in a `changes()` table.

    The sheets abbreviate — 대전조 for 대전조차장 — so fall back to the longest
    key that prefixes the name. Longest matters: both 대전 and 대전조 prefix
    대전조차장 and only one of them is the junction.
    """
    if nm in ch:
        return ch[nm]
    best = None
    for k in ch:
        if nm.startswith(k) and (best is None or len(k) > len(best)):
            best = k
    return ch[best] if best else None


def main():
    table, _ = LN.resolve()
    known = by_line()
    print("%-11s %-9s %8s %8s %8s   %s"
          % ("line", "at", "before", "after", "step", "as a share"))
    print("-" * 66)
    for canon in sorted(table):
        spec = table[canon]
        if "error" in spec or canon not in known:
            continue
        ch = changes(canon, spec["types"])
        for nm, (a, b) in sorted(ch.items()):
            mark = "" if a != b else "   (no step -- forbids one)"
            print("%-11s %-9s %8d %8d %+8d   %s%s"
                  % (canon, nm, a, b, b - a,
                     "%.0f%%" % (100.0 * (b - a) / a) if a else "-", mark))
    missing = [c for c in sorted(table)
               if "error" not in table[c] and c not in known]
    print("\nno section data: %s" % ", ".join(missing))


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
