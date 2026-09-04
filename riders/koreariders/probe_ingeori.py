# -*- coding: utf-8 -*-
"""What does the yearbook's 선별 인거리 actually measure?

It matters because japanriders anchors each segment's magnitude to a published
per-line figure, and the plan here was to do the same with 선별 인거리 /
영업거리. That only works if 인거리 is attributed to the *track* a passenger
rode over. If instead a whole journey is charged to every line it touches, the
figure is not a density and cannot anchor anything.

The test: divide each line's 인거리 by its 통과인원 to get the average distance
per user, and compare that with the line's own length. A connector a few km long
whose average user apparently rides 30 km is not measuring track.

    python probe_ingeori.py
"""

import io
import os
import sys
import zipfile

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
YEARBOOK = os.path.join(HERE, "data", "korail_yearbook_2022_excel.zip")

# 영업거리 (km) from 8. 시설 table 4, for the lines worth showing.
LENGTH = {
    "경부고속선": 398.3, "호남고속선": 182.3, "경부선": 441.7, "호남선": 252.5,
    "전라선": 180.4, "장항선": 154.4, "중앙선": 331.4, "경전선": 277.7,
    "동해선": 244.4, "영동선": 188.9, "태백선": 85.8, "충북선": 113.2,
    "경북선": 115.0, "정선선": 38.7, "경원선": 94.4, "광주선": 14.9,
    "대구선": 29.0, "강릉선": 120.7, "오송선": 4.2, "익산북연결": 3.4,
    "광주송북연": 2.2, "행신선": 3.4, "시흥연결선": 3.9, "대전북연결": 4.2,
    "건천연결선": 4.2, "미전선": 1.6, "부전선": 2.2, "가야선": 3.9,
}


def read(sheet, col):
    with zipfile.ZipFile(YEARBOOK) as z:
        target = None
        for i in z.infolist():
            n = i.filename
            if not (i.flag_bits & 0x800):
                n = n.encode("cp437").decode("cp949")
            if n == "1.지역간철도/4. 수송(여객)_완.xlsx":
                target = i.filename
        raw = z.read(target)
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[sheet]
    out, total = {}, None
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row,
                            min_col=1, max_col=2, values_only=True):
        if not row[0] or not str(row[0]).strip():
            continue
        k = str(row[0]).strip()
        v = 0.0 if row[1] in (None, "-", "") else float(row[1])
        if k == "합계":
            total = v
        else:
            out[k] = v
    wb.close()
    return out, total


def main():
    passing, passing_total = read("5", 2)     # 통과인원
    pkm, pkm_total = read("4", 2)             # 인거리

    print("stated 합계   통과인원 %15.0f   인거리 %18.0f"
          % (passing_total, pkm_total))
    print("sum of rows   통과인원 %15.0f   인거리 %18.0f"
          % (sum(passing.values()), sum(pkm.values())))
    print("rows / 합계             %6.2fx                    %6.2fx"
          % (sum(passing.values()) / passing_total, sum(pkm.values()) / pkm_total))

    print("\n%-12s %14s %16s %9s %9s %7s"
          % ("line", "통과인원", "인거리", "평균km", "영업거리", "비율"))
    print("-" * 74)
    for k in sorted(pkm, key=lambda x: -pkm[x]):
        if pkm[k] <= 0 or passing.get(k, 0) <= 0:
            continue
        avg = pkm[k] / passing[k]
        L = LENGTH.get(k)
        ratio = ("%7.1f" % (avg / L)) if L else "      ?"
        print("%-12s %14.0f %16.0f %9.1f %9s %s"
              % (k, passing[k], pkm[k], avg, ("%.1f" % L) if L else "?", ratio))


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
