# -*- coding: utf-8 -*-
"""The yearbook, parsed, plus the one thing it cannot supply: which name means
which line.

The 철도통계연보 names its lines three different ways depending on which table
you are in -- `경부고속본선` in the distance table, `경부고속` in the station
roster, `경부고속선` in the traffic table -- and OSM has a fourth opinion
(`경부본선` alongside `경부선` on the same route). None of the four is derivable
from the others, so LINES below is a hand table, in the manner of japanriders'
build_names.py. Everything else here is read out of the workbook.

    python lines.py          # print the resolved table and what is missing
"""

import io
import os
import sys
import zipfile

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
D = os.path.join(HERE, "data")
YEARBOOK = os.path.join(D, "korail_yearbook_2022_excel.zip")

PASSENGER = "1.지역간철도/4. 수송(여객)_완.xlsx"
FACILITY = "1.지역간철도/8. 시설_완.xlsx"

# canonical -> (traffic-table name, distance-table name, roster name, OSM names)
# A None means that table has no row for the line; the pipeline copes as long as
# the distance and OSM entries are there.
LINES = {
    "경부고속선": ("경부고속선", "경부고속본선", "경부고속", ["경부고속선"]),
    "호남고속선": ("호남고속선", "호 남 고 속 본 선", "호남고속", ["호남고속선"]),
    "수서고속선": ("수서고속선", "수서평택 고속선", "수서평택선", ["수서평택고속선"]),
    "경부선": ("경부선", "경부선", "경부선", ["경부선", "경부본선"]),
    "호남선": ("호남선", "호남선", "호남선", ["호남선"]),
    "전라선": ("전라선", "전라선", "전라선", ["전라선"]),
    "장항선": ("장항선", "장항선", "장항선", ["장항선"]),
    "중앙선": ("중앙선", "중앙선", "중앙선", ["중앙선"]),
    "경전선": ("경전선", "경전선", "경전선", ["경전선"]),
    "동해선": ("동해선", "동해선", "동해선", ["동해선", "동해본선"]),
    "영동선": ("영동선", "영동선", "영동선", ["영동선", "영동본선"]),
    "태백선": ("태백선", "태백선", "태백선", ["태백선"]),
    "충북선": ("충북선", "충북선", "충북선", ["충북선"]),
    "경북선": ("경북선", "경북선", "경북선", ["경북선"]),
    "경원선": ("경원선", "경원선", "경원선", ["경원선"]),
    "경의선": ("경의선", "경의선", "경의선", ["경의선"]),
    "경춘선": ("경춘선", "경춘선", None, ["경춘선"]),
    "광주선": ("광주선", "광주선", "광주선", ["광주선"]),
    "대구선": ("대구선", "대구선", "대구선", ["대구선"]),
    "정선선": ("정선선", "정선선", "정선선", ["정선선"]),
    "강릉선": ("강릉선", "경강선(원주-강릉)", "경강선(강릉선)", ["경강선"]),
    "중부내륙선": ("중부내륙선", "중부내륙선", "중부내륙선", ["중부내륙선"]),
    "삼척선": (None, "삼척선", "삼척선", ["삼척선"]),
    "진해선": (None, "진해선", "진해선", ["진해선"]),
    "여천선": (None, "여천선", "여천선", ["여천선"]),
    "문경선": (None, "문경선", "문경선", ["문경선"]),
}

# The distance table's endpoints are the legal extent of the line, which is not
# always where trains run to or what OSM calls the place. Overrides win.
ENDS = {
    # 영동선 legally ends at a signal box; 강릉 is where the track and the
    # station data stop.
    "영동선": ("영주", "강릉"),
    # 중앙선's 모량 end is a junction with 동해선 and has no station; 경주 is the
    # last station with traffic.
    "중앙선": ("청량리", "경주"),
    # The distance table still says 장항, but 장항선 has run through to 익산
    # since 군산선 was absorbed -- the roster and station table both agree.
    "장항선": ("천안", "익산"),
    # 광주선's own table calls the junction end 동송정, which is a signal point.
    "광주선": ("광주송정", "광주"),
    # 강릉선 branches off 중앙선 at 서원주, not at 원주 station.
    "강릉선": ("서원주", "강릉"),
    # 동해선's 부산진 end is a freight junction; 부전 is the passenger terminus.
    "동해선": ("부전", "영덕"),
    # 호남선 legally starts at 대전조차장, which is not a passenger station and
    # has no OSM node; 서대전 is the first one trains call at.
    "호남선": ("서대전", "목포"),
    # 태백선 ends at 백산 junction, again no station; 태백 is the last stop.
    "태백선": ("제천", "태백"),
    # 대구선 starts at 가천 junction, no OSM node; 동대구 is where trains start.
    "대구선": ("동대구", "영천"),
    # 수서평택고속선 ends at a junction with 경부고속선, not at 평택 station.
    "수서고속선": ("수서", "평택지제"),
    # 경부고속선 keeps 서울 even though its own metals start south of 광명: the
    # corridor search bridges 서울~광명 over 경부선 at a penalty, and 서울 is the
    # line's largest single source of traffic -- dropping it cost a third of the
    # reconstruction.
    "경부고속선": ("서울", "부산"),
}

# Which train types run on each line. The point is the parallel pairs: 경부선 and
# 경부고속선 share 서울, 대전, 동대구 and 부산, and sheet 8's combined counts
# cannot tell which set of metals a passenger rode. Only the high-speed services
# use the 고속선, so the type split separates them.
#
# It does not separate everything. 서울's KTX arrivals are 경부, 호남, 전라 and
# 강릉 KTX together, and this table hands all of them to every line that lists
# KTX -- see README.md.
ALL_TYPES = ["KTX", "SRT", "새마을", "ITX-새마을", "무궁화", "통근"]
CONVENTIONAL = ["새마을", "ITX-새마을", "무궁화", "통근"]
TYPES = {
    "경부고속선": ["KTX", "SRT"],
    "호남고속선": ["KTX", "SRT"],
    "수서고속선": ["SRT"],
    "강릉선": ["KTX"],
    "중부내륙선": ["KTX"],
    "경부선": CONVENTIONAL,
    "호남선": CONVENTIONAL,
    "경의선": CONVENTIONAL,
    "경원선": CONVENTIONAL,
    "경춘선": CONVENTIONAL,
    "충북선": CONVENTIONAL,
    "경북선": CONVENTIONAL,
    "대구선": CONVENTIONAL,
    "정선선": CONVENTIONAL,
    "태백선": CONVENTIONAL,
    "영동선": CONVENTIONAL,
    "장항선": CONVENTIONAL,
    # These carry high-speed services over their own conventional metals.
    "전라선": ALL_TYPES,
    "중앙선": ALL_TYPES,
    "경전선": ALL_TYPES,
    "동해선": ALL_TYPES,
    "광주선": ALL_TYPES,
}


def _open(member):
    with zipfile.ZipFile(YEARBOOK) as z:
        for i in z.infolist():
            n = i.filename
            if not (i.flag_bits & 0x800):
                n = n.encode("cp437").decode("cp949")
            if n == member:
                return io.BytesIO(z.read(i.filename))
    raise SystemExit("no %s in the yearbook zip" % member)


def _num(v):
    return 0.0 if v in (None, "-", "") or isinstance(v, str) else float(v)


def station_flows():
    """역별 승하차 -> {station: (하행승차, 하행하차, 상행승차, 상행하차)}."""
    wb = openpyxl.load_workbook(_open(PASSENGER), read_only=True, data_only=True)
    ws = wb["8"]
    out = {}
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row,
                            min_col=2, max_col=10, values_only=True):
        st = row[0]
        if not st or not str(st).strip() or str(st).strip() == "합계":
            continue
        v = [_num(x) for x in row[1:]]
        out[str(st).strip()] = (v[0], v[1], v[4], v[5])
    wb.close()
    return out


# Sheet 9 carries KTX and SRT in one table, split by a 열차종 label in column A
# and shifted one column right of the others; sheets 10-13 are one train type
# each, station name in column A.
TYPE_SHEETS = {"9": None, "10": "새마을", "11": "ITX-새마을",
               "12": "무궁화", "13": "통근"}


def station_flows_by_type():
    """{train type: {station: (하행승차, 하행하차, 상행승차, 상행하차)}}.

    Sheet 8 is the total over all train types, which is what breaks the trunk
    lines: 서울's arrivals are KTX riding 경부고속선 *and* 무궁화 riding 경부선,
    and cumulating the sum along either one is meaningless. Splitting by train
    type separates the parallel pair, since only the high-speed services use the
    고속선.
    """
    wb = openpyxl.load_workbook(_open(PASSENGER), read_only=True, data_only=True)
    out = {}
    for sheet, fixed in TYPE_SHEETS.items():
        ws = wb[sheet]
        col = 2 if fixed is None else 1        # sheet 9 has 열차종 in column A
        kind = fixed
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row,
                                min_col=1, max_col=col + 7, values_only=True):
            if fixed is None and row[0] and str(row[0]).strip():
                kind = str(row[0]).strip()
            st = row[col - 1]
            if not st or not str(st).strip() or str(st).strip() == "합계":
                continue
            v = [_num(x) for x in row[col:col + 7]]
            out.setdefault(kind, {})[str(st).strip()] = (v[0], v[1], v[4], v[5])
    wb.close()
    return out


def line_passing():
    """선별 통과인원 -> {traffic-table name: 명/년}."""
    wb = openpyxl.load_workbook(_open(PASSENGER), read_only=True, data_only=True)
    ws = wb["5"]
    out = {}
    for row in ws.iter_rows(min_row=9, max_row=ws.max_row,
                            min_col=1, max_col=2, values_only=True):
        if row[0] and str(row[0]).strip():
            out[str(row[0]).strip()] = _num(row[1])
    wb.close()
    return out


def rosters():
    """노선 -> {station, ...} from the facility table."""
    wb = openpyxl.load_workbook(_open(FACILITY), read_only=True, data_only=True)
    ws = wb["2"]
    out, cur = {}, None
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row,
                            min_col=3, max_col=4, values_only=True):
        ln, st = row
        if ln and str(ln).strip():
            cur = str(ln).strip()
        if cur and st and str(st).strip():
            out.setdefault(cur, set()).add(str(st).strip())
    wb.close()
    return out


def distances():
    """영업선로별 철도거리 -> {name: (from, to, km)}, km summed over track types."""
    wb = openpyxl.load_workbook(_open(FACILITY), read_only=True, data_only=True)
    ws = wb["4"]
    out = {}
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row,
                            min_col=1, max_col=11, values_only=True):
        name = row[0]
        if not name or not str(name).strip():
            continue
        name = " ".join(str(name).split())
        a, b = row[3], row[5]
        km = sum(_num(x) for x in row[7:11])
        if a and b and km > 0:
            out[name] = ("".join(str(a).split()), "".join(str(b).split()), km)
    wb.close()
    return out


def bad_anchor(station, roster, flows):
    """Can this end station anchor the reconstruction?

    The anchor asserts that everything alights here in 하행 and boards here in
    상행, which is only true at a station belonging to this line and nowhere
    else. The roster assigns every station exactly one *home* line, so an end
    station whose home is some other line is a junction shared with it -- 익산's
    home is 호남선, and its 591k arrivals are mostly 호남선's, not 장항선's.
    A station with no 승하차 row at all cannot anchor anything either.
    """
    if station not in flows:
        return "no 승하차 row"
    if roster and station not in roster:
        return "belongs to another line"
    return None


def resolve():
    """Everything the build needs, per canonical line name.

    `last` is the end the reconstruction anchors on, so a true terminus is
    preferred; if only one end is a junction the pair is swapped to put the
    clean one last. `clean_end` false means both ends are junctions and the
    profile's level will have to be solved rather than measured.
    """
    flows, passing = station_flows(), line_passing()
    by_type = station_flows_by_type()
    rost, dist = rosters(), distances()

    out = {}
    for canon, (fname, dname, rname, osm) in LINES.items():
        if dname not in dist:
            out[canon] = {"error": "no distance row named %r" % dname}
            continue
        a, b, km = dist[dname]
        a, b = ENDS.get(canon, (a, b))
        roster = rost.get(rname, set()) if rname else set()
        # Sum only the train types that use this line's metals.
        kinds = TYPES.get(canon, ALL_TYPES)
        lf = {}
        for k in kinds:
            for st, v in by_type.get(k, {}).items():
                p = lf.get(st, (0.0, 0.0, 0.0, 0.0))
                lf[st] = tuple(p[i] + v[i] for i in range(4))
        bad_a, bad_b = (bad_anchor(a, roster, lf), bad_anchor(b, roster, lf))
        # Put the usable end last, since that is the one the anchor reads.
        if bad_b and not bad_a:
            a, b, bad_a, bad_b = b, a, bad_b, bad_a
        out[canon] = {
            "first": a, "last": b, "length_km": km,
            "clean_end": bad_b is None, "why": bad_b,
            "ways": osm,
            "roster": roster,
            "passing": passing.get(fname, 0.0) if fname else 0.0,
            "types": kinds,
            "flows": lf,
        }
    return out, flows


def main():
    table, flows = resolve()
    print("%-11s %9s %8s %-9s %-9s %8s %12s"
          % ("line", "length", "anchor", "first", "last", "roster", "통과인원"))
    print("-" * 74)
    for canon in LINES:
        r = table[canon]
        if "error" in r:
            print("%-11s  %s" % (canon, r["error"]))
            continue
        print("%-11s %8.1f %8s %-9s %-9s %8d %12.0f  %s"
              % (canon, r["length_km"], "clean" if r["clean_end"] else "junction",
                 r["first"][:9], r["last"][:9], len(r["roster"]), r["passing"],
                 r["why"] or ""))


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
