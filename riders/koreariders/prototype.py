# -*- coding: utf-8 -*-
"""Reconstruct per-segment passenger load on one intercity line, from the
철도통계연보's station table, and check it against the same yearbook's line
totals.

Korea publishes no 輸送密度 equivalent -- no single throughput figure per rail
segment. What it does publish, in `4. 수송(여객)` sheet 8, is **역별 여객 승하차
실적 split 상행 / 하행**: for every station, how many people boarded and
alighted in each direction. Cumulating that along a line gives the load on each
segment, which is the same quantity Japan publishes outright.

The awkward part is junctions. A junction station's counts mix every line
through it, so its net is meaningless for any one of them. The way round is to
cumulate *inward from the far terminus*:

    load(n-1, n) = 하차 at the terminus          (nothing boards there)
    load(i-1, i) = load(i, i+1) - 승차_i + 하차_i  for i = n-1 .. 1

which never reads station 0 -- the junction end -- at all. Through traffic
crossing the junction is not observed, but it does not need to be: it falls out
of the recursion as whatever is left over at the far end of the line.

Two checks decide whether this is any good, neither of which enters the
calculation:

  * the 하행 and 상행 profiles are built from disjoint columns of the source, so
    how closely they mirror each other tests the method rather than the data;
  * 통과인원 -- how many people used the line at all -- is published per line and
    can be rebuilt from the finished profile.

The obvious third check, 선별 인거리 / 영업거리, turns out not to be a check at
all: see probe_ingeori.py. That column is not attributed to the track a
passenger rode over, so it is not a density and cannot anchor anything.

Lines that junction at *both* ends (장항선, 경북선) have no clean terminus to
anchor on. The anchor error is a constant, though -- the shape comes from
interior stations and survives -- so the published 통과인원 pins the one number
per line that is missing. That is the same split japanriders uses: shape from
one source, magnitude from another.

    python prototype.py                # 전라선
    python prototype.py --line 장항선
    python prototype.py --all          # one summary row per line
"""

import argparse
import heapq
import io
import json
import math
import os
import sys
import zipfile

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
D = os.path.join(HERE, "data")
YEARBOOK = os.path.join(D, "korail_yearbook_2022_excel.zip")
RAILWAYS = os.path.join(D, "osm_railways.json")
STATIONS = os.path.join(D, "osm_stations.json")

PASSENGER_XLSX = "1.지역간철도/4. 수송(여객)_완.xlsx"

# One entry per line we can currently reconstruct. `ways` are the OSM
# `railway=rail` name values to assemble the corridor from; `first`/`last` are
# its end stations, and the shortest path between them through the named track
# is what km is measured along. `length_km` is the yearbook's 영업거리
# (8. 시설 table 4), which the OSM chainage is rescaled to match.
#
# `clean_end` says whether `last` is a true terminus. The reconstruction anchors
# on it -- everything alights there in 하행, everything boards there in 상행 --
# so a line that junctions at both ends has no anchor and cannot be done this
# way on its own.
LINES = {
    "전라선": {
        "ways": ["전라선"], "first": "익산", "last": "여수엑스포",
        "length_km": 180.4, "clean_end": True,
    },
    "장항선": {
        "ways": ["장항선"], "first": "천안", "last": "익산",
        "length_km": 154.4, "clean_end": False,
    },
    "경북선": {
        "ways": ["경북선"], "first": "김천", "last": "영주",
        "length_km": 115.0, "clean_end": False,
    },
    "정선선": {
        "ways": ["정선선"], "first": "민둥산", "last": "아우라지",
        "length_km": 38.7, "clean_end": True,
    },
}


# ---------------------------------------------------------------- geometry


def haversine(a, b):
    R = 6371.0088
    la1, lo1, la2, lo2 = map(math.radians, (a[0], a[1], b[0], b[1]))
    h = (math.sin((la2 - la1) / 2) ** 2
         + math.cos(la1) * math.cos(la2) * math.sin((lo2 - lo1) / 2) ** 2)
    return 2 * R * math.asin(math.sqrt(h))


PENALTY = 40.0


def build_graph(ways, target):
    """Track ways -> {node: [(node, cost_km, real_km), ...]}, keyed on coords.

    Chaining ways end to end does not survive contact with a real network:
    stations sit on double track, lines carry sidings and triangles, and a
    greedy chain wanders off down whichever branch it meets first. A graph plus
    a shortest path between the two end stations picks the through route and
    ignores everything hanging off it.

    A line's own named track is not quite continuous either -- station throats
    and junction approaches are tagged with the crossing line's name, or with
    none at all -- so every named way goes into the graph and the ones that are
    not this line's are charged `PENALTY` times their length. That is cheap
    enough to hop a 200 m gap through 익산 station and far too expensive to
    follow 경전선 out of 순천.
    """
    g = {}
    for w in ways:
        mine = (w.get("tags") or {}).get("name") in target
        mult = 1.0 if mine else PENALTY
        pts = [(p["lat"], p["lon"]) for p in w.get("geometry", [])]
        for a, b in zip(pts, pts[1:]):
            ka = (round(a[0], 6), round(a[1], 6))
            kb = (round(b[0], 6), round(b[1], 6))
            if ka == kb:
                continue
            d = haversine(a, b)
            g.setdefault(ka, []).append((kb, d * mult, d, mine))
            g.setdefault(kb, []).append((ka, d * mult, d, mine))
    return g


def nearest_node(g, pt):
    return min(g, key=lambda n: haversine(n, pt))


def shortest_path(g, src, dst):
    dist = {src: 0.0}
    prev = {}
    pq = [(0.0, src)]
    seen = set()
    while pq:
        d, u = heapq.heappop(pq)
        if u in seen:
            continue
        seen.add(u)
        if u == dst:
            break
        for v, cost, real, mine in g.get(u, ()):
            nd = d + cost
            if nd < dist.get(v, 1e18):
                dist[v] = nd
                prev[v] = (u, real, mine)
                heapq.heappush(pq, (nd, v))
    if dst not in dist:
        return None, 0.0
    path, cur, foreign = [dst], dst, 0.0
    while cur != src:
        cur, real, mine = prev[cur]
        if not mine:
            foreign += real
        path.append(cur)
    return path[::-1], foreign


def chainage(line):
    """Cumulative km at each vertex."""
    out = [0.0]
    for i in range(len(line) - 1):
        out.append(out[-1] + haversine(line[i], line[i + 1]))
    return out


def project(pt, line, cum):
    """Nearest point on the polyline: (distance_km, chainage_km)."""
    best = (1e9, 0.0)
    for i in range(len(line) - 1):
        a, b = line[i], line[i + 1]
        # Local flat approximation is ample at these latitudes and segment sizes.
        kx = math.cos(math.radians(a[0]))
        ax, ay = a[1] * kx, a[0]
        bx, by = b[1] * kx, b[0]
        px, py = pt[1] * kx, pt[0]
        dx, dy = bx - ax, by - ay
        L2 = dx * dx + dy * dy
        t = 0.0 if L2 == 0 else max(0.0, min(1.0, ((px - ax) * dx + (py - ay) * dy) / L2))
        qx, qy = ax + t * dx, ay + t * dy
        d = math.hypot(px - qx, py - qy) * 111.32
        if d < best[0]:
            best = (d, cum[i] + t * (cum[i + 1] - cum[i]))
    return best


# ---------------------------------------------------------------- yearbook


def station_flows():
    """역별 여객 승하차 실적 -> {station: (하행승차, 하행하차, 상행승차, 상행하차)}."""
    with zipfile.ZipFile(YEARBOOK) as z:
        name = None
        for i in z.infolist():
            n = i.filename
            if not (i.flag_bits & 0x800):
                n = n.encode("cp437").decode("cp949")
            if n == PASSENGER_XLSX:
                name = i.filename
        raw = z.read(name)
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb["8"]
    out = {}
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row,
                            min_col=2, max_col=10, values_only=True):
        st = row[0]
        if not st or not str(st).strip() or str(st).strip() == "합계":
            continue
        vals = [0 if v in (None, "-", "") else float(v) for v in row[1:]]
        # columns: 하행 승차·하차·계·인거리, 상행 승차·하차·계·인거리
        out[str(st).strip()] = (vals[0], vals[1], vals[4], vals[5])
    wb.close()
    return out


def line_roster():
    """8. 시설 table 2 -> {line: {station, ...}}, the yearbook's own idea of
    which stations belong to which line. Alphabetical, so it settles membership
    but never order."""
    with zipfile.ZipFile(YEARBOOK) as z:
        name = None
        for i in z.infolist():
            n = i.filename
            if not (i.flag_bits & 0x800):
                n = n.encode("cp437").decode("cp949")
            if n == "1.지역간철도/8. 시설_완.xlsx":
                name = i.filename
        raw = z.read(name)
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb["2"]
    out, cur = {}, None
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row,
                            min_col=3, max_col=4, values_only=True):
        ln, st = row
        if ln and str(ln).strip():
            cur = str(ln).strip()
        if cur and st and str(st).strip():
            out.setdefault(cur, set()).add(str(st).strip())
    wb.close()
    return out


def line_totals():
    """선별 여객 수송실적 -> {line: (통과인원, 인거리)} from sheets 5 and 4."""
    with zipfile.ZipFile(YEARBOOK) as z:
        name = None
        for i in z.infolist():
            n = i.filename
            if not (i.flag_bits & 0x800):
                n = n.encode("cp437").decode("cp949")
            if n == PASSENGER_XLSX:
                name = i.filename
        raw = z.read(name)
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    out = {}
    for sheet, idx in (("5", 0), ("4", 1)):
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=9, max_row=ws.max_row,
                                min_col=1, max_col=2, values_only=True):
            if not row[0] or not str(row[0]).strip():
                continue
            k = str(row[0]).strip()
            v = 0.0 if row[1] in (None, "-", "") else float(row[1])
            rec = out.setdefault(k, [0.0, 0.0])
            rec[idx] = v
    wb.close()
    return out


# ---------------------------------------------------------------- main


def build(line_name, quiet=False):
    spec = LINES[line_name]

    with io.open(STATIONS, encoding="utf-8") as f:
        nodes = json.load(f)["elements"]
    named = {}
    for n in nodes:
        nm = (n.get("tags", {}) or {}).get("name")
        if nm:
            named.setdefault(nm, []).append((n["lat"], n["lon"]))

    with io.open(RAILWAYS, encoding="utf-8") as f:
        ways = json.load(f)["elements"]
    target = set(spec["ways"])
    g = build_graph(ways, target)

    for end in ("first", "last"):
        if spec[end] not in named:
            raise SystemExit("no OSM station node named %s" % spec[end])
    # Anchor on a node of the line's *own* track, so the path cannot start by
    # wandering off down a crossing line at a junction station.
    own = build_graph([w for w in ways
                       if (w.get("tags") or {}).get("name") in target], target)
    src = nearest_node(own, named[spec["first"]][0])
    dst = nearest_node(own, named[spec["last"]][0])
    poly, foreign = shortest_path(g, src, dst)
    if poly is None:
        raise SystemExit("%s track is not connected between %s and %s"
                         % (spec["ways"], spec["first"], spec["last"]))
    cum = chainage(poly)
    if not quiet: print("corridor %-14s %6.1f km, %d nodes, %.2f km borrowed from other lines"
          % ("+".join(spec["ways"]), cum[-1], len(poly), foreign))

    # Snap every station node to the corridor; keep the closest node per name.
    snapped = {}
    for n in nodes:
        nm = (n.get("tags", {}) or {}).get("name")
        if not nm:
            continue
        d, km = project((n["lat"], n["lon"]), poly, cum)
        if d > 0.30:
            continue
        if nm not in snapped or d < snapped[nm][0]:
            snapped[nm] = (d, km)

    if spec["first"] not in snapped or spec["last"] not in snapped:
        raise SystemExit("could not snap the line's end stations: %s / %s"
                         % (spec["first"], spec["last"]))
    k0, k1 = snapped[spec["first"]][1], snapped[spec["last"]][1]
    lo, hi = min(k0, k1), max(k0, k1)
    on_line = sorted(((km, nm, d) for nm, (d, km) in snapped.items()
                      if lo - 0.2 <= km <= hi + 0.2))
    if k0 > k1:                       # make km increase in the 하행 direction
        on_line = [(hi - km, nm, d) for km, nm, d in on_line][::-1]
    else:
        on_line = [(km - lo, nm, d) for km, nm, d in on_line]

    # OSM chainage runs a touch short of the published 영업거리; rescale so the
    # segment lengths sum to the yearbook's figure.
    scale = spec["length_km"] / on_line[-1][0]
    on_line = [(km * scale, nm, d) for km, nm, d in on_line]
    if not quiet: print("%d stations snapped over %.1f km (OSM chainage x%.4f to match 영업거리)"
          % (len(on_line), on_line[-1][0], scale))

    flows = station_flows()
    roster = line_roster().get(line_name, set())

    # Snapping by distance alone picks up whatever sits near the track, and at
    # 아산 what sits near the track is 천안아산 -- a 경부고속선 KTX station about
    # 100 m away, whose 720k arrivals would be cumulated into 장항선 as though
    # they had come off a 무궁화. A station that reports traffic but is not on
    # this line's roster is somebody else's and is dropped.
    #
    # The line's own end stations are exempt: a junction like 익산 belongs to
    # 호남선 on the roster but is still where 장항선 physically stops, and
    # dropping it would cut 12 km off the line. Its counts are never read as
    # boardings -- only as the anchor, whose error is a constant the published
    # 통과인원 can absorb.
    ends = {spec["first"], spec["last"]}
    intruders = [nm for _, nm, _ in on_line
                 if nm in flows and roster and nm not in roster and nm not in ends]
    if intruders:
        if not quiet: print("dropped, on the track but not on the %s roster: %s"
              % (line_name, ", ".join(intruders)))
        on_line = [t for t in on_line if t[1] not in set(intruders)]

    missing = [nm for _, nm, _ in on_line if nm not in flows]
    if missing:
        if not quiet: print("no yearbook row for: %s" % ", ".join(missing))

    # A station the chain misses is a hole in the cumulation: its boardings are
    # never added, so every load upstream of it comes out too low. Check the
    # chain against the yearbook's own roster for the line.
    chained = {nm for _, nm, _ in on_line}
    dropped = sorted(roster - chained)
    if dropped:
        if not quiet: print("on the yearbook's %s roster but not snapped to the corridor:" % line_name)
        for nm in dropped:
            f = flows.get(nm)
            traffic = "no 승하차 row" if f is None else "%.0f 명/년 승차" % (f[0] + f[2])
            if not quiet: print("    %-10s %s" % (nm, traffic))
    return spec, on_line, flows, missing


def reconstruct(on_line, flows, down=True):
    """Cumulate inward from the clean end. Returns loads per segment.

    Station 0 is the junction end and station n the terminus, so *both*
    directions must be anchored at n and walked back towards 0 -- anchoring the
    상행 profile at station 0 would read the junction's polluted counts, which
    is the one thing this is built to avoid.

        하행 (0 -> n)   everything alights at n:  load(n-1) = 하차_down(n)
        상행 (n -> 0)   everything boards at n:   load(n-1) = 승차_up(n)

    and in both cases load(i-1) = load(i) -/+ (boarded - alighted) at station i,
    the sign following which way the trains are moving.
    """
    stops = [(km, nm) for km, nm, _ in on_line]
    n = len(stops) - 1

    def bo(nm):
        f = flows.get(nm, (0, 0, 0, 0))
        return (f[0], f[1]) if down else (f[2], f[3])

    loads = [0.0] * n
    b_end, a_end = bo(stops[n][1])
    loads[n - 1] = a_end if down else b_end
    for i in range(n - 1, 0, -1):
        b, a = bo(stops[i][1])
        loads[i - 1] = (loads[i] - b + a) if down else (loads[i] + b - a)
    return loads


def summarise():
    """One row per line: how well it reconstructs, and what it took."""
    rows = []
    for name in LINES:
        spec, on_line, flows, _ = build(name, quiet=True)
        down = reconstruct(on_line, flows, down=True)
        up = reconstruct(on_line, flows, down=False)
        stops = [nm for _, nm, _ in on_line]
        f = lambda nm: flows.get(nm, (0, 0, 0, 0))
        users = down[0] + up[-1] + sum(f(nm)[0] + f(nm)[2] for nm in stops[1:-1])
        tot = line_totals().get(name, [0, 0])[0]
        gap = max(abs(down[i] - up[i]) / max(down[i], up[i], 1)
                  for i in range(len(down)))
        length = on_line[-1][0]
        pkm = sum((down[i] + up[i]) * (on_line[i + 1][0] - on_line[i][0])
                  for i in range(len(down)))
        if not spec["clean_end"] and tot:
            d = (tot - users) / 2.0
            pkm = sum((down[i] + up[i] + 2 * d) * (on_line[i + 1][0] - on_line[i][0])
                      for i in range(len(down)))
        rows.append((name, len(on_line), spec["clean_end"], users, tot,
                     users / tot if tot else 0, 100 * gap, pkm / length / 365.0))

    print("\n%-8s %5s %7s %12s %12s %7s %8s %9s"
          % ("line", "stops", "anchor", "통과인원", "yearbook", "ratio",
             "mirror", "수송밀도"))
    print("-" * 76)
    for name, n, clean, users, tot, ratio, gap, density in rows:
        print("%-8s %5d %7s %12.0f %12.0f %7.3f %7.1f%% %8.0f"
              % (name, n, "clean" if clean else "junction", users, tot,
                 ratio, gap, density))
    print("\n'ratio' is the reconstruction's own 통과인원 over the published one;")
    print("for junction-anchored lines it is what gets solved away, so only the")
    print("clean-anchor rows are a test. 'mirror' is the worst disagreement")
    print("between the independently built 하행 and 상행 profiles.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--line", default="전라선")
    ap.add_argument("--all", action="store_true",
                    help="one summary row per configured line")
    args = ap.parse_args()

    if args.all:
        summarise()
        return

    spec, on_line, flows, missing = build(args.line)
    down = reconstruct(on_line, flows, down=True)
    up = reconstruct(on_line, flows, down=False)

    print("\n%-12s %7s %9s %9s %9s %9s" % ("station", "km", "하행승차", "하행하차",
                                           "상행승차", "상행하차"))
    print("-" * 60)
    for km, nm, _ in on_line:
        f = flows.get(nm)
        if f is None:
            print("%-12s %7.1f %9s %9s %9s %9s" % (nm, km, "-", "-", "-", "-"))
        else:
            print("%-12s %7.1f %9.0f %9.0f %9.0f %9.0f" % (nm, km, f[0], f[1], f[2], f[3]))

    print("\n%-14s %8s   %9s %9s %9s" % ("segment", "km", "하행", "상행", "합계/일"))
    print("-" * 56)
    pkm = 0.0
    for i in range(len(on_line) - 1):
        (k0, a, _), (k1, b, _) = on_line[i], on_line[i + 1]
        seg = k1 - k0
        both = down[i] + up[i]
        pkm += both * seg
        print("%-14s %8.1f   %9.0f %9.0f %9.0f"
              % ((a + "-" + b)[:14], seg, down[i], up[i], both / 365.0))

    length = on_line[-1][0]
    print("-" * 56)
    print("reconstructed 인거리        %15.0f 인-km/년   (on this line's track)" % pkm)
    print("reconstructed 수송밀도      %15.0f 명/일" % (pkm / length / 365.0))

    # 통과인원 -- how many people used the line at all -- is the one line-level
    # figure in the yearbook that is a plain count rather than an attribution.
    # Everyone who used the line either entered at one end or boarded at an
    # interior station, so it can be rebuilt from the profile:
    #
    #   통과인원 = down[0] + up[n-1] + interior boardings, both directions
    #
    # The two end stations contribute only through the loads, never through
    # their own 승차, which at a junction is mostly people leaving along some
    # other line entirely.
    stops = [nm for _, nm, _ in on_line]
    f = lambda nm: flows.get(nm, (0, 0, 0, 0))
    interior = sum(f(nm)[0] + f(nm)[2] for nm in stops[1:-1])
    users = down[0] + up[-1] + interior

    tot = line_totals().get(args.line)
    print("\nreconstructed 통과인원      %15.0f 명/년" % users)
    if tot:
        print("yearbook 통과인원           %15.0f 명/년" % tot[0])
        print("ratio                       %15.3f" % (users / tot[0]))

    # The mirror test: the two directions are reconstructed from disjoint
    # columns of the source, so how closely they track is a check on the method
    # rather than on the data.
    gap = max(abs(down[i] - up[i]) / max(down[i], up[i], 1)
              for i in range(len(down)))
    print("하행/상행 worst disagreement %14.2f %%" % (100 * gap))

    if not spec["clean_end"] and tot:
        # No true terminus, so the anchor -- "everything alights here" -- reads a
        # junction's whole traffic, not this line's share, and lifts the profile
        # by a constant. The differences between adjacent segments still come
        # from interior stations and are unaffected, so the shape is intact and
        # only one number per direction is wrong. 통과인원 moves by exactly 2d
        # when both profiles shift by d, which pins it.
        d = (tot[0] - users) / 2.0
        print("\nboth ends are junctions, so the level is unanchored.")
        print("shifting both profiles by %.0f 명/년 to meet the published 통과인원:" % d)
        adj_down = [x + d for x in down]
        adj_up = [x + d for x in up]
        if min(min(adj_down), min(adj_up)) < 0:
            print("   ... which drives a segment negative -- the shape is wrong too.")
        pkm2 = sum((adj_down[i] + adj_up[i]) * (on_line[i + 1][0] - on_line[i][0])
                   for i in range(len(adj_down)))
        print("   %-14s %9s -> %9s  (%.0f 명/일)"
              % ("busiest seg", "%.0f" % (down[0] + up[0]),
                 "%.0f" % (adj_down[0] + adj_up[0]),
                 (adj_down[0] + adj_up[0]) / 365.0))
        i = min(range(len(adj_down)), key=lambda k: adj_down[k] + adj_up[k])
        print("   %-14s %9s -> %9s  (%.0f 명/일)"
              % ("quietest seg", "%.0f" % (down[i] + up[i]),
                 "%.0f" % (adj_down[i] + adj_up[i]),
                 (adj_down[i] + adj_up[i]) / 365.0))
        print("   수송밀도       %15.0f -> %.0f 명/일"
              % (pkm / length / 365.0, pkm2 / length / 365.0))


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
