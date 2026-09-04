# -*- coding: utf-8 -*-
"""Turn the KRIC and 인천교통공사 timetables into rows the pipeline already reads.

Writes data/timetable_extra.csv with exactly the columns of the 서울교통공사
timetable_raw.csv, so build_stations.py and build.py gain fifteen lines by
reading a second file rather than by growing a second code path.

Three things have to be sorted out on the way in.

**The KRIC file carries two row shapes.** Korail's lines (수인분당, 경의중앙,
경춘, 서해, 경강) write one row per station stop, like Seoul's file. The metro
operators write one row per *train*, with the whole station list and all the
times packed into single cells -- and each operator picked its own conventions.
All five are handled in parse_packed:

    신분당선     D19-광교+D18-광교중앙        D19-10:17+D18-10:21     '+', pair on '-'
    우이신설선    001-신설동+002-보문          001-:+002-5:31          '+', ':' means none
    김포골드라인   001-장기역+002-운양역        001-5:26:10+002-5:28:39 '+', names end in 역
    인천1호선     3125-예술회관,3126-인천터미널  3125-5:32:00,…          ',' throughout
    의정부       001-발곡+002-회룡            001-05:00/002-05:01     '/' between pairs
    공항철도     001-서울+002-인천공항1터미널   001+06:00+002+06:45     '+' for *both*

**KRIC truncates station names.** To about three characters, and it prefixes 신
where a name is already used elsewhere on the network. The OD is the naming
authority here -- it is what the ridership is keyed on -- so every stop is
resolved back to an OD name by resolve(), which tries exact match, then prefix,
then the 신 strip, and only then a hand-written alias. Anything it cannot
resolve is reported rather than guessed at, and the run ends by checking that
every station the OD has on a line was actually reached.

**인천2호선 has no times in the KRIC file**, only a headway, so it comes from
인천교통공사's own four CSVs -- one per direction per day type, wide format with
a column per station.

    python kric.py
"""

import collections
import csv
import io
import os
import re
import sys

import openpyxl

import lines as LR

HERE = os.path.dirname(os.path.abspath(__file__))
D = os.path.join(HERE, "data")
KRIC = os.path.join(D, "kric_urbanrail_timetable.xlsx")
OD = os.path.join(D, "od_2023-12-31.csv")
OUT = os.path.join(D, "timetable_extra.csv")

COLUMNS = ["고유번호", "호선", "역사코드", "역사명", "주중주말", "방향",
           "급행여부", "열차코드", "열차도착시간", "열차출발시간",
           "출발역", "도착역"]

# Synthetic 역사코드 for the new lines. Five digits keeps them clear of the
# four-digit 서울교통공사 codes and of the three-digit 역번호 in the hourly file,
# which build_stations.py compares against.
CODE_BASE = 70001

# Station names no rule can get to: initialisms and contractions rather than
# truncations. Keep this list short -- if it grows, the rule is wrong.
ALIAS = {
    "디엠시": "디지털미디어시티",
    "세종릉": "세종대왕릉",
    "경광주": "경기광주",
    "시흥청": "시흥시청",
    "도예촌": "신둔도예촌",
    "로데오": "압구정로데오",
    "남동인": "남동인더스파크",
    # KRIC keeps the old 경의선 name; the station sits between 강매 and 수색,
    # which is where 화전 is, and the OD has no 항공대 at all.
    "항공대": "화전",
    "센럴파크": "센트럴파크",   # a typo in the source, one stop
}

# KRIC 요일구분 -> the 주중주말 codes the Seoul timetable uses. A label naming
# more than one kind of day is emitted for each. 명절 (new year / chuseok) is
# not a day we ever draw.
DAYMAP = [
    ("평일", "DAY"),
    ("토요일", "SAT"), ("토+", "SAT"), ("주말", "SAT"),
    ("휴일", "END"), ("공휴일", "END"), ("일요일", "END"), ("주말", "END"),
]


def norm(s):
    """Cells arrive as str, int or None depending on the operator."""
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s).strip())


def base(s):
    """잠실(송파구청) -> 잠실, and settle the two middle dots on one.

    The parenthetical is not always last -- 김포골드라인 writes
    사우(김포시청)역 -- so strip it wherever it sits and let the resolver deal
    with the 역 that is left behind.
    """
    s = norm(s).replace("·", ".").replace("．", ".")
    return re.sub(r"\([^)]*\)", "", s)


def daytypes(label):
    lab = norm(label)
    if "명절" in lab:
        return []
    out = []
    for needle, code in DAYMAP:
        if needle in lab and code not in out:
            out.append(code)
    return out


def parse_time(tok):
    """'5:31', '05:31:00', ':' or blank -> seconds, or None."""
    tok = (tok or "").strip()
    if not tok or set(tok) <= set(": .-"):
        return None
    m = re.fullmatch(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", tok)
    if not m:
        return None
    h, mi, se = int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
    if mi > 59 or se > 59:
        return None
    return h * 3600 + mi * 60 + se


def fmt_time(secs):
    """Back to HH:MM:SS, letting the hour run past 24 the way Seoul's file does."""
    return "%02d:%02d:%02d" % (secs // 3600, secs % 3600 // 60, secs % 60)


def parse_packed(cell):
    """A packed cell -> {station code: value}. Handles all five conventions.

    Every operator separates entries with one of '+', ',' or '/'. Most then
    join the code to its value with '-'; 공항철도 alone uses its separator for
    that too, giving code, value, code, value, ... which is what the second
    branch is for.
    """
    if cell is None:
        return {}
    s = str(cell).strip()
    if not s:
        return {}
    toks = [t for t in re.split(r"[+,/]", s) if t.strip()]
    if not toks:
        return {}
    paired = [t for t in toks if "-" in t]
    if len(paired) >= max(1, len(toks) - 1):
        out = {}
        for t in toks:
            if "-" not in t:
                continue
            code, val = t.split("-", 1)
            out[code.strip()] = val.strip()
        return out
    if len(toks) % 2 == 0:
        return dict((toks[i].strip(), toks[i + 1].strip())
                    for i in range(0, len(toks), 2))
    return {}


def running_order(stops):
    """Put a packed train's stops in the order it actually calls at them.

    The packed cells list stations in the *line's* fixed order -- 신림선 always
    writes 001-샛강 first -- so the half of the trains running the other way
    have times that count down the cell. Decide by majority, because a single
    backward step is what crossing midnight looks like and must not flip the
    train.
    """
    ts = [(d if d is not None else a) for _, a, d in stops]
    up = down = 0
    prev = None
    for t in ts:
        if t is None:
            continue
        if prev is not None:
            if t > prev:
                up += 1
            elif t < prev:
                down += 1
        prev = t
    return stops[::-1] if down > up else stops


def spans_midnight(times):
    """True if a trip's stop times straddle 00:00.

    Korail's rows write after-midnight stops as 00:12 rather than 24:12, and
    those rows arrive unordered, so sorting them by face value turns a
    50-minute run into a 23-hour one. Same guard build.py applies to the Seoul
    file, needed here one step earlier because we do the grouping.
    """
    real = [t for t in times if t is not None]
    return bool(real) and max(real) - min(real) > 12 * 3600


def lift(v):
    """Small hours of a midnight-spanning trip, moved above 24:00."""
    return None if v is None else (v + 86400 if v < 4 * 3600 else v)


# The longest believable gap between two consecutive stops, and the only test
# applied to a stop time. It settles both questions at once:
#
#   backwards -- is this midnight, or a bad value? Lifting 00:05 over midnight
#     after 23:50 leaves 15 minutes, which is a train; lifting 05:32 after
#     17:37 leaves 12 hours, which is not.
#   forwards  -- 김포골드라인 stamps 5:32:39 on station 005 of every weekday
#     train, evidently the first train's time pasted down the column. On an
#     evening train that reads as a jump backwards, on a small-hours one as a
#     five-hour jump forwards, and only checking one direction leaves half of
#     them in.
#
# The bound is generous: the longest real hop is 공항철도 직통 running 서울 to
# 인천공항1터미널 non-stop in about 45 minutes.
MAX_HOP_S = 90 * 60


def monotonic(times):
    """Make a train's stop times run forward, unwrapping across midnight once.

    Treating every backward step as a wrap is what turned a single mistyped
    stop into a train arriving at 96:33 -- each bad value added another day to
    everything after it. A backward step is only a wrap if lifting it over
    midnight leaves a believable gap to the stop before; otherwise the time is
    bad and the stop is dropped rather than allowed to shift its neighbours.

    A run that starts before 03:00 belongs to the previous service day -- the
    Seoul file writes those as 24:xx and up, and the rest of the pipeline
    expects that, so push them there too.

    Returns (times, n_dropped).
    """
    out = []
    add = 0
    prev = None
    dropped = 0
    for t in times:
        if t is None:
            out.append(None)
            continue
        v = t + add
        if prev is not None:
            if v < prev and v + 86400 - prev <= MAX_HOP_S:
                add += 86400            # a real midnight crossing
                v = t + add
            elif v < prev or v - prev > MAX_HOP_S:
                # Not a time this train could have kept. Drop the stop and
                # leave prev where it was, so one bad value costs one stop
                # rather than every stop after it.
                out.append(None)
                dropped += 1
                continue
        out.append(v)
        prev = v
    first = next((t for t in out if t is not None), None)
    if first is not None and first < 3 * 3600:
        out = [None if t is None else t + 86400 for t in out]
    return out, dropped


# --------------------------------------------------------------------------
# reading
# --------------------------------------------------------------------------

Train = collections.namedtuple("Train", "line no day express stops")
# stops: [(kric_name, arr_secs, dep_secs), ...] in running order


def read_kric():
    """Both row shapes, yielding one Train per (train, day type)."""
    print("reading %s ..." % os.path.basename(KRIC))
    wb = openpyxl.load_workbook(KRIC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows)

    per_stop = collections.defaultdict(list)   # Korail shape, needs grouping
    trains = []
    n_rows = 0
    for r in rows:
        kline = r[2]
        L = LR.BY_KRIC.get(kline)
        if L is None:
            continue
        n_rows += 1
        express = norm(r[5]) in ("급행", "직통")
        days = daytypes(r[6])
        if not days:
            continue
        stop, arr, dep = r[7], r[8], r[9]
        if stop is None:
            continue
        s = str(stop)
        if not re.search(r"[+,]", s):
            # one row per stop -- collect and stitch together below
            t_arr, t_dep = parse_time(str(arr or "")), parse_time(str(dep or ""))
            if t_arr is None and t_dep is None:
                continue
            for day in days:
                per_stop[(L.id, day, norm(r[0]), norm(r[3]), norm(r[4]),
                          express)].append((t_dep if t_dep is not None else t_arr,
                                            base(s), t_arr, t_dep))
            continue

        # one row per train, everything packed
        names = parse_packed(stop)
        order = [t.split("-", 1)[0].strip()
                 for t in re.split(r"[+,/]", s) if "-" in t]
        arrs = parse_packed(arr)
        deps = parse_packed(dep)
        stops = []
        for code in order:
            nm = names.get(code)
            if not nm:
                continue
            a = parse_time(arrs.get(code))
            d = parse_time(deps.get(code))
            if a is None and d is None:
                a = d = None
            stops.append((base(nm), a, d))
        if sum(1 for _, a, d in stops if a is not None or d is not None) < 2:
            continue
        stops = running_order(stops)
        for day in days:
            trains.append(Train(L.id, norm(r[0]), day, express, stops))

    for key, rows_ in per_stop.items():
        line_id, day, no, _, _, express = key
        if spans_midnight([x[0] for x in rows_]):
            rows_ = [(lift(t), nm, lift(a), lift(d))
                     for (t, nm, a, d) in rows_]
        rows_.sort(key=lambda x: x[0])
        stops = [(nm, a, d) for _, nm, a, d in rows_]
        if len(stops) >= 2:
            trains.append(Train(line_id, no, day, express, stops))

    print("   %s rows in our lines -> %d train runs"
          % (format(n_rows, ","), len(trains)))
    return trains


def read_incheon2():
    """인천교통공사's wide CSVs: a column per station, a row per train."""
    files = [("incheon2_holiday_up.csv", "END"),
             ("incheon2_holiday_down.csv", "END"),
             ("incheon2_weekday_up.csv", "DAY"),
             ("incheon2_weekday_down.csv", "DAY")]
    trains = []
    for name, day in files:
        path = os.path.join(D, name)
        if not os.path.exists(path):
            print("   %s missing -- run fetch_schedules.py" % name)
            continue
        with io.open(path, encoding="cp949", errors="replace", newline="") as f:
            rows = list(csv.reader(f))
        hdr = [norm(h) for h in rows[0]]
        stations = hdr[3:]
        for i, row in enumerate(rows[1:]):
            if len(row) < 4:
                continue
            stops = []
            for j, cell in enumerate(row[3:]):
                t = parse_time(cell)
                if t is not None:
                    stops.append((base(stations[j]), t, t))
            if len(stops) >= 2:
                trains.append(Train("I2", "%s-%d" % (name[9:12], i), day,
                                    False, stops))
    print("   인천2호선: %d train runs" % len(trains))
    return trains


# --------------------------------------------------------------------------
# names
# --------------------------------------------------------------------------

def load_od_names():
    """OD station names, per line id and overall. The OD is the naming authority."""
    with io.open(OD, encoding="cp949", errors="replace", newline="") as f:
        rows = list(csv.DictReader(f))
    by_line = collections.defaultdict(set)
    allnames = set()
    for r in rows:
        for lab, st in ((norm(r["승차_호선"]), r["승차_역"]),
                        (norm(r["하차_호선"]), r["하차_역"])):
            n = base(st)
            allnames.add(n)
            lid = LR.OD_TO_ID.get(lab)
            if lid:
                by_line[lid].add(n)
    return by_line, allnames


class Resolver(object):
    """KRIC's truncated name -> the OD's name, or None.

    Order matters: try the name as given before stripping anything, and try
    prefix before the 신 strip, so that 신길온 finds 신길온천 rather than being
    read as a prefixed 길온.
    """

    def __init__(self, by_line, allnames):
        self.by_line = by_line
        self.allnames = allnames
        self.cache = {}
        self.log = collections.Counter()

    def _exact(self, cand, own):
        if cand in own:
            return cand
        if cand in self.allnames:
            return cand
        return None

    def _prefix(self, cand, own):
        for pool in (own, self.allnames):
            hits = [n for n in pool if n.startswith(cand)]
            if len(hits) == 1:
                return hits[0]
            if len(hits) > 1:
                exact = [n for n in hits if n == cand]
                if exact:
                    return exact[0]
        return None

    def resolve(self, name, line_id):
        key = (name, line_id)
        if key in self.cache:
            return self.cache[key]
        own = self.by_line.get(line_id, set())
        cands = [name]
        stripped = re.sub(r"역$", "", name)
        if stripped != name and len(stripped) >= 2:
            cands.append(stripped)
        for c in list(cands):
            nd = re.sub(r"^\d+", "", c)          # KRIC writes 1양원 for 양원
            if nd != c and len(nd) >= 2:
                cands.append(nd)

        for c in cands:
            r = self._exact(c, own)
            if r:
                self.log["exact"] += 1
                self.cache[key] = r
                return r
        for c in cands:
            r = self._prefix(c, own)
            if r:
                self.log["prefix"] += 1
                self.cache[key] = r
                return r
        for c in cands:                           # 신판교 -> 판교
            if c.startswith("신") and len(c) > 2:
                s = c[1:]
                r = self._exact(s, own) or self._prefix(s, own)
                if r:
                    self.log["sin-strip"] += 1
                    self.cache[key] = r
                    return r
        for c in cands:
            if c in ALIAS:
                self.log["alias"] += 1
                self.cache[key] = ALIAS[c]
                return ALIAS[c]
        self.log["unresolved"] += 1
        self.cache[key] = None
        return None


# --------------------------------------------------------------------------

def canonical_order(trains, line_id):
    """One running order for a line, from its longest run outward.

    Only used to label direction, so an imperfect merge costs a label, not a
    route.
    """
    seqs = [[nm for nm, _, _ in t.stops] for t in trains if t.line == line_id]
    seqs = [list(dict.fromkeys(s)) for s in seqs if len(s) > 1]
    if not seqs:
        return {}
    seqs.sort(key=len, reverse=True)
    order = list(seqs[0])
    known = set(order)
    for seq in seqs[1:]:
        for i, nm in enumerate(seq):
            if nm in known:
                continue
            before = next((seq[j] for j in range(i - 1, -1, -1)
                           if seq[j] in known), None)
            after = next((seq[j] for j in range(i + 1, len(seq))
                          if seq[j] in known), None)
            if before is not None:
                order.insert(order.index(before) + 1, nm)
            elif after is not None:
                order.insert(order.index(after), nm)
            else:
                order.append(nm)
            known.add(nm)
    return dict((nm, i) for i, nm in enumerate(order))


def main():
    if not os.path.exists(KRIC):
        raise SystemExit("missing %s -- run fetch_schedules.py first" % KRIC)

    trains = read_kric() + read_incheon2()

    print("resolving station names against the OD ...")
    by_line, allnames = load_od_names()
    res = Resolver(by_line, allnames)

    unresolved = collections.Counter()
    out_trains = []
    for t in trains:
        stops = []
        for nm, a, d in t.stops:
            full = res.resolve(nm, t.line)
            if full is None:
                unresolved[(t.line, nm)] += 1
                continue
            # A truncation can collide with the stop before it; keep the first.
            if stops and stops[-1][0] == full:
                continue
            stops.append((full, a, d))
        if len(stops) >= 2:
            out_trains.append(t._replace(stops=stops))

    print("   %s" % dict(res.log))
    if unresolved:
        # A name with no OD match is a station the network did not have on
        # 2023-12-31, or had no riders at: the 인천1호선 검단 extension opened
        # 2024-03, 경강선 성남 and 경의중앙 운천 have no OD rows at all. Dropping
        # the stop is the same clipping rule build_stations.py applies, so this
        # is a report rather than a warning -- but read it, because a genuine
        # parsing failure would surface here too.
        print("   no OD match, stop dropped -- %d names:" % len(unresolved))
        for (lid, nm), c in unresolved.most_common(40):
            print("      %-4s %-16s %d stops" % (lid, nm, c))

    # ---- direction, codes, output ----------------------------------------
    codes = {}
    next_code = [CODE_BASE]

    def code_of(line_id, name):
        key = (line_id, name)
        if key not in codes:
            codes[key] = "%05d" % next_code[0]
            next_code[0] += 1
        return codes[key]

    orders = dict((lid, canonical_order(out_trains, lid))
                  for lid in set(t.line for t in out_trains))

    rows = []
    seq_no = 0
    per_line = collections.Counter()
    dropped_times = 0
    bad_times = 0
    for t in out_trains:
        order = orders.get(t.line, {})
        pos = [order.get(nm) for nm, _, _ in t.stops]
        pos = [p for p in pos if p is not None]
        direction = "UP" if len(pos) < 2 or pos[-1] >= pos[0] else "DOWN"

        deps, nd = monotonic([d if d is not None else a for _, a, d in t.stops])
        arrs, na = monotonic([a if a is not None else d for _, a, d in t.stops])
        bad_times += max(nd, na)
        if all(x is None for x in deps):
            dropped_times += 1
            continue
        first = t.stops[0][0]
        last = t.stops[-1][0]
        for i, (nm, _, _) in enumerate(t.stops):
            a, d = arrs[i], deps[i]
            if a is None and d is None:
                continue
            seq_no += 1
            per_line[t.line] += 1
            rows.append({
                "고유번호": seq_no,
                "호선": t.line,
                "역사코드": code_of(t.line, nm),
                "역사명": nm,
                "주중주말": t.day,
                "방향": direction,
                "급행여부": "1" if t.express else "0",
                "열차코드": "%s-%s-%s" % (t.line, t.day, t.no),
                "열차도착시간": fmt_time(a) if a is not None else "",
                "열차출발시간": fmt_time(d) if d is not None else "",
                "출발역": first,
                "도착역": last,
            })

    with io.open(OUT, "w", encoding="cp949", errors="replace", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)

    print("\nwrote %s -- %s rows, %d stations"
          % (OUT, format(len(rows), ","), len(codes)))
    if dropped_times:
        print("   %d runs dropped for having no usable times" % dropped_times)
    if bad_times:
        print("   %d stop times dropped as implausible hops -- values the "
              "source got wrong, mostly 김포골드라인's 걸포북변 column" % bad_times)

    print("\n%-5s %-14s %8s %8s %7s   %s"
          % ("id", "line", "rows", "trains", "stns", "day types"))
    for lid in LR.EXTRA_IDS:
        L = LR.BY_ID[lid]
        tr = [t for t in out_trains if t.line == lid]
        days = collections.Counter(t.day for t in tr)
        nst = sum(1 for (l, _) in codes if l == lid)
        print("%-5s %-14s %8s %8d %7d   %s"
              % (lid, L.display, format(per_line[lid], ","), len(tr), nst,
                 dict(days)))

    # ---- did we reach every station the OD has on these lines? -----------
    print("\ncoverage against the OD:")
    total_missing = 0
    for lid in LR.EXTRA_IDS:
        want = by_line.get(lid, set())
        have = set(nm for (l, nm) in codes if l == lid)
        miss = sorted(want - have)
        total_missing += len(miss)
        flag = "" if not miss else "   MISSING: " + " ".join(miss)[:80]
        print("   %-4s %-14s %3d/%3d%s"
              % (lid, LR.BY_ID[lid].display, len(want & have), len(want), flag))
    print("\n%d OD stations on these lines have no stop in the timetable"
          % total_missing)


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
