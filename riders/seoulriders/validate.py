# -*- coding: utf-8 -*-
"""Check the network against the outside world, not against itself.

Every bug this project has found looked completely plausible on the map, so
the useful checks are the ones that compare what we built to a figure someone
else published. Three groups, in order of how much they would cost to get
wrong:

  --schedules   Did kric.py parse the timetables correctly? Compares each
                line's end-to-end journey time, service span and station count
                against published figures. A packed-cell parsed with the wrong
                delimiter, or times read as H:MM when they are H:MM:SS, shows
                up here as a journey time that is wrong by a factor.
  --geometry    Is every station where it should be? Walks each line's own
                stop order and measures the gaps. A station whose name was
                resolved to the wrong place lands far from its neighbours.
  --od          Is the IPF output sane? Compares fitted per-line daily volume
                against the OD's own totals, and prints the hourly profile of
                stations whose shape we can predict -- on a weekday an office
                district's *boardings* peak in the evening, not the morning.
  --coverage    What is actually in the OD? Measures how much of each line's
                traffic the source contains at all. This is the check that
                found the OD holds only trips touching the Seoul network.
  --congestion  Does the finished build agree with the operators' own published
                혼잡도? The only check here that tests the *output*: they
                measure riders-on-board as a share of 정원 per station, per
                direction, per half hour, for a typical 평일/토요일/일요일, and
                so do we. Lines 1-8 from 서울교통공사, 9호선 from its own operator
                and split 일반/급행 because on that line the two are not the
                same question. Shape is the part that matters -- a level offset
                shared by every line is a disagreement about 정원.

    python validate.py               # all of them
    python validate.py --schedules
    python validate.py --congestion --sample

Reference figures are the operators' own published journey times and service
spans, rounded to the minute. They are a sanity bound, not ground truth: a
10-15% difference is ordinary (we hold a 2026 timetable, they quote a typical
run), a 50% difference is a bug.
"""

import argparse
import collections
import csv
import io
import json
import math
import os
import re
import sys

import lines as LR

import daytype

HERE = os.path.dirname(os.path.abspath(__file__))
D = os.path.join(HERE, "data")


def _day():
    """The day build_od.py built, so this file checks the build that exists."""
    p = os.path.join(D, "od_hourly.npz")
    if not os.path.exists(p):
        return "nye", daytype.get("nye")
    import numpy as np
    with np.load(p, allow_pickle=True) as z:
        name = str(z["day"]) if "day" in z.files else "nye"
    return name, daytype.get(name)


DAY_NAME, DAY = _day()
SERVICE = DAY["service"]

# line id -> (end-to-end minutes, stations, first train, last train)
# Published by the operators; see the module docstring on how to read a miss.
REFERENCE = {
    "SB": (135, 63, "05:10", "24:30"),   # 청량리-인천, 수인분당선
    "GJ": (140, 57, "05:00", "24:30"),   # 문산-지평
    "AR": (66, 14, "05:20", "24:40"),    # 서울-인천공항2터미널, 일반열차
    "SN": (50, 16, "05:30", "24:30"),    # 신사-광교
    "UI": (23, 13, "05:30", "24:00"),    # 신설동-북한산우이
    "SL": (16, 11, "05:30", "24:30"),    # 샛강-관악산
    "I1": (57, 30, "05:30", "24:30"),    # 계양-국제업무지구 (2023 extent)
    "I2": (57, 27, "05:30", "24:30"),    # 검단오류-운연
    "GP": (32, 10, "05:30", "24:00"),    # 양촌-김포공항
    "GC": (80, 25, "05:10", "24:00"),    # 상봉-춘천
    "SH": (70, 21, "05:30", "24:00"),    # 일산-원시
    "GG": (50, 11, "05:30", "23:50"),    # 판교-여주
    "UJ": (21, 15, "05:00", "24:00"),    # 발곡-탑석
}

# A metro hop longer than this is worth a look; the commuter lines legitimately
# run further between stops, so they get their own bound.
# 신분당선 is a metro by service but was built for speed: the 청계산 tunnel
# between 청계산입구 and 판교 is 7.8 km, the longest gap on the network.
HOP_WARN_M = {"SB": 9000, "GJ": 12000, "GC": 14000, "SH": 9000, "GG": 16000,
              "AR": 20000, "1": 12000, "SN": 9000}
HOP_WARN_DEFAULT = 6000


def norm(s):
    return re.sub(r"\s+", "", (s or "").strip())


def base(s):
    """Station name without its trailing 부역명, e.g. 잠실(송파구청) -> 잠실."""
    return re.sub(r"\(.*?\)$", "", norm(s))


def read_cp949(name):
    p = os.path.join(D, name)
    if not os.path.exists(p):
        return []
    with io.open(p, encoding="cp949", errors="replace", newline="") as f:
        return list(csv.DictReader(f))


def hhmmss(s):
    s = (s or "").strip()
    if not s or s.count(":") != 2:
        return None
    h, m, sec = (int(x) for x in s.split(":"))
    return h * 3600 + m * 60 + sec


def hhmm(secs):
    return "%02d:%02d" % (secs // 3600, secs % 3600 // 60)


def metres(a, b):
    dy = (a[0] - b[0]) * 111320.0
    dx = (a[1] - b[1]) * 111320.0 * math.cos(math.radians(a[0]))
    return math.hypot(dx, dy)


def load_runs(service=SERVICE):
    """Every train run for a service day, as (line, [(code, dep_secs), ...])."""
    rows = read_cp949("timetable_raw.csv") + read_cp949("timetable_extra.csv")
    runs = collections.defaultdict(list)
    for r in rows:
        if service and norm(r["주중주말"]) != service:
            continue
        line = norm(r["호선"])
        if line not in LR.BY_ID:
            continue
        arr, dep = hhmmss(r["열차도착시간"]), hhmmss(r["열차출발시간"])
        if arr == 0:
            arr = None
        if dep == 0:
            dep = None
        t = dep if dep is not None else arr
        if t is None:
            continue
        key = (line, norm(r["방향"]), norm(r["열차코드"]))
        runs[key].append((t, norm(r["역사코드"])))
    out = collections.defaultdict(list)
    for (line, _, _), stops in runs.items():
        stops.sort()
        if len(stops) >= 2:
            out[line].append(stops)
    return out


def load_coords():
    with io.open(os.path.join(D, "stations.json"), encoding="utf-8") as f:
        net = json.load(f)
    coord, name = {}, {}
    for c in net["complexes"]:
        for p in c["platforms"]:
            coord[p["code"]] = (c["lat"], c["lon"])
            name[p["code"]] = c["name"]
    return net, coord, name


# --------------------------------------------------------------------------

def check_schedules(runs):
    print("=" * 78)
    print("SCHEDULES -- parsed timetable against the operators' published figures")
    print("=" * 78)
    print("%-4s %-14s %s" % ("", "", "end-to-end     stations      first        last"))
    print("%-4s %-14s %6s %6s  %4s %4s  %5s %5s  %5s %5s   %s"
          % ("id", "line", "ours", "pub", "ours", "pub",
             "ours", "pub", "ours", "pub", "trains"))
    problems = []
    for lid in LR.ALL_IDS:
        seqs = runs.get(lid)
        if not seqs:
            continue
        longest = max(seqs, key=len)
        e2e = (longest[-1][0] - longest[0][0]) / 60.0
        stations = len(set(c for s in seqs for _, c in s))
        first = min(s[0][0] for s in seqs)
        last = max(s[-1][0] for s in seqs)
        ref = REFERENCE.get(lid)
        if ref:
            r_e2e, r_st, r_first, r_last = ref
            flag = ""
            if abs(e2e - r_e2e) / max(r_e2e, 1) > 0.35:
                flag += " JOURNEY-TIME"
            if abs(stations - r_st) > 3:
                flag += " STATION-COUNT"
            if abs(first - hhmmss(r_first + ":00")) > 3600:
                flag += " FIRST-TRAIN"
            # An impossible last train is how a mishandled midnight shows up,
            # and it is invisible unless something asserts on it: 47:59 and
            # 234:36 both sat in this column for a while looking like data.
            if abs(last - hhmmss(r_last + ":00")) > 3600:
                flag += " LAST-TRAIN"
            if last > 28 * 3600:
                flag += " IMPOSSIBLE-LAST-TRAIN"
            if flag:
                problems.append((lid, flag.strip()))
            print("%-4s %-14s %6.0f %6d  %4d %4d  %5s %5s  %5s %5s   %5d%s"
                  % (lid, LR.DISPLAY[lid], e2e, r_e2e, stations, r_st,
                     hhmm(first), r_first, hhmm(last), r_last, len(seqs),
                     flag))
        else:
            flag = " IMPOSSIBLE-LAST-TRAIN" if last > 28 * 3600 else ""
            if flag:
                problems.append((lid, flag.strip()))
            print("%-4s %-14s %6.0f %6s  %4d %4s  %5s %5s  %5s %5s   %5d%s"
                  % (lid, LR.DISPLAY[lid], e2e, "-", stations, "-",
                     hhmm(first), "-", hhmm(last), "-", len(seqs), flag))
    print()
    if problems:
        print("   %d lines differ from the published figure by more than the "
              "tolerance:" % len(problems))
        for lid, flag in problems:
            print("      %-4s %s" % (lid, flag))
    else:
        print("   every line with a reference figure is within tolerance.")
    return problems


def check_geometry(runs, coord, name):
    print()
    print("=" * 78)
    print("GEOMETRY -- station spacing along each line's own stop order")
    print("=" * 78)
    print("A station resolved to the wrong place lands far from its neighbours.")
    print()
    print("%-4s %-14s %6s %7s %8s %9s   %s"
          % ("id", "line", "hops", "median", "longest", "over bound", "worst hop"))
    problems = []
    for lid in LR.ALL_IDS:
        seqs = runs.get(lid)
        if not seqs:
            continue
        longest = max(seqs, key=len)
        codes = [c for _, c in longest]
        hops = []
        for i in range(len(codes) - 1):
            a, b = coord.get(codes[i]), coord.get(codes[i + 1])
            if a and b and codes[i] != codes[i + 1]:
                hops.append((metres(a, b), codes[i], codes[i + 1]))
        if not hops:
            continue
        hops_m = sorted(h[0] for h in hops)
        med = hops_m[len(hops_m) // 2]
        worst = max(hops)
        bound = HOP_WARN_M.get(lid, HOP_WARN_DEFAULT)
        over = [h for h in hops if h[0] > bound]
        if over:
            problems.append((lid, over))
        print("%-4s %-14s %6d %6.0fm %7.0fm %9d   %s -> %s  %.1f km"
              % (lid, LR.DISPLAY[lid], len(hops), med, worst[0], len(over),
                 name.get(worst[1], "?"), name.get(worst[2], "?"),
                 worst[0] / 1000.0))
    print()
    if problems:
        print("   hops over the per-line bound (long is normal on commuter "
              "lines; a metro line's is not):")
        for lid, over in problems:
            for d, a, b in sorted(over, reverse=True)[:4]:
                print("      %-4s %-12s -> %-12s %6.1f km"
                      % (lid, name.get(a, "?"), name.get(b, "?"), d / 1000.0))
    else:
        print("   no station sits implausibly far from its neighbours.")
    return problems


def check_od(net):
    print()
    print("=" * 78)
    print("OD -- fitted hourly volumes against the source totals")
    print("=" * 78)
    import numpy as np

    p = os.path.join(D, "od_hourly.npz")
    if not os.path.exists(p):
        print("   no od_hourly.npz -- run build_od.py")
        return []
    z = np.load(p, allow_pickle=True)
    X, pairs = z["x"], z["pairs"]
    complexes = net["complexes"]

    # which line each complex belongs to, for a per-line rollup
    cx_lines = [sorted(set(pl["line"] for pl in c["platforms"]))
                for c in complexes]

    board = np.zeros(len(complexes))
    for k, (o, _d) in enumerate(pairs):
        board[int(o)] += X[k].sum()

    print("fitted daily boardings by line (a station on two lines is counted")
    print("once per line, so the total exceeds the trip count):")
    print()
    per = collections.Counter()
    for i, ls in enumerate(cx_lines):
        for l in ls:
            per[l] += board[i]
    print("   %-4s %-14s %12s" % ("id", "line", "boardings"))
    for lid in LR.ALL_IDS:
        if per.get(lid):
            print("   %-4s %-14s %12s"
                  % (lid, LR.DISPLAY[lid], format(int(per[lid]), ",")))

    # Shape check: a handful of stations whose profile we can predict.
    hours = list(z["hours"])
    print()
    print("hourly shape for stations whose profile we can predict:")
    idx = dict((c["name"], i) for i, c in enumerate(complexes))
    # What each of these should look like depends on the day, so say which is
    # expected rather than leaving a caption that was written for one build and
    # quietly stops being true in another.
    WATCH = {
        "nye": [("인천공항1터미널", "an airport: flat, no commuter peak"),
                ("잠실", "NYE crowd: a late spike"),
                ("홍대입구", "nightlife: builds through the evening"),
                ("여의도", "office: quiet on a Sunday")],
        "weekday": [("인천공항1터미널", "an airport: flat-ish, no evening rush"),
                    ("잠실", "residential and commercial: peaks at both ends"),
                    ("홍대입구", "nightlife: still builds through the evening"),
                    ("여의도", "office: boardings peak on the way *home*")],
    }
    watch = WATCH.get(DAY_NAME, WATCH["weekday"])
    prof = np.zeros((len(complexes), X.shape[1]))
    for k, (o, _d) in enumerate(pairs):
        prof[int(o)] += X[k]
    for nm, why in watch:
        i = idx.get(nm)
        if i is None:
            print("   %-14s (not in the network)" % nm)
            continue
        v = prof[i]
        if v.sum() <= 0:
            print("   %-14s (no boardings)" % nm)
            continue
        peak = v.max()
        bars = "".join("%d" % min(9, round(9 * x / peak)) for x in v)
        print("   %-14s %s   %s" % (nm, bars, why))
    print("   %-14s %s" % ("", "".join(str(h % 10) for h in hours)))
    print("   (digits are 0-9 relative to that station's own peak hour)")
    return []


def check_coverage():
    """How much of a line's own traffic does the OD contain at all?

    A line whose riders mostly travel within it should show a high share of
    within-line trips: 2호선 is 55%. A line showing ~0% is not quiet, it is
    absent -- the source only holds trips that touch the Seoul network, so the
    outer operators appear as feeders and their internal traffic is missing.
    This is a property of the OD, not of anything we do to it, and it is the
    single most important caveat on the map.
    """
    print()
    print("=" * 78)
    print("COVERAGE -- how much of each line's traffic is in the OD at all")
    print("=" * 78)
    rows = read_cp949("od_2023-12-31.csv")
    if not rows:
        print("   no OD file")
        return []
    board = collections.Counter()
    internal = collections.Counter()
    total = 0
    for r in rows:
        n = int(r["총_승객수"])
        o, d = norm(r["승차_호선"]), norm(r["하차_호선"])
        total += n
        board[o] += n
        if o == d:
            internal[o] += n

    print("%-16s %11s %11s %8s   %s"
          % ("OD 호선 label", "boardings", "within-line", "share", ""))
    problems = []
    for lab in sorted(board, key=lambda l: -board[l]):
        share = 100.0 * internal[lab] / max(board[lab], 1)
        note = ""
        if board[lab] > 1500 and share < 5.0:
            note = "  <- essentially only its trips to and from Seoul"
            problems.append((lab, "OD-COVERAGE"))
        print("%-16s %11s %11s %7.1f%%%s"
              % (lab, format(board[lab], ","), format(internal[lab], ","),
                 share, note))
    print()
    print("   Lines flagged above are not quiet -- their internal traffic is")
    print("   not in the source. 인천1호선's real daily ridership is an order")
    print("   of magnitude above what the OD holds for it.")
    return problems


# --------------------------------------------------------------------------
# against 혼잡도
# --------------------------------------------------------------------------

CONGESTION = os.path.join(D, "congestion_raw.csv")
CONGESTION9 = os.path.join(D, "congestion_line9.xlsx")

# 9호선's file is keyed by station *name* -- it carries no 역번호 -- and its
# 평일/휴일 split is coarser than the three-way one for lines 1-8. A Saturday
# has to borrow 휴일, which is the file's own limit, not ours.
LINE9_DAY = {"평일": "평일", "토요일": "휴일", "일요일": "휴일"}

# 서울교통공사 numbers its stations in the 하행 direction, so the direction of a
# segment is the sign of the station-number step. 2호선's loop is labelled
# 내선/외선 instead, and 외선순환 is the one that runs 시청 -> 을지로입구, i.e.
# increasing. Anything else on the loop's branches keeps 상선/하선.
DIR_UP, DIR_DOWN = ("상선", "내선"), ("하선", "외선")


def load_congestion(day_label):
    """(line, station number, direction) -> [percent per hour, 05:00-25:00]."""
    if not os.path.exists(CONGESTION):
        return None
    with io.open(CONGESTION, encoding="cp949", errors="replace", newline="") as f:
        rows = list(csv.DictReader(f))
    if not rows:
        return None
    cols = [c for c in rows[0]
            if c not in ("구분", "호선", "역번호", "역명", "상하구분")]
    # '5시30분' .. '00시30분', on the half hour. Fold to the hours the build
    # uses; the two halves of an hour are equally weighted, which is what an
    # "average over the trains in the hour" means when headways are even.
    slot = {}
    for c in cols:
        h = int(re.match(r"(\d+)시", c).group(1))
        slot.setdefault(h if h >= 5 else h + 24, []).append(c)

    out = {}
    for r in rows:
        if norm(r["구분"]) != day_label:
            continue
        line = norm(r["호선"]).replace("호선", "")
        num = norm(r["역번호"]).lstrip("0")
        d = norm(r["상하구분"])
        per = {}
        for h, cs in slot.items():
            v = [float(r[c]) for c in cs if (r[c] or "").strip()]
            if v:
                per[h] = sum(v) / len(v)
        out[(line, num, d)] = per
    return out


def load_congestion9(day_label):
    """9호선: (station name, direction, express) -> [percent per hour].

    A different shape from the lines 1-8 CSV and it needs its own reader: eight
    sheets named 상선일반(평일) and so on, stations down the side, half-hour
    columns across the top, and no station numbers at all. Keyed by name, which
    is safe here because 9호선 shares none of its station names with itself.
    """
    if not os.path.exists(CONGESTION9):
        return None
    try:
        import openpyxl
    except ImportError:
        print("congestion: 9호선 needs openpyxl (pip install openpyxl) -- skipped")
        return None

    want = LINE9_DAY.get(day_label, "평일")
    wb = openpyxl.load_workbook(CONGESTION9, read_only=True, data_only=True)
    out = {}
    for sheet in wb.sheetnames:
        m = re.match(r"(상선|하선)(일반|급행)\((평일|휴일)\)", norm(sheet))
        if not m or m.group(3) != want:
            continue
        direction, kind = m.group(1), m.group(2)
        express = kind == "급행"
        cols = None
        for row in wb[sheet].iter_rows(values_only=True):
            if row is None or all(c is None for c in row):
                continue
            head = norm(str(row[0] or ""))
            if head == "구분":
                # '05:30~05:59' .. fold the two halves of an hour together
                cols = []
                for c in row[1:]:
                    mm = re.match(r"(\d+):", str(c or ""))
                    cols.append(int(mm.group(1)) if mm else None)
                continue
            if cols is None or not head:
                continue
            per = collections.defaultdict(list)
            for c, hh in zip(row[1:], cols):
                if hh is None or c is None:
                    continue
                try:
                    per[hh if hh >= 5 else hh + 24].append(float(c))
                except (TypeError, ValueError):
                    continue
            out[(head, direction, express)] = dict(
                (h, sum(v) / len(v)) for h, v in per.items() if v)
    wb.close()
    return out or None


def check_congestion(stats_path):
    """Compare the build's train loads against 서울교통공사's published 혼잡도.

    This is the only check here that tests the *output* rather than an input.
    혼잡도 is riders on board as a percentage of 정원, averaged over the trains
    passing a station in a window -- exactly what the map draws, published for
    a typical 평일/토요일/일요일 over lines 1-8 within 서울교통공사's own
    boundary. Our equivalent is the segment's riders-per-hour over its
    trains-per-hour, against the line's capacity.

    Read the *shape* first and the level second. The shape is the thing a
    day-type change is supposed to fix, and it is measured on both sides. The
    level carries our capacity assumption (160 a car) and theirs, so a
    consistent 10-20% offset across every line means the two definitions of
    정원 differ, not that the routing is wrong.
    """
    pub = load_congestion(DAY["congestion"])
    if pub is None:
        print("congestion: no data/congestion_raw.csv -- run fetch_ridership.py")
        return []
    pub9 = load_congestion9(DAY["congestion"])
    if not os.path.exists(stats_path):
        print("congestion: no %s -- run build.py" % os.path.basename(stats_path))
        return []
    with io.open(stats_path, encoding="utf-8") as f:
        stats = json.load(f)
    if "n" not in (stats["segments"][0] if stats["segments"] else {}):
        print("congestion: %s predates the per-segment train counts -- "
              "re-run build.py" % os.path.basename(stats_path))
        return []

    print("comparing train loads against 혼잡도 (%s, %d rows%s) ..."
          % (DAY["congestion"], len(pub),
             ", + %d for 9호선" % len(pub9) if pub9 else
             ", 9호선 absent -- run fetch_ridership.py"))
    if pub9 and not any("hx" in s for s in stats["segments"]):
        print("   NOTE: %s predates the express split, so 9호선's 급행 cannot be"
              % os.path.basename(stats_path))
        print("   separated from its 일반. Re-run build.py.")
        pub9 = None
    if stats["build"].get("sample", 1) > 1:
        print("   NOTE: --sample %d build. Ignore every level below. A sampled"
              % stats["build"]["sample"])
        print("   run keeps only the segments near the origins it kept, and "
              "those\n   carry their riders in full, so the ratio column is "
              "neither 1x nor\n   1/%d. The shape correlations are still worth "
              "reading." % stats["build"]["sample"])

    hours = stats["hours"]
    cap = dict((k, v.get("capacity") or 0)
               for k, v in stats["line_meta"].items())
    seoul_lines = set(str(i) for i in range(1, 9))

    def collect(flip):
        """(line, hour) -> [(ours %, theirs %, riders), ...]."""
        out = collections.defaultdict(list)
        matched = unmatched = 0
        for s in stats["segments"]:
            line = s["line"]
            if line not in seoul_lines or not cap.get(line):
                continue
            a, b = norm(s.get("ca", "")), norm(s.get("cb", ""))
            if not a.isdigit() or not b.isdigit():
                continue
            down = (int(b) - int(a)) > 0
            want = DIR_UP if down == bool(flip) else DIR_DOWN
            per = None
            for d in want:
                per = pub.get((line, a.lstrip("0"), d))
                if per is not None:
                    break
            if per is None:
                unmatched += 1
                continue
            matched += 1
            for i, h in enumerate(hours):
                if h not in per or not s["n"][i]:
                    continue
                out[(line, h)].append(
                    (100.0 * (s["h"][i] / s["n"][i]) / cap[line],
                     per[h], s["h"][i]))

        # 9호선, from its own file, and split the way its operator splits it.
        # Comparing our blended average against either sheet would be
        # meaningless: on this line the 급행 is much fuller than the 일반 it
        # overtakes, which is the whole reason both are published.
        if pub9:
            c9 = cap.get("9") or 0
            for s in stats["segments"]:
                if s["line"] != "9" or not c9:
                    continue
                a, b = norm(s.get("ca", "")), norm(s.get("cb", ""))
                if not a.isdigit() or not b.isdigit():
                    continue
                down = (int(b) - int(a)) > 0
                dirs = DIR_UP if down == bool(flip) else DIR_DOWN
                name = base(s.get("a", ""))
                nx = s.get("nx") or [0] * len(hours)
                hx = s.get("hx") or [0.0] * len(hours)
                for express in (True, False):
                    per9 = None
                    for d in dirs:
                        per9 = pub9.get((name, d, express))
                        if per9 is not None:
                            break
                    if per9 is None:
                        unmatched += 1
                        continue
                    matched += 1
                    label = "9급행" if express else "9일반"
                    for i, h in enumerate(hours):
                        # express subset, or what is left after taking it out
                        n = nx[i] if express else s["n"][i] - nx[i]
                        r = hx[i] if express else s["h"][i] - hx[i]
                        if h not in per9 or n <= 0 or r <= 0:
                            continue
                        out[(label, h)].append(
                            (100.0 * (r / n) / c9, per9[h], r))
        return out, matched, unmatched

    rows, matched, unmatched = collect(False)
    print("   %d segments matched a published station-direction, %d not"
          % (matched, unmatched))
    print("   (the misses are the Korail through-running sections, which are "
          "outside\n    서울교통공사's boundary and so outside the 혼잡도 file)")
    if not rows:
        return []

    # 상선/하선 is assigned from the station-number step, on the rule that Seoul
    # numbers in the 하행 direction. That rule is an assumption, and a silent
    # swap would leave every correlation positive -- both directions are busy
    # at both rushes -- just worse. So try it both ways and say which won.
    def cellwise_r(rs):
        o, t = [], []
        for cells in rs.values():
            for ours, theirs, w in cells:
                o.append(ours)
                t.append(theirs)
        return _pearson(o, t)

    r_ok = cellwise_r(rows)
    r_flip = cellwise_r(collect(True)[0])
    verdict = "as assigned" if r_ok >= r_flip else "SWAPPED -- see DIR_UP/DIR_DOWN"
    print("   direction check: 상선/하선 as assigned r=%.3f, flipped r=%.3f  -> %s"
          % (r_ok, r_flip, verdict))

    def agg(keys):
        """Rider-weighted mean congestion, ours and theirs, over some cells.

        Weighted rather than plain: an empty branch segment at 03% and 시청 at
        150% are one number each in the file, but they are not one question
        each. Weighting by the riders we route makes the summary say what the
        map looks like.
        """
        o = t = w = 0.0
        for k in keys:
            for ours, theirs, riders in rows.get(k, ()):
                o += ours * riders
                t += theirs * riders
                w += riders
        return (o / w, t / w, w) if w > 0 else (0.0, 0.0, 0.0)

    all_hours = sorted(set(h for _, h in rows))
    lines_seen = sorted(set(l for l, _ in rows),
                        key=lambda l: (1, l) if l.startswith("9")
                        else (0, LR.order_key(l)))

    print("\n   network profile, weighted by riders (lines 1-8, 서울교통공사)")
    print("   %5s %8s %10s %7s" % ("hour", "ours", "published", "ratio"))
    for h in all_hours:
        o, t, w = agg([(l, h) for l in lines_seen])
        if w <= 0:
            continue
        bar = "#" * int(round(t / 2.0))
        print("   %02d:00 %7.1f%% %9.1f%% %6.2fx  %s"
              % (h % 24, o, t, (o / t) if t else float("nan"), bar))

    print("\n   per line, over the whole day")
    print("   %-4s %8s %10s %7s %8s" % ("line", "ours", "published", "ratio", "corr"))
    problems = []
    for l in lines_seen:
        o, t, w = agg([(l, h) for h in all_hours])
        if w <= 0 or t <= 0:
            continue
        # Correlation of the hourly shapes, which is the part that is measured
        # on both sides and independent of any capacity assumption.
        ov, tv = [], []
        for h in all_hours:
            oh, th, wh = agg([(l, h)])
            if wh > 0:
                ov.append(oh)
                tv.append(th)
        corr = _pearson(ov, tv)
        print("   %-4s %7.1f%% %9.1f%% %6.2fx %7.2f"
              % (l, o, t, o / t, corr))
        if corr < 0.85:
            problems.append((l, "CONGESTION-SHAPE %.2f" % corr))

    if problems:
        print("\n   A shape correlation below 0.85 means the build's rush hour")
        print("   is not where 서울교통공사 measures it. That is a real problem;")
        print("   a level offset shared by every line is a capacity question.")
    return problems


def _pearson(a, b):
    n = len(a)
    if n < 3:
        return float("nan")
    ma, mb = sum(a) / n, sum(b) / n
    va = sum((x - ma) ** 2 for x in a) ** 0.5
    vb = sum((x - mb) ** 2 for x in b) ** 0.5
    if va == 0 or vb == 0:
        return float("nan")
    return sum((x - ma) * (y - mb) for x, y in zip(a, b)) / (va * vb)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--coverage", action="store_true")
    ap.add_argument("--schedules", action="store_true")
    ap.add_argument("--geometry", action="store_true")
    ap.add_argument("--od", action="store_true")
    ap.add_argument("--congestion", action="store_true")
    ap.add_argument("--sample", action="store_true",
                    help="check stats.sample.json instead of stats.json")
    args = ap.parse_args()
    everything = not (args.schedules or args.geometry or args.od or
                      args.coverage or args.congestion)

    bad = []
    print("loading timetables and network ...")
    runs = load_runs()
    net, coord, name = load_coords()

    # The 김포골드라인 weekday bug lived for a day because every check ran on
    # END service only, which is the one we draw. A bad row on any service day
    # is a bad parse, so sweep all of them.
    print("   checking every service day for impossible times ...")
    for svc in ("DAY", "SAT", "END"):
        bad_svc = []
        for lid, seqs in load_runs(svc).items():
            worst = max((s[-1][0] for s in seqs), default=0)
            if worst > 28 * 3600:
                bad_svc.append((lid, worst))
        if bad_svc:
            for lid, worst in bad_svc:
                print("      %s %-4s last stop at %.1fh   IMPOSSIBLE"
                      % (svc, lid, worst / 3600.0))
            bad += [(lid, "%s IMPOSSIBLE-TIME" % svc) for lid, _ in bad_svc]
        else:
            print("      %s: clean" % svc)

    print("   %d lines with END-service runs, %d platforms placed\n"
          % (len(runs), len(coord)))

    if everything or args.schedules:
        bad += check_schedules(runs)
    if everything or args.geometry:
        bad += check_geometry(runs, coord, name)
    if everything or args.od:
        check_od(net)
    if everything or args.coverage:
        bad += check_coverage()
    if everything or args.congestion:
        print()
        bad += check_congestion(os.path.join(
            D, "stats.sample.json" if args.sample else "stats.json"))

    print()
    print("=" * 78)
    print("%d thing(s) flagged for a look." % len(bad))


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
