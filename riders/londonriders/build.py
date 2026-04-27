"""
Build trains.json: synthesized train timelines + per-train rider counts for the
TfL rail network on a typical Tue/Wed/Thu (NUMBAT 2024 TWT day type).

Pipeline:
1. Load geometry.json (stations, branches, line drawings).
2. Load Link_Frequencies (trains/15min per (line, dir, link)) from outputs.xlsx.
3. For each branch, spawn N trains per 15-min band using the first link's frequency.
4. Each train traverses the branch with timings derived from distance / mode-speed.
5. Walk per-line OD CSVs; for each (line, origin, dest, qhr, vol), find a branch
   and a train pass that covers it, attribute boardings/alightings.
6. Write trains.json compatible with the nycriders viewer.
"""
import argparse
import csv
import json
import math
import os
import random
import time as time_module
from collections import defaultdict

import openpyxl

GEOMETRY_JSON = "geometry.json"
OUT_JSON = "trains.json"
OUTPUTS_XLSX = "data/numbat_twt_outputs.xlsx"
OD_PER_LINE_DIR = "data/od_per_line"

# qhr 1 = 0500-0515. NUMBAT day starts at 0500 and wraps past midnight to 0445.
QHR_START_HOUR = 5
QHR_LEN_S = 15 * 60  # 900s
DAY_START_S = QHR_START_HOUR * 3600

# Effective inter-station speeds (m/s) including dwell allowance.
# Tuned so a typical 1 km tube hop takes ~110 s door-to-door.
MODE_SPEED = {
    "u": 9.0,    # Underground
    "d": 8.5,    # DLR
    "o": 12.0,   # Overground (single-mode)
    "or": 12.0,  # Overground (multi-mode)
    "er": 17.0,  # Elizabeth
    "rr": 13.0,  # National Rail (rare in TfL OD)
    "t": 6.0,    # Trams
}
DEFAULT_SPEED = 10.0
DWELL_S = 25  # extra dwell time at each intermediate stop


def load_geometry():
    print(f"Loading {GEOMETRY_JSON}...")
    with open(GEOMETRY_JSON) as f:
        geo = json.load(f)
    print(f"  {len(geo['stations'])} stations, {len(geo['branches'])} branches")
    return geo


def load_link_frequencies():
    """Read Link_Frequencies sheet. Returns list of dicts:
        { line, dir, order, from_mnlc, to_mnlc, qhr_freq: [96 floats] }
    """
    print(f"Loading {OUTPUTS_XLSX} Link_Frequencies...")
    wb = openpyxl.load_workbook(OUTPUTS_XLSX, read_only=True, data_only=True)
    ws = wb["Link_Frequencies"]

    rows = ws.iter_rows(values_only=True)
    next(rows); next(rows)  # skip 2 title rows
    header = list(next(rows))
    idx = {h: i for i, h in enumerate(header) if h is not None}

    # 96 quarter-hour columns
    qhr_keys = [k for k in header if isinstance(k, str) and len(k) == 9 and k[4] == "-"]
    if len(qhr_keys) != 96:
        # fallback: take last 96 string keys
        qhr_keys = [k for k in header if isinstance(k, str) and "-" in k][-96:]
    qhr_idx = [idx[k] for k in qhr_keys]
    print(f"  {len(qhr_idx)} qhr columns ({qhr_keys[0]} ... {qhr_keys[-1]})")

    out = []
    for row in rows:
        if row[idx["Link"]] is None:
            continue
        try:
            from_mnlc = int(row[idx["From NLC"]])
            to_mnlc = int(row[idx["To NLC"]])
        except (TypeError, ValueError):
            continue
        freq = [float(row[i] or 0) for i in qhr_idx]
        out.append({
            "line": row[idx["Line"]],
            "dir": row[idx["Dir"]],
            "order": int(row[idx["Order"]] or 0),
            "from_mnlc": from_mnlc,
            "to_mnlc": to_mnlc,
            "qhr_freq": freq,
        })
    print(f"  {len(out)} link frequency rows")
    return out


def build_branch_index(branches):
    """Build (line, dir) -> list of branches, sorted by length descending."""
    by_ld = defaultdict(list)
    for br in branches:
        by_ld[(br["line"], br["dir"])].append(br)
    for k in by_ld:
        by_ld[k].sort(key=lambda b: -len(b["mnlcs"]))
    return by_ld


def link_freq_lookup(link_freqs):
    """(line_desc, dir, from_mnlc, to_mnlc) -> qhr_freq array.
    Keys are line *descriptions* (Bakerloo, Central, ...) since that's what
    Link_Frequencies uses, plus 2-letter direction codes.
    """
    out = {}
    for lf in link_freqs:
        out[(lf["line"], lf["dir"], lf["from_mnlc"], lf["to_mnlc"])] = lf["qhr_freq"]
    return out


# Line code -> Link_Frequencies line description. Built from Lines sheet
# but with manual fixes for sheets that name lines slightly differently.
LINE_CODE_TO_DESC_OVERRIDES = {
    "HAM": "H&C and Circle",
    "EZL": "Elizabeth Line",
    "ELL": "LO Windrush",
    "GOB": "LO Suffragette",
    "NLL": "LO Mildmay",
    "URL": "LO Liberty",
    "WEL": "LO Lioness",
    "WAG": "LO Weaver",
}


def load_line_descriptions():
    """Read Lines sheet, return dict line_code -> description."""
    wb = openpyxl.load_workbook("data/numbat_definitions.xlsx",
                                read_only=True, data_only=True)
    ws = wb["Lines"]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    out = {}
    for row in rows:
        code = row[1]
        desc = row[2]
        if code and desc:
            out[code] = desc
    out.update(LINE_CODE_TO_DESC_OVERRIDES)
    return out


# Branches use 1-letter dir codes (N, S, E, W, I, O, U, D).
# Link_Frequencies uses 2-letter (NB, SB, EB, WB, IB, OB, UP, DN) for most lines
# and IR/OR for circle-style routes (Inner Rail / Outer Rail).
DIR_2LETTER_VARIANTS = {
    "N": ["NB"], "S": ["SB"], "E": ["EB"], "W": ["WB"],
    "U": ["UP"], "D": ["DN"],
    "I": ["IB", "IR"], "O": ["OB", "OR"],
}


def synthesize_trains(branches, link_freqs, stations, line_colors, line_descs):
    """For each branch, spawn N trains per 15-min band from first-link freq.
    Returns list of train dicts: {line, color, branch_id, dir, mnlcs, times, boardings, alightings}
    """
    print("Synthesizing trains from link frequencies...")
    freq_by_link = link_freq_lookup(link_freqs)

    # Count how many branches share each (line_desc, dir, from, to) link.
    # Link_Frequencies reports the *combined* trains/15min on shared track,
    # so splitting by share-count avoids spawning N copies on a junction.
    link_share = defaultdict(int)
    for br in branches:
        if len(br["mnlcs"]) < 2:
            continue
        desc = line_descs.get(br["line"], br["line"])
        for d2 in DIR_2LETTER_VARIANTS.get(br["dir"], [br["dir"]]):
            for k in range(len(br["mnlcs"]) - 1):
                key = (desc, d2, br["mnlcs"][k], br["mnlcs"][k + 1])
                if key in freq_by_link:
                    link_share[key] += 1
                    break

    trains = []
    skipped_no_freq = 0
    for br in branches:
        if len(br["mnlcs"]) < 2:
            continue
        line = br["line"]
        dir1 = br["dir"]
        desc = line_descs.get(line, line)
        first_from = br["mnlcs"][0]
        first_to = br["mnlcs"][1]

        # Walk the branch link-by-link, take the first one with freq data,
        # and divide by the number of branches that picked the same link
        # (so a 2-way junction with 24 trains/15min gives each branch 12).
        freq = None
        used_key = None
        for k in range(min(len(br["mnlcs"]) - 1, 8)):
            fa, fb = br["mnlcs"][k], br["mnlcs"][k + 1]
            for d2 in DIR_2LETTER_VARIANTS.get(dir1, [dir1]):
                key = (desc, d2, fa, fb)
                cand = freq_by_link.get(key)
                if cand is not None:
                    freq = cand
                    used_key = key
                    break
            if freq is not None:
                break
        if freq is None:
            skipped_no_freq += 1
            continue
        share = max(1, link_share.get(used_key, 1))
        if share > 1:
            freq = [v / share for v in freq]

        # Compute travel time per link from distance and mode speed.
        speed = MODE_SPEED.get(br["mode"], DEFAULT_SPEED)
        link_times = []
        for d in br["distances_m"]:
            t = max(20, d / speed) + DWELL_S
            link_times.append(t)
        # Cumulative arrival time at each stop relative to spawn at stop 0.
        arr_offsets = [0]
        cum = 0
        for t in link_times:
            cum += t
            arr_offsets.append(cum)

        color = line_colors.get(line, "#808183")

        # For each 15-min band, spawn N trains evenly across the band.
        for qhr in range(96):
            n_raw = freq[qhr]
            if n_raw <= 0:
                continue
            # Round freq to int trains; for fractional, use stochastic rounding
            # so e.g. 0.5 trains spawns ~half the time across many qhrs.
            n_int = int(n_raw)
            if random.random() < (n_raw - n_int):
                n_int += 1
            if n_int < 1:
                continue
            band_start_s = DAY_START_S + qhr * QHR_LEN_S
            spacing = QHR_LEN_S / n_int
            for i in range(n_int):
                spawn_s = band_start_s + (i + random.random() * 0.5) * spacing
                trains.append({
                    "line": line,
                    "color": color,
                    "branch_id": br["branch_id"],
                    "dir": dir1,
                    "qhr_spawn": qhr,
                    "mnlcs": br["mnlcs"],
                    "stop_times": [round(spawn_s + off) for off in arr_offsets],
                    "boardings": defaultdict(float),
                    "alightings": defaultdict(float),
                })

    print(f"  {len(trains)} train runs synthesized "
          f"({skipped_no_freq} branches skipped — no Link_Frequencies match)")
    return trains


def index_trains_by_line_branch(trains):
    """(line, branch_id) -> list of train indices, sorted by spawn time."""
    by_lb = defaultdict(list)
    for i, t in enumerate(trains):
        by_lb[(t["line"], t["branch_id"])].append(i)
    for k in by_lb:
        by_lb[k].sort(key=lambda i: trains[i]["stop_times"][0])
    return by_lb


def build_branch_path_index(branches):
    """For each (line, dir, branch), give mnlc -> position in branch.
    Returns: branches_by_line[line] -> list of (branch, dir, mnlc_to_pos)."""
    out = defaultdict(list)
    for br in branches:
        pos = {mnlc: i for i, mnlc in enumerate(br["mnlcs"])}
        out[br["line"]].append({
            "branch_id": br["branch_id"],
            "dir": br["dir"],
            "pos": pos,
            "mnlcs": br["mnlcs"],
        })
    return out


# Map per-line CSV file basenames -> list of candidate NUMBAT line codes.
# OD rows are searched against branches on any of the candidate lines, so files
# that aggregate multiple lines (ssl = sub-surface = DIS+HAM+MET) can match.
PER_LINE_TO_CODES = {
    "bak": ["BAK"],
    "bakwdc": ["BAK", "WEL"],  # Bakerloo + Watford DC (shared with Lioness)
    "cen": ["CEN"],
    "dis": ["DIS"],
    "dlr": ["DLR"],
    "ezl": ["EZL"],
    "ham": ["HAM"],
    "jub": ["JUB"],
    "loa": ["WAG"],      # LO Anglia (West Anglia routes -> Weaver)
    "loe": ["ELL"],      # LO East London -> Windrush
    "log": ["GOB"],      # LO Gospel Oak - Barking -> Suffragette
    "lon": ["NLL"],      # LO North London -> Mildmay
    "lor": ["URL"],      # LO Romford -> Liberty
    "met": ["MET"],
    "nor": ["NOR"],
    "pic": ["PIC"],
    "ssl": ["DIS", "HAM", "MET"],  # Sub-surface lines combined
    "ssp": ["PIC", "DIS", "HAM", "MET"],  # Piccadilly route-choice incl SSL share
    "vic": ["VIC"],
    "wat": ["WAC"],
}


def assign_riders(trains, branches, stations, sample_rate=1.0):
    """For each per-line OD CSV row, find a branch covering origin->dest
    and distribute riders across trains in that qhr."""
    print(f"Assigning riders (sample_rate={sample_rate})...")
    branches_by_line = build_branch_path_index(branches)
    trains_by_lb = index_trains_by_line_branch(trains)

    total_assigned = 0
    total_unassigned = 0
    total_riders_assigned = 0.0
    total_riders_unassigned = 0.0
    rng = random.Random(42)

    for fname in sorted(os.listdir(OD_PER_LINE_DIR)):
        if not fname.endswith(".csv"):
            continue
        base = fname[:-4]
        line_codes = PER_LINE_TO_CODES.get(base)
        if not line_codes:
            print(f"  skip {fname} (no line mapping)")
            continue

        path = os.path.join(OD_PER_LINE_DIR, fname)
        # Candidate branches across all lines this CSV represents.
        candidates = []
        for lc in line_codes:
            for br in branches_by_line.get(lc, []):
                candidates.append((lc, br))
        if not candidates:
            print(f"  skip {fname}: no branches for lines {line_codes}")
            continue

        n_assigned = 0
        n_unassigned = 0
        riders_assigned = 0.0
        riders_unassigned = 0.0

        with open(path) as f:
            reader = csv.DictReader(f)
            for row in reader:
                if rng.random() > sample_rate:
                    continue
                try:
                    o = int(row["mode_mnlc_o"])
                    d = int(row["mode_mnlc_d"])
                    # qhr in per-line OD CSVs is 1-indexed from midnight
                    # (qhr=1 -> 0000-0015), not 5am-based like Link_Frequencies.
                    qhr = int(row["qhr"]) - 1
                    vol = float(row["vol"])
                except (KeyError, ValueError):
                    continue
                if vol <= 0 or o == d:
                    continue
                # qhr 0..~104 (midnight-based, includes early-morning of next day)
                if not (0 <= qhr < 104):
                    continue

                # Find a (line, branch) where both o and d are present, d after o.
                best_line = None
                best_branch = None
                best_o_pos = best_d_pos = None
                for lc, br in candidates:
                    op = br["pos"].get(o)
                    dp = br["pos"].get(d)
                    if op is None or dp is None:
                        continue
                    if dp <= op:
                        continue
                    if best_branch is None or (dp - op) < (best_d_pos - best_o_pos):
                        best_line = lc
                        best_branch = br
                        best_o_pos = op
                        best_d_pos = dp

                # Multi-branch fallback: chain two branches that share a transfer
                # station. Find branch A containing o and branch B containing d
                # where some station is on both — rider boards at o on A and we
                # also assign to a B train at the transfer point.
                transfer_branch = None
                transfer_o_pos = transfer_t_pos_a = transfer_t_pos_b = transfer_d_pos = None
                if best_branch is None:
                    a_branches = [(lc, br, br["pos"][o]) for lc, br in candidates if o in br["pos"]]
                    b_branches = [(lc, br, br["pos"][d]) for lc, br in candidates if d in br["pos"]]
                    best_chain_len = None
                    for la, ba, op in a_branches:
                        for lb, bb, dp in b_branches:
                            if ba is bb:
                                continue
                            # find best transfer: a station that's after o on A and before d on B
                            for mn, ta in ba["pos"].items():
                                if ta <= op:
                                    continue
                                tb = bb["pos"].get(mn)
                                if tb is None or tb >= dp:
                                    continue
                                chain = (ta - op) + (dp - tb)
                                if best_chain_len is None or chain < best_chain_len:
                                    best_chain_len = chain
                                    best_line = la
                                    best_branch = ba
                                    best_o_pos = op
                                    best_d_pos = ta  # alight at transfer on A
                                    transfer_branch = bb
                                    transfer_t_pos_b = tb
                                    transfer_d_pos = dp
                                    transfer_line = lb

                if best_branch is None:
                    n_unassigned += 1
                    riders_unassigned += vol
                    continue

                # Find trains on this (line, branch) that arrive at origin during qhr.
                lb_key = (best_line, best_branch["branch_id"])
                tids = trains_by_lb.get(lb_key, [])
                if not tids:
                    n_unassigned += 1
                    riders_unassigned += vol
                    continue

                # qhr from midnight, no DAY_START_S offset.
                band_start = qhr * QHR_LEN_S
                band_end = band_start + QHR_LEN_S

                # Trains where stop_times[best_o_pos] is in [band_start, band_end)
                matching = []
                for tid in tids:
                    t = trains[tid]
                    if best_o_pos >= len(t["stop_times"]):
                        continue
                    arr = t["stop_times"][best_o_pos]
                    if band_start <= arr < band_end:
                        matching.append(tid)

                if not matching:
                    # Fall back: distribute across all trains in band by spawn time
                    for tid in tids:
                        t = trains[tid]
                        if band_start <= t["stop_times"][0] < band_end:
                            matching.append(tid)

                if not matching:
                    n_unassigned += 1
                    riders_unassigned += vol
                    continue

                # Distribute vol across matching trains. Spread randomly across
                # trains so we don't overload one.
                per_train = vol / len(matching)
                for tid in matching:
                    t = trains[tid]
                    t["boardings"][best_o_pos] += per_train
                    t["alightings"][best_d_pos] += per_train

                # Multi-branch trip: also assign to a train on the connecting
                # branch, boarding at the transfer station ~2 min after we
                # alighted there.
                if transfer_branch is not None:
                    tlb_key = (transfer_line, transfer_branch["branch_id"])
                    ttids = trains_by_lb.get(tlb_key, [])
                    transfer_band = (band_start + 2 * 60, band_end + 8 * 60)
                    tmatching = []
                    for tid in ttids:
                        t = trains[tid]
                        if transfer_t_pos_b >= len(t["stop_times"]):
                            continue
                        arr = t["stop_times"][transfer_t_pos_b]
                        if transfer_band[0] <= arr < transfer_band[1]:
                            tmatching.append(tid)
                    if tmatching:
                        per = vol / len(tmatching)
                        for tid in tmatching:
                            t = trains[tid]
                            t["boardings"][transfer_t_pos_b] += per
                            t["alightings"][transfer_d_pos] += per

                n_assigned += 1
                riders_assigned += vol

        total_assigned += n_assigned
        total_unassigned += n_unassigned
        total_riders_assigned += riders_assigned
        total_riders_unassigned += riders_unassigned
        coverage = riders_assigned / max(1, riders_assigned + riders_unassigned) * 100
        line_label = "+".join(line_codes)
        print(f"  {fname:14s} -> {line_label:12s}  "
              f"{n_assigned:6d} assigned, {n_unassigned:6d} unassigned  "
              f"({riders_assigned:>11,.0f} / {riders_unassigned + riders_assigned:>11,.0f} riders, "
              f"{coverage:5.1f}% covered)")

    overall_coverage = total_riders_assigned / max(1, total_riders_assigned + total_riders_unassigned) * 100
    print(f"  TOTAL: {total_assigned} assigned, {total_unassigned} unassigned, "
          f"{total_riders_assigned:,.0f} riders ({overall_coverage:.1f}% coverage)")


def build_output(trains, geo, line_colors):
    """Construct trains.json. Compatible-ish with nycriders viewer."""
    print("Building output...")
    stations_in = geo["stations"]

    # Per-station: set of lines that serve it (for station tooltips).
    station_routes = defaultdict(set)
    for br in geo["branches"]:
        for mnlc in br["mnlcs"]:
            station_routes[str(mnlc)].add(br["line"])

    # Build trains list. Drop trains with 0 riders to save bytes.
    trains_out = []
    for t in trains:
        boardings = t["boardings"]
        alightings = t["alightings"]
        if not boardings:
            continue
        riders_on = 0.0
        timeline = []
        for i, mnlc in enumerate(t["mnlcs"]):
            s = stations_in.get(str(mnlc))
            if not s:
                continue
            b = boardings.get(i, 0.0)
            a = alightings.get(i, 0.0)
            riders_on += b - a
            riders_on = max(0.0, riders_on)
            timeline.append([
                t["stop_times"][i],
                round(s["lat"], 5),
                round(s["lon"], 5),
                round(riders_on),
                round(b),
            ])
        # Drop ghost trains (e.g. spurious trains from over-counted shared
        # links at junctions). Real Tube trains routinely carry 200+ riders;
        # < 8 means almost certainly a duplicated/synthetic artifact.
        if not timeline or max(r[3] for r in timeline) < 8:
            continue
        trains_out.append({
            "route": t["line"],
            "color": t["color"],
            "timeline": timeline,
        })
    print(f"  {len(trains_out)} trains with riders")

    # Lines for drawing on the map
    lines_out = []
    for ln in geo["lines"]:
        lines_out.append({
            "route": ln["line"],
            "color": ln["color"],
            "coords": ln["coords"],
        })

    # Stations list (waiting timeline empty — not modeled). Index by integer
    # position so the viewer can reference si -> station; tooltip uses .name + .routes.
    waiting_out = []
    station_index_by_mnlc = {}
    for mnlc in sorted(stations_in.keys(), key=lambda m: int(m)):
        s = stations_in[mnlc]
        station_index_by_mnlc[mnlc] = len(waiting_out)
        waiting_out.append({
            "name": s["name"],
            "lat": s["lat"],
            "lon": s["lon"],
            "routes": sorted(station_routes.get(mnlc, [])),
            "timeline": [],
        })

    return {
        "lines": lines_out,
        "trains": trains_out,
        "waiting": waiting_out,
        "transfer_walks": [],
        "origin_walks": [],
        "dest_walks": [],
        "line_colors": line_colors,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--sample", type=float, default=1.0,
                        help="OD sample rate, 0.0-1.0 (default 1.0)")
    args = parser.parse_args()
    random.seed(123)

    t0 = time_module.time()
    geo = load_geometry()
    line_freqs = load_link_frequencies()
    line_descs = load_line_descriptions()
    line_colors = geo["line_colors"]

    branches = geo["branches"]
    trains = synthesize_trains(branches, line_freqs, geo["stations"],
                               line_colors, line_descs)

    assign_riders(trains, branches, geo["stations"], sample_rate=args.sample)

    out = build_output(trains, geo, line_colors)

    print(f"Writing {OUT_JSON}...")
    with open(OUT_JSON, "w") as f:
        json.dump(out, f)
    size_mb = os.path.getsize(OUT_JSON) / 1024 / 1024
    print(f"Done! {OUT_JSON} is {size_mb:.1f} MB ({time_module.time() - t0:.0f}s elapsed)")


if __name__ == "__main__":
    main()
