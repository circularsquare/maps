"""
Aggregate per-segment ridership from the MLIT 12th Census line-segment OD data.

Input:
  data/line_segment_riders.xlsx  — columns: 路線コード, 乗車駅コード, 初乗り乗継, 降車駅コード, 最終降車乗継, 人員
  data/codes/line_codes_kanto.xlsx
  data/codes/station_codes_kanto.xlsx

Output:
  data/processed/segments_kanto.json — { segments: [{line_code, line_name, op_name, from_code, from_name, to_code, to_name, riders}], stations: {...}, lines: {...} }

Station codes are zero-padded strings; ridership xlsx stores them as ints, so we zfill(5).
Stations on a line are ordered by station code (confirmed monotonic in peek).
Yamanote is stored as a linear sequence but is physically a loop; we handle loop lines specially
by picking the shorter direction when a trip could go either way.
"""

import collections
import io
import json
import os
import sys
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).parent
DATA = ROOT / 'data'
CODES = DATA / 'codes'
OUT = DATA / 'processed'
OUT.mkdir(exist_ok=True)

LOOP_LINES = {'010'}  # 山手線 (Yamanote) — only Kanto loop line


def load_stations():
    wb = openpyxl.load_workbook(CODES / 'station_codes_kanto.xlsx', read_only=True, data_only=True)
    ws = wb['駅コード']
    by_line = collections.defaultdict(list)   # line_code -> [(station_code, station_name)]
    meta = {}                                  # station_code -> {name, line_code, line_name, op_name, op_code}
    for r in list(ws.iter_rows(values_only=True))[1:]:
        code, op, line, stn, opc, linec = r
        if code is None:
            continue
        by_line[linec].append((code, stn))
        meta[code] = {'name': stn, 'line_code': linec, 'line_name': line, 'op_name': op, 'op_code': opc}
    # codes are strings like '00101'; sort ensures station order along the line
    for k in by_line:
        by_line[k].sort(key=lambda x: int(x[0]))
    return by_line, meta


def load_lines():
    wb = openpyxl.load_workbook(CODES / 'line_codes_kanto.xlsx', read_only=True, data_only=True)
    ws = wb['路線コード']
    out = {}
    for r in list(ws.iter_rows(values_only=True))[1:]:
        lc, op, line, opc = r
        if lc is None:
            continue
        out[lc] = {'op_name': op, 'line_name': line, 'op_code': opc}
    return out


def station_order_index(by_line):
    """line_code -> {station_code: position}"""
    return {lc: {c: i for i, (c, _) in enumerate(lst)} for lc, lst in by_line.items()}


def aggregate_segments(by_line, idx):
    """Return dict: (line_code, from_code, to_code) -> riders. from/to stored in line-order (from<to)."""
    loads = collections.defaultdict(int)
    wb = openpyxl.load_workbook(DATA / 'line_segment_riders.xlsx', read_only=True, data_only=True)
    ws = wb['線別駅間移動人員']
    skipped = {'unknown_station': 0, 'cross_line': 0, 'self_loop': 0, 'non_numeric': 0}
    total = 0

    for row in ws.iter_rows(values_only=True):
        lc_raw, board, _b_t, alight, _a_t, riders, *_ = row
        if not isinstance(lc_raw, int) or not isinstance(board, int) or not isinstance(alight, int):
            skipped['non_numeric'] += 1
            continue
        if not isinstance(riders, (int, float)):
            skipped['non_numeric'] += 1
            continue

        lc = f'{lc_raw:03d}'
        b = f'{board:05d}'
        a = f'{alight:05d}'

        line = by_line.get(lc)
        if line is None:
            skipped['unknown_station'] += 1
            continue
        stn_idx = idx[lc]
        if b not in stn_idx or a not in stn_idx:
            skipped['cross_line'] += 1
            continue
        if b == a:
            skipped['self_loop'] += 1
            continue

        i, j = stn_idx[b], stn_idx[a]
        n = len(line)

        if lc in LOOP_LINES:
            # pick the shorter arc
            forward = (j - i) % n
            backward = (i - j) % n
            if forward <= backward:
                path = [(i + k) % n for k in range(forward + 1)]
            else:
                path = [(i - k) % n for k in range(backward + 1)]
        else:
            if i < j:
                path = list(range(i, j + 1))
            else:
                path = list(range(j, i + 1))

        for k in range(len(path) - 1):
            p, q = path[k], path[k + 1]
            # canonicalize undirected segment by smaller stop-index first
            if p > q:
                p, q = q, p
            fc = line[p][0]
            tc = line[q][0]
            loads[(lc, fc, tc)] += int(riders)
        total += 1

    print(f'processed rows: {total}  skipped: {skipped}', file=sys.stderr)
    return loads


def main():
    by_line, station_meta = load_stations()
    line_meta = load_lines()
    idx = station_order_index(by_line)
    loads = aggregate_segments(by_line, idx)

    segments = []
    for (lc, fc, tc), riders in loads.items():
        lm = line_meta.get(lc, {})
        segments.append({
            'line_code': lc,
            'line_name': lm.get('line_name'),
            'op_name': lm.get('op_name'),
            'from_code': fc,
            'from_name': station_meta[fc]['name'],
            'to_code': tc,
            'to_name': station_meta[tc]['name'],
            'riders': riders,
        })
    segments.sort(key=lambda s: -s['riders'])

    payload = {
        'segments': segments,
        'stations': station_meta,
        'lines': line_meta,
    }
    out = OUT / 'segments_kanto.json'
    out.write_text(json.dumps(payload, ensure_ascii=False), encoding='utf-8')
    print(f'wrote {out}  segments={len(segments)}  stations={len(station_meta)}  lines={len(line_meta)}')
    print('\ntop 15 segments by riders:')
    for s in segments[:15]:
        print(f'  {s["riders"]:>8}  {s["op_name"]} {s["line_name"]}  {s["from_name"]} → {s["to_name"]}')


if __name__ == '__main__':
    main()
