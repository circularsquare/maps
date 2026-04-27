import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook('riders/tokyoriders/data/station_od.xlsx', read_only=True, data_only=True)
print(f"Sheets ({len(wb.sheetnames)}):")
for s in wb.sheetnames:
    ws = wb[s]
    print(f"  - {s!r}  max_row={ws.max_row}  max_col={ws.max_column}")

print("\n--- Each sheet, top 10 rows ---")
for s in wb.sheetnames:
    ws = wb[s]
    print(f"\n=== {s} ===")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 10: break
        print(row)

