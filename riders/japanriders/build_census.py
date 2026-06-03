"""
v2 census stitch — give the three metropolitan areas a real per-segment taper.

In the national base map lines are coarse: every non-JR line is a single
uniform feature, and many JR lines are one feature per route (gtfs-gis.jp
publishes one 輸送密度 per line section). This script replaces those, inside the
Tokyo / Nagoya / Osaka census regions, with station-to-station segments cut
from the real geometry.

Reads:
  data/tsukajinin_20260117.geojson          national base (via build.build_base)
  data/census/ekikan_tsuka_shutoken.xlsx    12th 大都市交通センサス 駅間通過人員, 首都圏
  data/census/ekikan_tsuka_chukyo.xlsx                                       中京圏
  data/census/ekikan_tsuka_kinki.xlsx                                        近畿圏
  ../tokyoriders/data/N02-24/UTF-8/N02-24_Station.geojson   国土数値情報 — station coords

Method ("reshape"):
  駅間通過人員 gives each segment's two-way passing volume (人/日, 2015 survey).
  Each census segment is cut from the real gtfs-gis.jp geometry by projecting
  its two station coordinates (resolved against N02) onto the nearest geometry
  strand. Each gtfs feature's census segments are then scaled together so their
  length-weighted average equals that feature's gtfs-gis.jp 輸送密度 — the taper
  *shape* is the census's, the *magnitude* stays on the base map's fiscal year.

  A census line is matched to gtfs features by line name within its region's
  bounding box; gtfs's bundled features (「京王線・高尾線・相模原線」) are indexed
  by each ・-component, and a JR route split into sections matches all of them.
  Per-segment cutting then routes each segment to the right feature/section by
  geometry, so collisions (JR 奈良線 vs 近鉄 奈良線) sort themselves out.

  Census lines with no 戸籍 line of their own — JR operational services
  (京浜東北・根岸線, 湘南新宿ライン, 上野東京ライン, 埼京線, the split 常磐/総武
  locals) — match no feature and are skipped: their traffic is already inside
  the 戸籍 lines' gtfs 輸送密度.

  A stitched line keeps the part of its geometry the census never reached —
  the extent outside the metro area — as uniform base-density segments, so a
  line running past the census boundary (飯田線, 近鉄大阪線) isn't truncated.

Writes:
  data/segments.geojson   (national base, the 3 census regions stitched)

Run directly: python build_census.py
"""
import collections
import io
import json
import re
import sys
import unicodedata
import warnings
from pathlib import Path

import openpyxl
from shapely.geometry import LineString, Point, box as shp_box, shape
from shapely.ops import linemerge, substring

import build  # build.build_base()

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
# off-geometry projections produce a NaN that cut_segment already guards against
warnings.filterwarnings('ignore', 'invalid value encountered in line_locate_point')

ROOT = Path(__file__).parent
CENSUS_DIR = ROOT / 'data' / 'census'
N02_STATION = ROOT.parent / 'tokyoriders' / 'data' / 'N02-24' / 'UTF-8' / 'N02-24_Station.geojson'
OUT = ROOT / 'data' / 'segments.geojson'

# census region -> (file, bounding box lon_min, lat_min, lon_max, lat_max)
REGIONS = [
    ('首都圏', 'ekikan_tsuka_shutoken.xlsx', (138.2, 34.6, 141.5, 37.8)),
    ('中京圏', 'ekikan_tsuka_chukyo.xlsx',   (135.5, 33.6, 138.5, 36.8)),
    ('近畿圏', 'ekikan_tsuka_kinki.xlsx',    (133.8, 33.2, 136.9, 35.9)),
]

# (region, census line) -> exact gtfs line name, for names plain matching can't
# bridge. Filled in iteratively from the build's "no gtfs match" report.
LINE_ALIAS = {
    ('首都圏', 'ブルーライン'): '1号線・3号線',
    ('首都圏', 'グリーンライン'): '4号線',
    ('首都圏', 'つくばエクスプレス'): '常磐新線',
    ('首都圏', 'みなとみらい線'): 'みなとみらい21線',
    ('首都圏', '東京モノレール羽田空港線'): '東京モノレール羽田線',
    ('首都圏', '箱根登山鉄道線'): '鉄道線',
    ('首都圏', '成田スカイアクセス線'): '成田空港線',
    ('中京圏', 'あおなみ線'): '西名古屋港線',
    ('近畿圏', '西神・山手線'): '西神線・山手線・北神線',
    ('首都圏', '日暮里・舎人ライナー'): '日暮里・舎人ライナー',   # ・ is in the name, not a bundle
    # gtfs-gis.jp ships all of Kintetsu's Kansai network as one feature labelled
    # 「大阪・京都・奈良線」, but the geometry covers every branch too. Without
    # these aliases each branch's census line ends up in the "no gtfs match"
    # bucket and the branch renders at the bundle's flat 全線 average.
    ('近畿圏', '難波線'):   '大阪・京都・奈良線',
    ('近畿圏', '南大阪線'): '大阪・京都・奈良線',
    ('近畿圏', '吉野線'):   '大阪・京都・奈良線',
    ('近畿圏', '橿原線'):   '大阪・京都・奈良線',
    ('近畿圏', '天理線'):   '大阪・京都・奈良線',
    ('近畿圏', '生駒線'):   '大阪・京都・奈良線',
    ('近畿圏', '道明寺線'): '大阪・京都・奈良線',
    ('近畿圏', '信貴線'):   '大阪・京都・奈良線',
    ('近畿圏', '御所線'):   '大阪・京都・奈良線',
    ('近畿圏', '長野線'):   '大阪・京都・奈良線',
    ('近畿圏', 'けいはんな線'): '大阪・京都・奈良線',
    # Same story for the Nagoya-side Kintetsu bundle.
    ('中京圏', '湯の山線'): '名古屋・三重線',
    ('中京圏', '鈴鹿線'):   '名古屋・三重線',
    # Same pattern — branches gtfs bundled into a parent feature whose label
    # doesn't mention them. Geometric cutting (MAX_OFFSET) rejects wrong
    # assignments, so wrong aliases just drop quietly rather than misroute.
    ('近畿圏', '今津線'):     '神戸線・宝塚線・伊丹線',
    ('近畿圏', '甲陽線'):     '神戸線・宝塚線・伊丹線',
    ('近畿圏', '箕面線'):     '神戸線・宝塚線・伊丹線',
    ('近畿圏', '武庫川線'):   '本線・阪神なんば線',
    ('近畿圏', '宇治線'):     '京阪本線・鴨東線・中之島線',
    ('近畿圏', '交野線'):     '京阪本線・鴨東線・中之島線',
    ('近畿圏', '汐見橋線'):   '高野線',
    ('近畿圏', '彩都線'):     '大阪モノレール線・国際文化公園都市モノレール線（彩都線）',
    ('近畿圏', 'ポートアイランド線ループ'): 'ポートアイランド線・六甲アイランド線',
    ('首都圏', '豊島線'):     '池袋線・西武秩父線',
    ('首都圏', '狭山線'):     '池袋線・西武秩父線',
    ('首都圏', '西武有楽町線'): '池袋線・西武秩父線',
    ('首都圏', '西武園線'):   '新宿線・拝島線',
    ('首都圏', '競馬場線'):   '新宿線・拝島線',
    ('首都圏', '国分寺線'):   '新宿線・拝島線',
    ('首都圏', '動物園線'):   '山口線',
    ('首都圏', '千葉都市モノレール１号線'): '1号線・2号線',
    ('首都圏', '千葉都市モノレール２号線'): '1号線・2号線',
    # JR support-line branches share strands with their parent 戸籍 line.
    ('首都圏', '南武支線'):       '南武線',
    ('首都圏', '成田支線（１）'): '成田線',
    ('首都圏', '成田支線（２）'): '成田線',
    ('近畿圏', '和田岬支線'):     '山陽線',
    ('近畿圏', '東羽衣支線'):     '阪和線',
    # 鉄 vs 鐵 — census uses the regular form, gtfs uses the old form.
    ('首都圏', '小湊鉄道線'):     '小湊鐵道線',
}

# JR operational services that aren't 戸籍 lines themselves — they ride trackage
# owned by one or more cadastral lines, and the census reports them under the
# service name. Without these mappings the per-segment census shape is dropped
# entirely AND the cadastral section the service rode on goes uncovered (the
# Tohoku Main Line south of Omiya — served by Keihin-Tohoku / Ueno-Tokyo Line /
# Shonan-Shinjuku trains — had no census shape at all). List the cadastral line
# names of every line the service physically uses; the existing geometric
# routing then assigns each station to whichever it's closest to. One service
# is one ridership stream, so the stitched shape under-portrays the aggregate
# unless other services on the same trackage are mapped too.
#
# (region, census line) -> list of gtfs line names the service runs on
OPERATIONAL_TRACKAGE = {
    # Keihin-Tohoku / Negishi: Omiya → Tokyo → Yokohama → Ofuna
    ('首都圏', '京浜東北・根岸線'): ['東北本線', '東海道本線', '根岸線'],
    # Ueno-Tokyo Line: through-service connecting Utsunomiya/Takasaki (Tohoku
    # Main Line north of Tokyo) with Tokaido rapid trains via Tokyo Station.
    # Two census entries (1) and (2) likely cover the two directions.
    ('首都圏', '上野東京ライン（1）'): ['東北本線', '東海道本線'],
    ('首都圏', '上野東京ライン（2）'): ['東北本線', '東海道本線'],
    # Saikyo Line: Osaki → Shinjuku → Ikebukuro → Akabane → Omiya. Yamanote
    # Line trackage between Osaki and Ikebukuro is intentionally excluded — the
    # Yamanote 戸籍 line's published 輸送密度 is for the Yamanote loop service
    # alone, so feeding Saikyo's lower per-segment values into the Yamanote
    # reshape would inflate the Yamanote service's per-segment numbers (the
    # weighted average drops, k rises, everything multiplies up). Saikyo's
    # Osaki~Ikebukuro stretch goes uncovered as a result; the cadastral Yamanote
    # still gets its own loop-service census.
    ('首都圏', '埼京線'): ['赤羽線', '東北本線'],
    # Shonan-Shinjuku Line — same reasoning as Saikyo for the Yamanote stretch.
    ('首都圏', '湘南新宿ライン'): ['東北本線', '赤羽線', '横須賀線', '東海道本線'],
    # Yamanote Line: the 戸籍 line covers only the western arc Shinagawa↔Tabata
    # via Shibuya/Shinjuku/Ikebukuro. The eastern arc (Tabata↔Tokyo↔Shinagawa)
    # runs on Tohoku Main Line + Tokaido Main Line trackage. The census reports
    # the full loop under "山手線"; this entry lets the eastern-arc stations
    # route onto Tohoku/Tokaido while keeping the western arc on Yamanote.
    ('首都圏', '山手線'): ['山手線', '東北本線', '東海道本線'],
    # Joban Line rapid + local both ride Joban Line trackage on parallel tracks
    ('首都圏', '常磐線快速'):     ['常磐線'],
    ('首都圏', '常磐線各駅停車'): ['常磐線'],
    # Sobu Line local — the rapid service is already covered via the plain
    # '総武線' census entry matching Sobu Main Line by name.
    ('首都圏', '総武線各駅停車'): ['総武本線'],
    # Nankai Koya Line: census reports the full operational service starting at
    # 難波, but cadastrally 高野線 only begins at 岸里玉出 — Nankai trains share
    # 南海本線 trackage between 難波 and 岸里玉出. Without this, the 難波-今宮戎
    # stretch projects onto the 汐見橋線 strand (the only nearby 高野線 geometry)
    # and the north of the line renders disconnected from the main trunk.
    ('近畿圏', '高野線'): ['高野線', '南海本線'],
}

# census station name (2015) -> N02 current name, for stations renamed 2015–2024
STATION_ALIASES = {
    '松原団地': '獨協大学前', '仲木戸': '京急東神奈川', '花月園前': '花月総持寺',
    '新逗子': '逗子・葉山', '産業道路': '大師橋', '南町田': '南町田グランベリーパーク',
    '西武遊園地': '多摩湖', '遊園地西': '多摩湖', '佐貫': '龍ケ崎市',
    '船の科学館': '東京国際クルーズターミナル', '国際展示場正門': '東京ビッグサイト',
}

# A census-truncated line keeps its outer (non-metro) extent as a "remainder".
# Without an entry here that remainder stays at the line's whole-line 輸送密度 —
# fine when the census covered little of the line, very wrong when it covered
# the busy metro half (東武伊勢崎線's rural tail then outweighed its own trunk).
# REMAINDER_ANCHOR overrides it with the traffic the line really carries past
# the census boundary; the remainder is then filled max(anchor, nearest census
# value):
#   - line MERGES into another — flow conservation at the junction:
#     tributary = downstream − Σ(other tributaries), worked out by hand from the
#     JR companies' published 区間別平均通過人員 (a junction of any number of
#     lines is just more terms in the subtraction);
#   - line ends at a TERMINUS — the value the census taper reached at its last
#     surveyed segment, i.e. continue flat to the end.
# The build's "uncovered remainders" report prints a census-based suggestion for
# every un-anchored line; use it to pick which are worth adding.
#
# Operators where gtfs-gis.jp ships the network as a single bundled feature
# (only one 輸送密度 for everything) but the operator publishes per-line daily
# boards. Used to split the bundle's total person-km proportional to boards so
# each line gets its own uniform density instead of all rendering at the
# operator-wide average. Trip-length-within-operator is assumed uniform — fine
# for small metros where every line carries similar character traffic.
#
# Values are 千人/day boards from operator releases / 日本地下鉄協会 publications.
# Year doesn't have to match the gtfs base year; only the *ratio* between lines
# matters, and metro line shares are stable year-over-year.
#
# (gtfs op, gtfs line) -> { component line name : boards/day in thousands }
PER_LINE_BOARDS = {
    # 仙台市 FY2023 — 仙台市交通局 / Wikipedia
    ('仙台市', '南北線・東西線'): {
        '南北線': 195,
        '東西線':  77,
    },
    # 福岡市 FY2024 — subway.city.fukuoka.lg.jp 運輸実績. 空港+箱崎 reported
    # jointly at 378k; split 91:9 per JSA's FY2016 ratio (geometry unchanged).
    ('福岡市', '空港線・箱崎線・七隈線'): {
        '空港線': 345,
        '箱崎線':  34,
        '七隈線': 145,
    },
}

# (gtfs op, gtfs line) -> remainder 輸送密度 (人/日)
REMAINDER_ANCHOR = {
    # 湖西線 → 近江塩津, merges into 北陸線 toward 敦賀 (junction conservation):
    #   北陸線 近江塩津~敦賀 22,554 − 北陸線 米原~近江塩津 9,281 = 13,273   (JR西 2024)
    ('西日本旅客鉄道', '湖西線'): 13273,
    # 東武伊勢崎線 → 伊勢崎 terminus; census taper reached 3,252 at 館林.
    ('東武鉄道', '伊勢崎線'): 3252,
}

SMALL_TO_LARGE_KANA = str.maketrans({
    'ッ': 'ツ', 'ャ': 'ヤ', 'ュ': 'ユ', 'ョ': 'ヨ', 'ァ': 'ア', 'ィ': 'イ',
    'ゥ': 'ウ', 'ェ': 'エ', 'ォ': 'オ', 'ヵ': 'カ', 'ヶ': 'ケ',
})

# a census segment whose stations sit further than this from every candidate
# strand doesn't belong on that geometry (e.g. census 山手線 is the whole loop,
# the gtfs feature only the western 戸籍 arc).
MAX_OFFSET = 0.015      # degrees, ≈ 1.5 km
K_MIN, K_MAX = 0.3, 4.0  # plausible census-2015 vs gtfs-FY ratio
# Metro census is a commuter survey — it doesn't capture tourist / leisure /
# non-commuter traffic, so lines that run into mountains, temples, or resort
# areas (Nankai Koya south of Hashimoto, Hakone Tozan, Eizan Kurama branch,
# etc.) come out spuriously thin past the suburban boundary. Pieces with raw
# census volume below this threshold get a soft floor at LOW_VOL_FLOOR_FRAC of
# the feature's whole-line D, so they don't disappear visually. Some lines
# (Tobu Tojo north of Ogawamachi — the metro region happens to extend into a
# genuinely sparse rural area) are exempted because the thin display is real.
LOW_VOL_THRESHOLD = 2000
LOW_VOL_FLOOR_FRAC = 0.08
LOW_VOL_FLOOR_EXEMPT = {
    ('東武鉄道', '東上本線'),
}


def norm_line(n):
    n = unicodedata.normalize('NFKC', n or '').strip()
    # subway lines: 「N号線(御堂筋線)」 / 「N号線東山線」 — keep the proper name
    m = re.match(r'^\d+号線[（(]?([^（()）]+線)[)）]?$', n)
    if m:
        return m.group(1)
    n = re.sub(r'[（(].*?[)）]', '', n)        # drop （１） branch suffixes etc.
    n = re.sub(r'^\d+号線', '', n)             # drop bare 「N号線」 prefix
    if n.endswith('本線'):
        n = n[:-2] + '線'
    return n.strip()


def line_components(line):
    """Split a (possibly bundled) gtfs line name into component line names.
    Within a bundle, gtfs abbreviates 「大阪線・京都線・奈良線」 to
    「大阪・京都・奈良線」 — re-add 線 to the bare components."""
    parts = [p.strip() for p in line.split('・') if p.strip()]
    if len(parts) == 1:
        return parts                                  # not a bundle — leave as-is
    return [p if '線' in p else p + '線' for p in parts]


def norm_stn(s):
    return unicodedata.normalize('NFKC', s or '').strip().translate(SMALL_TO_LARGE_KANA)


def in_box(x, y, box):
    return box[0] <= x <= box[2] and box[1] <= y <= box[3]


# ───────────────────────────── census parsing ────────────────────────────
def parse_census(path):
    """Census xlsx -> ordered [(line_name, [(station, down通過, up通過)])]."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out, cur = [], None
    for row in ws.iter_rows(values_only=True):
        c0 = row[0].strip() if isinstance(row[0], str) else row[0]
        if c0 is None or c0 == '駅名' or str(c0).startswith('駅別発着'):
            continue
        if all(v is None for v in row[1:7]):           # line-header row
            cur = (c0, [])
            out.append(cur)
        elif cur is not None and isinstance(row[15], (int, float)):
            cur[1].append((c0, row[15] or 0, row[18] or 0))   # 合計下り通過, 合計上り通過
    return out


def census_segments(stations):
    """[(stn, down, up)] -> [(from, to, two-way volume)]."""
    return [(stations[i][0], stations[i + 1][0], stations[i][1] + stations[i + 1][2])
            for i in range(len(stations) - 1)]


# ─────────────────────────── station coordinates ─────────────────────────
def load_n02():
    feats = json.loads(N02_STATION.read_text(encoding='utf-8'))['features']
    rows = []
    for f in feats:
        c = f['geometry']['coordinates']
        mid = c[len(c) // 2]
        rows.append((norm_line(f['properties']['N02_003']),
                     norm_stn(f['properties']['N02_005']), (mid[0], mid[1])))
    return rows


def region_coords(n02_rows, box):
    """Within the region box: (norm_line, norm_stn) -> coord, and norm_stn -> coord."""
    exact, loose = {}, {}
    for ln, st, (x, y) in n02_rows:
        if not in_box(x, y, box):
            continue
        exact.setdefault((ln, st), (x, y))
        loose.setdefault(st, (x, y))
    return exact, loose


def station_coord(name, line, exact, loose):
    cands = [norm_stn(name)]
    if name in STATION_ALIASES:
        cands.append(norm_stn(STATION_ALIASES[name]))
    ln = norm_line(line)
    for s in cands:
        if (ln, s) in exact:
            return exact[(ln, s)]
    for s in cands:
        if s in loose:
            return loose[s]
    return None


def resolve_line_segments(stations, line, exact, loose):
    """Pair each census segment with its station coordinates. Stations whose
    coordinate can't be resolved are dropped and the segments either side merge
    (mean volume) so the line stays continuous. Returns (segments, n_merged)."""
    raw = census_segments(stations)
    pts = [station_coord(s[0], line, exact, loose) for s in stations]
    kept = [j for j in range(len(stations)) if pts[j] is not None]
    out = []
    for a, b in zip(kept, kept[1:]):
        vols = [raw[j][2] for j in range(a, b)]
        out.append((stations[a][0], stations[b][0],
                    sum(vols) / len(vols), pts[a], pts[b]))
    return out, (len(stations) - 1) - len(out)


# ──────────────────────────── geometry cutting ───────────────────────────
def merged_parts(geom):
    g = shape(geom)
    if g.geom_type == 'LineString':
        return [g]
    m = linemerge(g)
    return list(m.geoms) if m.geom_type == 'MultiLineString' else [m]


def split_base_bundles(base):
    """Replace each PER_LINE_BOARDS bundle in `base` with one feature per
    component line, density proportional to boards. Total person-km is
    conserved across the split. Uses N02 station-on-line membership to assign
    each strand to its dominant line."""
    if not PER_LINE_BOARDS:
        return base

    # N02 station index: (operator, normalized-line) -> [coords]
    stns = collections.defaultdict(list)
    for f in json.loads(N02_STATION.read_text(encoding='utf-8'))['features']:
        p = f['properties']
        c = f['geometry']['coordinates']
        mid = c[len(c) // 2]
        op = unicodedata.normalize('NFKC', p['N02_004'] or '').strip()
        stns[(op, norm_line(p['N02_003']))].append(mid)

    out = []
    for f in base:
        gp = f['properties']
        cfg = PER_LINE_BOARDS.get((gp['op'], gp['line']))
        if cfg is None or gp['density'] is None:
            out.append(f)
            continue

        strands = merged_parts(f['geometry'])
        # Assign each strand to the component line whose N02 stations dominate it.
        line_strands = collections.defaultdict(list)
        for s in strands:
            best, best_score = None, 0
            for line_name in cfg:
                coords = stns.get((gp['op'], norm_line(line_name)), [])
                score = sum(1 for c in coords if s.distance(Point(c)) < 0.001)
                if score > best_score:
                    best, best_score = line_name, score
            if best is not None:
                line_strands[best].append(s)
            else:
                # Strand couldn't be classified — keep it bundled.
                line_strands.setdefault('__unassigned__', []).append(s)

        total_strand_len = sum(s.length for s in strands)
        total_boards = sum(cfg.values())
        total_pkm = gp['density'] * gp['km']        # bundle person-km/day

        for line_name, ss in line_strands.items():
            line_strand_len = sum(s.length for s in ss)
            # Bundle km is sliced proportional to each line's share of geometry.
            line_km = gp['km'] * line_strand_len / total_strand_len
            if line_name == '__unassigned__':
                density = gp['density']
                disp_name = gp['line']
            else:
                share = cfg[line_name] / total_boards
                density = (total_pkm * share) / line_km if line_km > 0 else 0
                disp_name = line_name
            for s in ss:
                props = dict(gp)
                props['line'] = disp_name
                props['density'] = round(density, 1)
                props['km'] = round(gp['km'] * s.length / total_strand_len, 3)
                # null out per-year detail since we're substituting an aggregate;
                # the split is only valid for the displayed year.
                for k in list(props):
                    if k.startswith('y') and k[1:].isdigit():
                        props[k] = None
                out.append({'type': 'Feature', 'geometry': s.__geo_interface__,
                            'properties': props})
    return out


def centroid(geom):
    g = shape(geom)
    c = g.centroid
    return c.x, c.y


def all_close_projections(strand, point, threshold):
    """All positions on `strand` where it comes within `threshold` of `point`.
    Groups consecutive close segments and keeps each run's local minimum, so a
    strand that revisits the same junction multiple times (loop + branch bundle
    like 名城線・名港線) yields multiple candidate positions instead of just the
    one shapely.project would return."""
    coords = list(strand.coords)
    out = []
    cum = 0.0
    in_run = False
    best = (float('inf'), 0.0)
    for i in range(len(coords) - 1):
        seg = LineString([coords[i], coords[i + 1]])
        seg_len = seg.length
        if seg_len > 0:
            t = seg.project(point)
            d = seg.interpolate(t).distance(point)
        else:
            t, d = 0.0, Point(coords[i]).distance(point)
        if d < threshold:
            if d < best[0]:
                best = (d, cum + t)
            in_run = True
        else:
            if in_run:
                out.append(best[1])
                best = (float('inf'), 0.0)
                in_run = False
        cum += seg_len
    if in_run:
        out.append(best[1])
    return out


def cut_segment(strands, a, b):
    """`strands` is [(LineString, owner)]. Cut the a–b segment from the nearest
    strand. Returns (piece, owner, strand, lo, hi) or None (degenerate / off
    every strand)."""
    pa, pb = Point(a), Point(b)
    strand, owner = min(strands, key=lambda so: max(so[0].distance(pa), so[0].distance(pb)))
    if min(strand.distance(pa), strand.distance(pb)) > MAX_OFFSET:
        return None
    lo, hi = sorted((strand.project(pa), strand.project(pb)))
    if lo != lo or hi != hi or hi - lo < 1e-7:       # NaN guard + degenerate
        return None
    piece = substring(strand, lo, hi)
    # If the naïve cut is much longer than the straight-line distance, the
    # strand probably revisits one of the points (a junction in a bundle that
    # traces a loop plus a branch, e.g. 名城線・名港線 passes through 金山 once
    # on each lap). Search all close projections and pick the pair giving the
    # shortest cut.
    if piece.length > 5 * pa.distance(pb):
        # Use a tighter threshold here than MAX_OFFSET: a station 1.5km off the
        # strand is plausibly the right strand, but a 1.5km "alternative" run on
        # a strand whose far end happens to lie ~1.5km from the station (大阪環状線:
        # 新今宮 station is 1.5km from 天王寺-the-strand-end) is a false positive
        # that yields nonsense short cuts.
        ALT_OFFSET = 0.005
        cands_a = all_close_projections(strand, pa, ALT_OFFSET)
        cands_b = all_close_projections(strand, pb, ALT_OFFSET)
        best = (piece.length, lo, hi)
        EPS = 1e-6
        for ap in cands_a or [lo]:
            for bp in cands_b or [hi]:
                lo2, hi2 = sorted((ap, bp))
                cut_len = hi2 - lo2
                if cut_len > EPS and cut_len < best[0]:
                    best = (cut_len, lo2, hi2)
        if best[0] < piece.length:
            lo, hi = best[1], best[2]
            piece = substring(strand, lo, hi)
    # Loop seam: a closed-loop strand (大阪環状線: 天王寺 → ... → 新今宮, missing
    # only the short 天王寺-新今宮 stretch) has its seam segment cut as nearly
    # the whole strand even after the alternative-projection search, because
    # neither station has a close projection at the other end of the strand.
    # Substitute a straight connector between the strand's endpoints (the real
    # physical gap) and mark the whole strand as covered. Strand-shape guard
    # (endpoints close) prevents this firing on non-loop bundles like
    # 名城線・名港線, whose strand endpoints are 6 km apart at different stations.
    coords = list(strand.coords)
    endgap = Point(coords[0]).distance(Point(coords[-1]))
    if (piece.length > 5 * pa.distance(pb)
            and piece.length > 0.5 * strand.length
            and endgap < MAX_OFFSET):
        # Don't claim coverage of the whole strand here — census might be
        # genuinely missing other segments along it (大阪環状線's census lists
        # 大阪 → 福島 → ... → 桜ノ宮 → 天満 with no 天満 → 大阪 wrap-around,
        # leaving a real ~1 km gap mid-loop). Marking (0, length) covered
        # would hide that as no remainder. Returning (None, None) tells the
        # caller to skip the coverage record — the seam connector still
        # renders as a piece, but actual missing-census gaps stay uncovered
        # and become remainders filled by the interior-gap rule.
        return LineString([coords[0], coords[-1]]), owner, strand, None, None
    return piece, owner, strand, lo, hi


def uncovered(length, intervals, eps=1e-4):
    """Complement of `intervals` within [0, length] — the spans of a strand the
    census never surveyed. eps (~10 m) drops slivers between adjacent pieces."""
    gaps, cur = [], 0.0
    for lo, hi in sorted(intervals):
        if lo - cur > eps:
            gaps.append((cur, lo))
        cur = max(cur, hi)
    if length - cur > eps:
        gaps.append((cur, length))
    return gaps


def census_floor(span, pieces, k):
    """Continuing-on density for a remainder span: the display density (raw
    census × k) of the geometrically nearest census piece with real data — what
    the census found where its coverage ends. Zero-value census pieces are
    boundary data gaps, not measurements, so they are skipped. Distance is
    spatial, so it still finds the right piece when a line merged into several
    geometry strands. 0.0 if the feature carries no census. `pieces` is
    [(piece, vol, fr, to)]."""
    best = None
    for piece, vol, *_ in pieces:
        if vol <= 0:
            continue
        d = span.distance(piece)
        if best is None or d < best[0]:
            best = (d, vol)
    return round(best[1] * k, 1) if best else 0.0


# ─────────────────────────────── main ────────────────────────────────────
def main():
    # keep_empty=True: also keep features gtfs-gis.jp never measured (density=None),
    # so census detail can be stitched onto their geometry directly.
    base, _ = build.build_base(keep_empty=True)
    base = split_base_bundles(base)
    n02 = load_n02()

    # per-region gtfs indexes + coordinate indexes
    regions = []
    for name, fname, box in REGIONS:
        comp = collections.defaultdict(list)   # norm_line(component) -> [features]
        exact_line = {}                        # full gtfs line name -> feature
        box_poly = shp_box(*box)
        # Index by geometry-intersects-box, not centroid-in-box. A line that
        # crosses the region but has its centroid elsewhere (東海道新幹線's
        # centroid is in 静岡; 山陽本線's is in 広島) was previously invisible
        # to the region's census, so e.g. 首都圏 東海道新幹線 hit no-match
        # despite the line obviously running through Tokyo.
        for f in base:
            if not shape(f['geometry']).intersects(box_poly):
                continue
            line = f['properties']['line']
            exact_line[line] = f
            for part in line_components(line):
                comp[norm_line(part)].append(f)
        cex, cloose = region_coords(n02, box)
        regions.append({'name': name, 'file': fname, 'box': box,
                        'comp': comp, 'exact': exact_line, 'cex': cex, 'cloose': cloose})

    def match(reg, line):
        """Census line -> list of gtfs features it covers."""
        if (reg['name'], line) in LINE_ALIAS:
            f = reg['exact'].get(LINE_ALIAS[(reg['name'], line)])
            return [f] if f else []
        if (reg['name'], line) in OPERATIONAL_TRACKAGE:
            out, seen = [], set()
            for trackage in OPERATIONAL_TRACKAGE[(reg['name'], line)]:
                for f in reg['comp'].get(norm_line(trackage), []):
                    if id(f) not in seen:
                        seen.add(id(f))
                        out.append(f)
            return out
        keys = {norm_line(line)}
        bare = re.sub(r'[（(].*?[)）]', '', unicodedata.normalize('NFKC', line)).strip()
        if bare.endswith('本線'):
            keys.add('線')                     # 京成本線 / 京浜急行本線 / bare 本線
        out, seen = [], set()
        for k in keys:
            for f in reg['comp'].get(k, []):
                if id(f) not in seen:
                    seen.add(id(f))
                    out.append(f)
        return out

    # Group census lines so every group contains a connected component of gtfs
    # features (two features are connected if any census line matches both).
    # Without this, the Kintetsu bundle 大阪・京都・奈良線 was being claimed by
    # census 奈良線's group (which also pulls in JR奈良) and then re-claimed by
    # 大阪線's group — whichever group ran first won, and the other group's
    # segments were silently dropped (line 378's `if id(f) in replaced`).
    skipped = []
    parent = {}   # id(feat) -> id(feat)  (union-find)
    feat_by_id = {}
    line_recs = []
    def find(i):
        while parent[i] != i:
            parent[i] = parent[parent[i]]
            i = parent[i]
        return i
    def union(i, j):
        ri, rj = find(i), find(j)
        if ri != rj:
            parent[ri] = rj
    for reg in regions:
        for line, stations in parse_census(CENSUS_DIR / reg['file']):
            feats = match(reg, line)
            if not feats:
                skipped.append((reg['name'], line, 'no gtfs match'))
                continue
            for f in feats:
                parent.setdefault(id(f), id(f))
                feat_by_id[id(f)] = f
            for f in feats[1:]:
                union(id(feats[0]), id(f))
            line_recs.append((reg, line, stations, [id(f) for f in feats]))

    groups = {}
    for reg, line, stations, ids in line_recs:
        root = find(ids[0])
        g = groups.setdefault(root, {'feats': {}, 'lines': []})
        for i in ids:
            g['feats'][i] = feat_by_id[i]
        # carry the per-line feat ids so routing can restrict each census line
        # to *its* matched cadastrals — without this, the supergroup's strands
        # all compete and a station shared between two cadastrals (e.g. 代々木
        # on both Yamanote and Chuo) can route its census segment to the wrong
        # parent line on geometric proximity alone.
        g['lines'].append((reg, line, stations, ids))

    replaced, stitched, report = set(), [], []
    remainder_report = []
    total_merged = total_remainder = 0
    for g in groups.values():
        feats = list(g['feats'].values())
        # strands cached per feature so the per-line strand lists below are
        # cheap to assemble (merged_parts is the costly call).
        feat_strands = {id(f): merged_parts(f['geometry']) for f in feats}

        by_feat = collections.defaultdict(list)        # id(feat) -> [(piece, vol, fr, to)]
        coverage = collections.defaultdict(list)       # id(strand) -> [(lo, hi)] census-covered spans
        for reg, line, stations, ids in g['lines']:
            # only this census line's matched cadastrals — not the whole
            # supergroup — get to compete in cut_segment's nearest-strand pick.
            line_strands = [(s, feat_by_id[i]) for i in ids for s in feat_strands[i]]
            segs, m = resolve_line_segments(stations, line, reg['cex'], reg['cloose'])
            total_merged += m
            for fr, to, vol, a, b in segs:
                cut = cut_segment(line_strands, a, b)
                if cut is not None:
                    piece, owner, strand, lo, hi = cut
                    by_feat[id(owner)].append((piece, vol, fr, to))
                    if lo is not None:
                        coverage[id(strand)].append((lo, hi))

        for f in feats:
            if id(f) in replaced:
                continue
            gp = f['properties']
            pieces = by_feat.get(id(f))
            if not pieces or len(pieces) < 2:
                continue
            census_len = sum(pc.length for pc, *_ in pieces)
            census_total = sum(v * pc.length for pc, v, *_ in pieces)
            census_density = census_total / census_len
            tag = f'{gp["op"]} {gp["line"]}'

            # km is spread over the line's whole geometry length, so the census
            # pieces and the remainder below together sum back to gp['km'].
            f_strands = feat_strands[id(f)]
            full_len = sum(s.length for s in f_strands) or census_len
            uncovered_len = full_len - census_len

            # Mode pick — two ways to make stitched + remainder reconcile with
            # the operator's whole-line total person-km (gp['density'] × L):
            #   reshape — scale every census piece by k = D / census_avg so the
            #     surveyed average matches D; fill the remainder with D. Shape
            #     comes from census, magnitude from operator. Works for any k in
            #     [K_MIN, K_MAX], but distorts segment magnitudes when k != 1.
            #   residual — leave the operator's whole-line baseline (D) on every
            #     piece (covered + uncovered) and add the census's deviation from
            #     a flat line on top of it. Concretely: each piece's density is
            #     `census_vol + offset` where offset = (D×L − Σ vol·len) / L, and
            #     the uncovered baseline is `offset` alone. The line's total
            #     person-km stays at D × L; census peaks rise above the baseline
            #     while the uncovered tail relaxes to (something ≤ D), with the
            #     gap covered by the census's actual measurements. Only valid
            #     when leftover > 0 (census measured less in total than the
            #     operator's flat baseline implies); bundle features whose
            #     published average is below their busiest surveyed sections
            #     (Sotetsu, Keihan) get a negative offset and fall back to reshape.
            EPS = 1e-6
            residual_offset = None      # leftover_per_meter, added to every piece in residual mode
            if gp['density'] is None:
                kind = 'direct'          # no gtfs 輸送密度: use census values as-is
                k = 1.0
            else:
                base_total = gp['density'] * full_len
                leftover = base_total - census_total
                if leftover > EPS * base_total:
                    kind = 'residual'
                    k = 1.0
                    residual_offset = leftover / full_len
                else:
                    kind = 'ok'
                    k = gp['density'] / census_density if census_density else 1.0
                    if not (K_MIN <= k <= K_MAX):
                        report.append(('skip', tag, len(pieces), k))
                        continue

            for pc, vol, fr, to in pieces:
                stitched.append({
                    'type': 'Feature',
                    'geometry': pc.__geo_interface__,
                    'properties': {
                        'op': gp['op'], 'line': gp['line'],
                        'from': build.normalize_name(gp['op'], fr),
                        'to':   build.normalize_name(gp['op'], to),
                        # base_from / base_to / base_km = the original gtfs-gis.jp
                        # base section's endpoints + full km. In national/base mode
                        # the section is the published-throughput unit, so right-
                        # click groups every cut piece sharing these.
                        'base_from': gp['from'], 'base_to': gp['to'],
                        'base_km': gp['km'],
                        'kind': gp['kind'], 'remark': gp['remark'],
                        'km': round((gp['km'] or 0) * pc.length / full_len, 3),
                        # 3 density views — index.html "data view" toggle picks one:
                        #   d_base = the line's whole-line gtfs-gis.jp value (same
                        #     across every cut piece of this line);
                        #   d_census = the raw 2015 census passing volume, before
                        #     the reshape × k scaling (so the un-reshaped shape
                        #     is visible side-by-side with d_base for diagnosis);
                        #   density = the stitched value (vol × k) — what the
                        #     map shows by default.
                        'density': round(
                            max(vol * k + (residual_offset or 0.0),
                                LOW_VOL_FLOOR_FRAC * gp['density'])
                            if (vol < LOW_VOL_THRESHOLD
                                and gp['density'] is not None
                                and (gp['op'], gp['line']) not in LOW_VOL_FLOOR_EXEMPT)
                            else vol * k + (residual_offset or 0.0),
                            1),
                        'd_base': gp['density'],
                        'd_census': round(vol, 1),
                        'year': gp['year'] if gp['year'] is not None else 2015,
                        'detail': 'census',
                    },
                })

            # The census only surveys each metropolitan area, so a line running
            # past the census boundary (飯田線, 近鉄大阪線, …) loses its outer
            # portion when the original feature is replaced. Keep that un-surveyed
            # remainder — but the base map filled it with the line's *whole-line*
            # 輸送密度, badly wrong when the surveyed part is the busy metro half
            # (湖西線's rural north then claimed 33,000/day). Fill it instead with
            # max(junction anchor, census floor):
            #   anchor — what the line hands off where it merges into another,
            #            from REMAINDER_ANCHOR (flow conservation at the junction);
            #   floor  — the census's own lowest surveyed value on this line, so
            #            the remainder never drops below where the census left off.
            # Skipped for census-direct lines — they have no gtfs density.
            # The census only surveys each metropolitan area, so a line running
            # past the census boundary (飯田線, 近鉄大阪線, …) keeps its outer
            # extent as a "remainder". Fill priority (see the branch below):
            # interior gap / short tip → adjacent census; REMAINDER_ANCHOR →
            # max(anchor, census); a long endpoint stub with a credible census
            # edge → min(whole-line, census edge), i.e. continue flat without
            # inflating; otherwise (no/low census edge) → the whole-line default.
            # Skipped for census-direct lines — they have no gtfs density to fall
            # back on. `floor` is recorded for every remainder so the report can
            # suggest anchors.
            if gp['density'] is not None:
                anchor = REMAINDER_ANCHOR.get((gp['op'], gp['line']), 0.0)
                for s in f_strands:
                    EPS_GAP = 1e-4
                    # an interior gap is one with census coverage on BOTH sides;
                    # endpoint gaps sit at the very start or end of the strand
                    # (the rural tail past the census or a terminus stub).
                    for lo, hi in uncovered(s.length, coverage.get(id(s), [])):
                        is_endpoint = (lo < EPS_GAP) or (hi > s.length - EPS_GAP)
                        span = substring(s, lo, hi)
                        floor = census_floor(span, pieces, k)
                        # Short uncovered tails at a strand's end (the last few
                        # hundred metres past a terminus, where the strand
                        # extends slightly beyond the station midpoint) should
                        # inherit the adjacent census density — otherwise they
                        # jump abruptly to the whole-line average and render as
                        # a thick island at the line's tip. Long remainders
                        # (the rural extent past the census region) keep the
                        # whole-line average (or anchor).
                        SHORT_TAIL = 0.005   # ~500 m
                        # In residual mode, uncovered baseline is the offset;
                        # in reshape, it's the operator's whole-line density.
                        # SHORT_TAIL inherits from the adjacent census piece —
                        # in residual mode that means the offset-adjusted value,
                        # since floor here is raw vol (k=1) and the displayed
                        # density on the neighbour is vol + offset.
                        if residual_offset is not None:
                            default = residual_offset
                            floor_display = floor + residual_offset
                        else:
                            default = gp['density']
                            floor_display = floor
                        # Interior gap (census on both sides): use the neighbour
                        # density unconditionally — a 1 km uncovered patch mid
                        # loop falling to the bare offset/baseline reads as an
                        # unrealistic thin segment between thick census pieces
                        # (the 名城線 loop just above 大曽根 was 1,828 vs the 60k
                        # of the segments around it). Endpoint gaps still use
                        # SHORT_TAIL for tip artefacts and otherwise fall back
                        # to anchor / default for long rural extensions.
                        if not is_endpoint and floor > 0:
                            fill = floor_display
                        elif span.length < SHORT_TAIL and floor > 0:
                            fill = floor_display
                        elif anchor:
                            fill = max(anchor, floor_display)
                        elif floor >= LOW_VOL_THRESHOLD:
                            # A long endpoint stub of an otherwise census-covered
                            # line (Midosuji's tail to 新金岡; 東海道線 past 彦根
                            # toward 米原): continue flat from the census edge
                            # rather than jumping to the busier whole-line average
                            # — a remainder past the boundary is an extension of
                            # the line, never busier than where the census left
                            # off. Capped at the whole-line average so a census
                            # edge that happens to read higher (武蔵野線, 相模本線)
                            # can't inflate the tail. A spuriously low census edge
                            # (rural lines the survey barely sampled — 伊勢線's 77)
                            # falls through to the whole-line default below.
                            fill = min(default, floor_display)
                        else:
                            fill = default
                        rkm = round((gp['km'] or 0) * (hi - lo) / full_len, 3)
                        # Identity-grouping: every remainder on a line used to
                        # share gp['from']/gp['to'], so right-clicking one
                        # highlighted every remainder along the whole line. Pin
                        # each remainder to its nearest census piece's from/to
                        # so it groups with that one adjacent segment instead.
                        rem_from, rem_to = gp['from'], gp['to']
                        best_d = float('inf')
                        for pc, _vol, cfr, cto in pieces:
                            d = span.distance(pc)
                            if d < best_d:
                                best_d = d
                                rem_from, rem_to = cfr, cto
                        stitched.append({
                            'type': 'Feature',
                            'geometry': span.__geo_interface__,
                            'properties': {
                                'op': gp['op'], 'line': gp['line'],
                                'from': build.normalize_name(gp['op'], rem_from),
                                'to':   build.normalize_name(gp['op'], rem_to),
                                'base_from': gp['from'], 'base_to': gp['to'],
                                'base_km': gp['km'],
                                'kind': gp['kind'], 'remark': gp['remark'],
                                'km': rkm,
                                # d_census = None: the remainder lies past the
                                # census, so there is no raw census number for it
                                'density': fill,
                                'd_base': gp['density'],
                                'd_census': None,
                                'year': gp['year'],
                                'detail': 'remainder',
                            },
                        })
                        total_remainder += 1
                        remainder_report.append((tag, rkm, floor, anchor, fill,
                                                 gp['density']))
            replaced.add(id(f))
            report.append((kind, tag, len(pieces), k))

    # drop features gtfs-gis.jp never measured that didn't get census detail.
    # base features that survive unstitched get d_base/d_census in line with the
    # stitched ones so the "data view" toggle can read the same property names.
    final = []
    for f in base:
        if id(f) in replaced or f['properties']['density'] is None:
            continue
        p = f['properties']
        p['d_base'] = p['density']
        p['d_census'] = None
        # uncut passthrough: section == segment, so the base-grouping fields
        # are identity with the displayed from/to/km.
        p['base_from'] = p['from']
        p['base_to'] = p['to']
        p['base_km'] = p['km']
        final.append(f)
    final += stitched
    OUT.write_text(json.dumps({'type': 'FeatureCollection', 'features': final},
                              ensure_ascii=False), encoding='utf-8')

    # ── report ──
    ok = [r for r in report if r[0] == 'ok']
    residual = [r for r in report if r[0] == 'residual']
    direct = [r for r in report if r[0] == 'direct']
    kskip = [r for r in report if r[0] == 'skip']
    n_base = sum(1 for f in base if f['properties']['density'] is not None)
    print(f'stitched {len(replaced)} gtfs features -> {len(stitched)} segments'
          f'  ({total_merged} census gaps merged, {total_remainder} un-surveyed remainders kept)')
    print(f'output: {len(final)} features ({n_base} base, {len(replaced)} replaced)\n')
    print(f'reshaped to gtfs density — k = D / census_avg ({len(ok)} lines):')
    for _, tag, nseg, k in sorted(ok, key=lambda r: -r[2]):
        print(f'  {nseg:>3} seg  k={k:>5.2f}  {tag}')
    if residual:
        print(f'\nresidual — raw census on covered, leftover person-km on remainder ({len(residual)} lines):')
        for _, tag, nseg, k in sorted(residual, key=lambda r: -r[2]):
            print(f'  {nseg:>3} seg  {tag}')
    if direct:
        print(f'\ncensus-direct — no gtfs 輸送密度, census values used as-is ({len(direct)} lines):')
        for _, tag, nseg, k in sorted(direct, key=lambda r: -r[2]):
            print(f'  {nseg:>3} seg  {tag}')
    if kskip:
        print(f'\nkept uniform — implausible reshape k ({len(kskip)}):')
        for _, tag, nseg, k in sorted(kskip, key=lambda r: -r[3]):
            print(f'  k={k:>6.1f}  ({nseg} seg)  {tag}')
    if remainder_report:
        agg = {}   # tag -> [total_km, longest_km, (floor, anchor, fill, old)]
        for tag, rkm, floor, anchor, fill, old in remainder_report:
            a = agg.setdefault(tag, [0.0, -1.0, None])
            a[0] += rkm
            if rkm > a[1]:
                a[1], a[2] = rkm, (floor, anchor, fill, old)
        n_anchored = sum(1 for v in agg.values() if v[2][1])
        print(f'\nuncovered remainders — outer extent kept, filled max(anchor, '
              f'census floor) ({len(remainder_report)} pieces, {len(agg)} lines, '
              f'{n_anchored} anchored). "was -> fill" is each line\'s longest '
              f'remainder piece; longest first:')
        for tag, (tot, lng, (floor, anchor, fill, old)) in sorted(
                agg.items(), key=lambda kv: -kv[1][0]):
            note = (f'[anchor {anchor:,.0f}]' if anchor
                    else f'(census suggests ~{floor:,.0f})')
            print(f'  {tot:>6.1f} km  {old:>9,.0f} -> {fill:>9,.0f}  {note}  | {tag}')
    print(f'\nno gtfs match ({len(skipped)} census lines):')
    for rname, line, _ in skipped:
        print(f'  {rname}  {line}')


if __name__ == '__main__':
    main()
