"""Romania — UAT boundaries for the RPL 2021 religion data.

Writes data/geo/ro/ro_uat.gpkg, 3,181 polygons, the single layer countries.py reads, and
prints the join against data/normalized/ro.csv both ways.

THE PROBLEM: the census file has no SIRUTA code. Rows are `MUNICIPIUL ALBA IULIA` and
`Abrud` and nothing else, so the only key available is (judeţ, name) — and Romanian place
names are not unique nationally. `Călăraşi` is a county AND three separate communes;
`Păuleşti` is a commune in both Prahova and Satu Mare.

THE BRIDGE: the Eurostat **LAU 2021 – NUTS 2021 correspondence table**, which carries
`NUTS 3 CODE`, `LAU CODE` (the SIRUTA code, which is what GISCO uses as `LAU_ID`) and
`LAU NAME NATIONAL` **including the type prefix** — `Municipiul Oradea`, `Oraş Aleşd`,
`Sânmartin` — in exactly the form the census writes them. Romanian NUTS3 regions ARE the
judeţe, one for one, so (NUTS3, name) and (judeţ, name) are the same key under a rename.

    https://ec.europa.eu/eurostat/documents/345175/501971/EU-27-LAU-2021-NUTS-2021.xlsx

The NUTS3-to-judeţ rename is DERIVED rather than hard-coded: each NUTS3's set of LAU names
is matched against each county's set of UAT names and the best overlap wins. That has to
come out a perfect bijection over all 42, and the script fails if it does not — which is a
stronger check than a hand-typed lookup, because a hand-typed lookup cannot notice that
the boundary file has moved a commune between counties.

## ş vs ș — the encoding trap that looks like a vintage problem

`ş` U+015F (s-cedilla) and `ș` U+0219 (s-comma-below) are different codepoints, and so are
`ţ`/`ț`. Romanian orthography wants comma-below; a great deal of software emits cedilla.
INS writes the census with comma-below and Eurostat writes the same names with cedilla.
Folding both to ASCII is what makes the join work at all — without it roughly a third of
Romanian place names miss, which reads exactly like a boundary-vintage mismatch and is not
one. `sources/ro.py:fold()` is the shared implementation.

Usage:
    python sources/ro_geo.py            # needs the GISCO zip and the correspondence xlsx
    python sources/ro_geo.py --fetch    # download both first
"""

import os
import sys
import zipfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from ro import fold  # noqa: E402  the same folding the census parser uses

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

GEO = os.path.join(ROOT, "data", "geo", "lau2021")
OUTER_ZIP = os.path.join(GEO, "ref-lau-2021-01m.shp.zip")
INNER_NAME = "LAU_RG_01M_2021_4326.shp.zip"
SHP_DIR = os.path.join(GEO, "shp4326")
CORR = os.path.join(GEO, "EU-27-LAU-2021-NUTS-2021.xlsx")

OUT_DIR = os.path.join(ROOT, "data", "geo", "ro")
OUT = os.path.join(OUT_DIR, "ro_uat.gpkg")
NORMALIZED = os.path.join(ROOT, "data", "normalized", "ro.csv")

LAU_URL = ("https://gisco-services.ec.europa.eu/distribution/v2/lau/download/"
           "ref-lau-2021-01m.shp.zip")
CORR_URL = ("https://ec.europa.eu/eurostat/documents/345175/501971/"
            "EU-27-LAU-2021-NUTS-2021.xlsx")

EXPECTED_UNITS = 3181
EXPECTED_COUNTIES = 42


def squash(s):
    """fold(), and also treat a hyphen as a space and collapse runs of spaces.

    INS and Eurostat disagree about hyphens in compound place names in both directions:
    `Piatra-Neamţ` vs `Piatra Neamţ`, `Slănic-Moldova` vs `Slănic Moldova`,
    `Ciceu-Mihăieşti` vs `Ciceu - Mihaieşti`. The census also has one stray space,
    `ORAŞ LEHLIU- GARĂ`.
    """
    return " ".join(fold(s).replace("-", " ").split())


def fetch():
    import requests
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    os.makedirs(GEO, exist_ok=True)
    for url, dest, minb in ((LAU_URL, OUTER_ZIP, 90_000_000),
                            (CORR_URL, CORR, 6_000_000)):
        if os.path.exists(dest) and os.path.getsize(dest) >= minb:
            print("already have", os.path.basename(dest))
            continue
        print("downloading", url)
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=900,
                         verify=False, stream=True)
        r.raise_for_status()
        with open(dest, "wb") as fh:
            for chunk in r.iter_content(1 << 20):
                fh.write(chunk)
        if os.path.getsize(dest) < minb:
            raise SystemExit(f"{dest} is short, expected >= {minb:,}")
        print(f"  {os.path.getsize(dest):,} bytes")


def unpack():
    if not os.path.exists(OUTER_ZIP):
        raise SystemExit(f"missing {OUTER_ZIP} -- run with --fetch first")
    inner = os.path.join(GEO, INNER_NAME)
    if not os.path.exists(inner):
        with zipfile.ZipFile(OUTER_ZIP) as z:
            z.extract(INNER_NAME, GEO)
    if not os.path.isdir(SHP_DIR):
        with zipfile.ZipFile(inner) as z:
            z.extractall(SHP_DIR)
    shps = [f for f in os.listdir(SHP_DIR) if f.endswith(".shp")]
    if not shps:
        raise SystemExit(f"no .shp under {SHP_DIR}")
    return os.path.join(SHP_DIR, shps[0])


def read_correspondence():
    import openpyxl
    if not os.path.exists(CORR):
        raise SystemExit(f"missing {CORR} -- run with --fetch first")
    wb = openpyxl.load_workbook(CORR, read_only=True, data_only=True)
    if "RO" not in wb.sheetnames:
        raise SystemExit(f"no RO sheet in the correspondence table: {wb.sheetnames}")
    ws = wb["RO"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        nuts3, code, name_nat = r[0], r[1], r[2]
        if not nuts3 or code is None or not name_nat:
            continue
        rows.append((str(nuts3).strip(), str(code).strip(),
                     " ".join(str(name_nat).split())))
    return rows


def main():
    import geopandas as gpd
    import pandas as pd

    if "--fetch" in sys.argv:
        fetch()
    shp = unpack()

    corr = read_correspondence()
    print(f"correspondence: {len(corr):,} RO rows, "
          f"{len({n for n, _, _ in corr})} NUTS3 regions")
    if len(corr) != EXPECTED_UNITS:
        raise SystemExit(f"correspondence has {len(corr)} RO rows, "
                         f"expected {EXPECTED_UNITS}")

    # ---- census side
    df = pd.read_csv(NORMALIZED, dtype={"geo_id": str}, low_memory=False)
    cen = df[df["geo_level"] == "uat"][["geo_id", "geo_name"]].drop_duplicates("geo_id")
    cen[["county", "uat"]] = cen["geo_id"].str.split("|", n=1, expand=True)
    print(f"census: {len(cen):,} UATs in {cen['county'].nunique()} counties")

    # ---- derive NUTS3 <-> county, by name-set overlap
    corr_by_nuts = {}
    for nuts3, code, name in corr:
        corr_by_nuts.setdefault(nuts3, set()).add(fold(name))
    cen_by_county = {}
    for _, r in cen.iterrows():
        cen_by_county.setdefault(r["county"], set()).add(fold(r["uat"]))

    pairs, used = {}, set()
    for county, names in sorted(cen_by_county.items(), key=lambda kv: -len(kv[1])):
        best, score = None, -1
        for nuts3, lnames in corr_by_nuts.items():
            if nuts3 in used:
                continue
            s = len(names & lnames)
            if s > score:
                best, score = nuts3, s
        pairs[county] = (best, score, len(names))
        used.add(best)

    bad = {c: v for c, v in pairs.items() if v[1] < v[2]}
    print(f"\n  NUTS3 <-> judet: {len(pairs)} pairs, "
          f"{len(pairs) - len(bad)} matching every UAT name")
    for c, (n, s, t) in sorted(bad.items(), key=lambda kv: kv[1][1] - kv[1][2])[:12]:
        print(f"      {c:<22} {n}  {s}/{t} names matched")
    if len(pairs) != EXPECTED_COUNTIES:
        raise SystemExit(f"{len(pairs)} counties, expected {EXPECTED_COUNTIES}")

    # ---- the join: (county, name) -> SIRUTA, in three passes
    cen["nuts3"] = cen["county"].map(lambda c: pairs[c][0])
    corr_by_nuts_full = {}
    for nuts3, code, name in corr:
        corr_by_nuts_full.setdefault(nuts3, []).append((name, code))

    resolved, by_pass = {}, {"exact": 0, "squashed": 0, "elimination": 0}
    leftovers = []
    for county, sub in cen.groupby("county"):
        nuts3 = pairs[county][0]
        lau = corr_by_nuts_full[nuts3]

        exact = {fold(n): c for n, c in lau}
        taken, remaining_cen = set(), []
        for _, r in sub.iterrows():
            code = exact.get(fold(r["uat"]))
            if code is not None and code not in taken:
                resolved[r["geo_id"]] = code
                taken.add(code)
                by_pass["exact"] += 1
            else:
                remaining_cen.append(r)

        # pass 2: hyphens and stray spaces only. `Piatra-Neamţ` vs `Piatra Neamţ`,
        # `Ciceu-Mihăieşti` vs `Ciceu - Mihaieşti`, `Lehliu- Gară` vs `Lehliu Gară`.
        if remaining_cen:
            squashed = {squash(n): c for n, c in lau if c not in taken}
            still = []
            for r in remaining_cen:
                code = squashed.get(squash(r["uat"]))
                if code is not None and code not in taken:
                    resolved[r["geo_id"]] = code
                    taken.add(code)
                    by_pass["squashed"] += 1
                else:
                    still.append(r)
            remaining_cen = still

        # pass 3: elimination. If ONE census UAT and ONE Eurostat LAU are left unmatched
        # inside a county, they are the same place and no spelling rule is needed —
        # `Râşca`/`Rişca`, `Suhurlui`/`Suhurului`, `Sfântu`/`Sfântul Gheorghe`,
        # `Dobroeşti`/`Dobroieşti` are all genuine variant spellings, not encodings.
        # Two or more left over is a guess, and this refuses to make it.
        if remaining_cen:
            free = [c for _, c in lau if c not in taken]
            if len(remaining_cen) == 1 and len(free) == 1:
                resolved[remaining_cen[0]["geo_id"]] = free[0]
                by_pass["elimination"] += 1
            else:
                for r in remaining_cen:
                    leftovers.append((county, r["uat"], len(remaining_cen), len(free)))

    cen["kod"] = cen["geo_id"].map(resolved)
    print(f"\n  resolved by: {by_pass['exact']:,} exact name, "
          f"{by_pass['squashed']} hyphen/space, {by_pass['elimination']} elimination")

    unresolved = cen[cen["kod"].isna()]
    print(f"  {'OK ' if unresolved.empty else 'BAD'} census UATs with no SIRUTA code: "
          f"{len(unresolved)}")
    for county, uat, nc, nf in leftovers[:15]:
        print(f"      {county} | {uat}   ({nc} census vs {nf} eurostat left in county)")
    if not unresolved.empty:
        raise SystemExit("name resolution FAILED")
    if cen["kod"].nunique() != len(cen):
        dup = cen[cen["kod"].duplicated(keep=False)].sort_values("kod")
        print(dup.head(10).to_string())
        raise SystemExit("two census UATs resolved to the same SIRUTA code")

    # ---- polygons
    gdf = gpd.read_file(shp)
    ro = gdf[gdf["CNTR_CODE"] == "RO"].copy()
    ro["kod"] = ro["LAU_ID"].astype(str).str.strip()
    print(f"\n  GISCO RO polygons: {len(ro):,}")

    geo_keys, cen_keys = set(ro["kod"]), set(cen["kod"])
    missing, extra = cen_keys - geo_keys, geo_keys - cen_keys
    print(f"  {'OK ' if not missing else 'BAD'} census UATs with no polygon: {len(missing)}")
    print(f"  {'OK ' if not extra else 'BAD'} polygons with no census UAT: {len(extra)}")
    for k in sorted(missing)[:10]:
        print("      no polygon:", k, cen[cen["kod"] == k]["geo_id"].iloc[0])
    for k in sorted(extra)[:10]:
        print("      no data   :", k, ro[ro["kod"] == k]["LAU_NAME"].iloc[0])
    if missing or extra:
        raise SystemExit("join FAILED")

    out = ro[["kod", "LAU_NAME", "POP_2021", "AREA_KM2", "geometry"]].rename(
        columns={"LAU_NAME": "name", "POP_2021": "pop_2021", "AREA_KM2": "area_km2"})
    os.makedirs(OUT_DIR, exist_ok=True)
    out.to_file(OUT, layer="uat", driver="GPKG")
    print("\nwrote", OUT, f"({len(out):,} polygons)")

    # the census carries no code, so countries.py needs the resolved map on disk
    lut = os.path.join(OUT_DIR, "ro_uat_lookup.csv")
    cen[["geo_id", "kod", "county", "uat"]].to_csv(lut, index=False, encoding="utf-8")
    print("wrote", lut, "-- (county|name) -> SIRUTA, which countries.py joins on")

    tot = out["pop_2021"].sum()
    print("\n  largest single polygons (the spec §8.2 caveat for Romania):")
    for _, r in out.nlargest(5, "pop_2021").iterrows():
        print(f"    {r['name']:<22} {r['pop_2021']:>10,.0f}  {r['area_km2']:>7.1f} km2  "
              f"{100 * r['pop_2021'] / tot:5.2f}% of RO in one unit")


if __name__ == "__main__":
    main()
