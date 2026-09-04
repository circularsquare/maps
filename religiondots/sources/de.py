"""Germany — Zensus 2022 religion, at Gemeinde.

Reads (or fetches) data/raw/de/ and writes data/normalized/de.csv.

Germany is the §3.9 trade-off at its most extreme, and it runs the opposite way to every
other country on the map.  The geography is the finest anywhere in the project — 10,786
Gemeinden, and destatis publishes THE SAME NUMBERS on a 100m INSPIRE grid with 3,088,036
populated cells.  The categories are three:

    Roemisch-katholische Kirche (oeffentlich-rechtlich)   20,746,959   25.1%
    Evangelische Kirche (oeffentlich-rechtlich)           19,127,360   23.1%
    Sonstige, keine, ohne Angabe                          42,845,220   51.8%

NOBODY WAS ASKED.  This is the fact that decides everything else about the German map.
Zensus 2022 has no religion question on the form; the numbers come out of the
Melderegister, which records religious-body membership because it determines church-tax
liability.  So `basis` is `roll` and not `self_id` — an institution's records, like
ASARB, and not a person's answer (spec §3.1).  It also means the data can only ever see
bodies that levy church tax, which is why the third category is more than half the
country.  Destatis says so itself in Datensatzbeschreibung_Religion_Gitterzellen.xlsx:

    "Fuer diese anderen oeffentlich-rechtlichen Religionsgesellschaften liegen nur in
     sehr begrenztem Umfang Eintraege im Melderegister vor, die die entsprechenden
     Zugehoerigkeiten nicht zuverlaessig statistisch abbilden koennen, weshalb auf den
     Nachweis verzichtet werden muss."

So "Sonstige, keine, ohne Angabe" holds, undifferentiated and unsplittable: Germany's
roughly four million Muslims, its Orthodox Christians, the Jewish communities, the
Freikirchen, everyone who belongs to nothing, and everyone the register has no entry
for.  It is NOT `unaffiliated` — nobody reported no religion — and it is not `other`,
which means a religion the source declined to name.  It maps to a node of its own; see
taxonomy/de2022.py.

WHY ZENSUS 2011 DOES NOT RESCUE IT.  2011 did ask, with two questions, and all three of
its problems compound (Statistische Monatshefte Niedersachsen 8/2014, sources/de.md §2):
Frage 8 — the one carrying Islam, Buddhism and Hinduism — was the only voluntary
question on the form and most people skipped it; it was also CONDITIONAL, put only to
people who had just said they belonged to no public-law body, which is the Northern
Ireland trap of spec §3.1 exactly; and the richer Frage 7 breakdown (Freikirchen,
Orthodoxe, Juedische Gemeinden) came from the ~10% household sample.  At Gemeinde level
2011 had THE SAME THREE CATEGORIES, because Gemeinde figures are counted register data:
"Dabei war lediglich eine Unterscheidung zwischen der Zugehoerigkeit zur
Roemisch-katholischen Kirche, zur Evangelischen Kirche oder zu Sonstigen moeglich."

WHAT THE FILE DOES WELL, and it is a lot.  There is essentially no suppression: across
10,786 Gemeinden only 109 / 61 / 8 cells are the true-zero dash, and nothing at all is
withheld.  The Gemeinde rows reconcile against the published national row to within the
disclosure noise.  Compare Estonia's base-10 rounding or New Zealand's confidentialised
cells — Germany hides nothing it has, it simply does not have much.

TWO TRAPS, both of them the playbook's (spec §12):

  * The download URL returns HTTP 200 with 71KB of HTML unless it carries
    `?__blob=publicationFile`.  §5a, in its natural habitat.
  * Counts are stored as text in some cells and as numbers in others, in one column of
    one sheet.  An `isinstance(v, (int, float))` filter silently drops 2.2 million
    people and every total still looks plausible.  Every cell goes through `cell()`.

Usage:
    python sources/de.py --fetch    download the xlsx (~840KB) if missing
    python sources/de.py            normalise from data/raw/de/
"""

import csv
import math
import os
import sys
import urllib.request

# Category names and Gemeinde names are German and the reconciliation prints them; a
# Windows console is cp1252 and dies on "oe with umlaut" at the PRINT, which looks like
# a data error (spec §12).
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
RAW = os.path.join(ROOT, "data", "raw", "de")
OUT = os.path.join(ROOT, "data", "normalized", "de.csv")

SOURCE_ID = "de_zensus_2022"
YEAR = 2022
# Melderegister church-tax records, not a census question.  See the module docstring.
BASIS = "roll"

XLSX_NAME = "religion_je_gemeinde.xlsx"
# The `?__blob=publicationFile` is load-bearing: without it destatis serves the landing
# PAGE with status 200, and the bytes are HTML that openpyxl rejects much later.
XLSX_URL = ("https://www.destatis.de/DE/Themen/Gesellschaft-Umwelt/Bevoelkerung/"
            "Zensus2022/Publikationen/Downloads-Publikationen/Sonderauswertungen/"
            "bevoelkerung_religionszugehoerigkeit_je_gemeinde.xlsx"
            "?__blob=publicationFile&v=3")
MIN_BYTES = 500_000

COLUMNS = ["geo_id", "geo_level", "geo_name", "source_category", "count",
           "basis", "year", "source_id", "note"]

SHEET = "Religion"
FIRST_DATA_ROW = 6          # 1-based; rows 1-5 are the two title lines and three header lines

# Column indexes into the sheet, 0-based.  Listed explicitly rather than strided,
# because every count column is followed by its percentage twin and taking the wrong
# one of each pair yields a map where every Gemeinde holds about 100 people and nothing
# else complains (spec §12, Croatia).
COL_AGS, COL_NAME, COL_LEVEL = 0, 1, 2
COL_POP = 3
CATEGORIES = [
    # (count column, share column, category name verbatim from the sheet header)
    (4, 5, "Römisch-katholische Kirche (öffentlich-rechtlich)"),
    (6, 7, "Evangelische Kirche (öffentlich-rechtlich)"),
    (8, 9, "Sonstige, keine, ohne Angabe"),
]
POP_CATEGORY = "Einwohnerzahl"
TOTAL_NOTE = "universe total, not a religion category"

DASH = "–"   # – = "Genau Null oder auf Null geändert"

# The published national row, from the "Bund" line of this same sheet.  Hard-coded so a
# reissue with different figures fails loudly instead of being normalised in silence.
NATIONAL = {
    "pop": 82_719_540,
    "Römisch-katholische Kirche (öffentlich-rechtlich)": 20_746_959,
    "Evangelische Kirche (öffentlich-rechtlich)": 19_127_360,
    "Sonstige, keine, ohne Angabe": 42_845_220,
}
N_GEMEINDEN = 10_786

# Deutsche im Ausland — Bundeswehr, police and diplomatic service posted abroad, with
# their families.  The sheet's own trailer says they are counted in the "Deutschland"
# row and in NO Gemeinde, so the Gemeinde sum is expected to fall short of the national
# row by roughly this much and the gap is a property of the file, not an error.
DEUTSCHE_IM_AUSLAND = 8_258

# Cell-Key method: the per-cell perturbation is bounded and small.  5 is a conservative
# bound rather than a published constant, and it is used only to turn "these two columns
# disagree" into "by how many people" — see the share check in check().  The observed
# worst case in the 2022 file is about 3 people, in a Gemeinde of 18.
CKM_MAX_DEVIATION = 5


def fetch():
    os.makedirs(RAW, exist_ok=True)
    dest = os.path.join(RAW, XLSX_NAME)
    if os.path.exists(dest) and os.path.getsize(dest) >= MIN_BYTES:
        print("already have", dest)
        return
    print("downloading", XLSX_URL)
    urllib.request.urlretrieve(XLSX_URL, dest)
    _validate(dest)


def _validate(path):
    """Assert size AND type AND content.  HTTP 200 is not a download (spec §12)."""
    size = os.path.getsize(path)
    if size < MIN_BYTES:
        raise SystemExit(f"{path} is {size:,} bytes, expected >= {MIN_BYTES:,} -- "
                         "destatis serves the landing PAGE with status 200 when the "
                         "URL is missing ?__blob=publicationFile")
    import zipfile
    if not zipfile.is_zipfile(path):
        raise SystemExit(f"{path} is not a zip, so not an xlsx -- most likely the HTML "
                         "landing page (spec §12, 'HTTP 200 is not a download')")
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    if SHEET not in sheets:
        raise SystemExit(f"{path} has no {SHEET!r} sheet, only {sheets}")
    print(f"  {size:,} bytes, sheets {sheets}")


def cell(v, where):
    """A count cell -> int, or None for the true-zero dash.

    Raises on anything unrecognised, so a new sentinel cannot appear silently
    (spec §12).  Handles the int/str split described in the module docstring: destatis
    stores some counts as numbers and some as text in the same column.
    """
    if isinstance(v, bool):
        raise ValueError(f"{where}: boolean count {v!r}")
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        if v != int(v):
            raise ValueError(f"{where}: non-integral count {v!r}")
        return int(v)
    if isinstance(v, str):
        s = v.strip().replace(" ", "").replace(" ", "")
        if s == DASH:
            return None                      # exactly zero, or changed to zero
        if s == ".":
            # "Zahlenwert unbekannt oder geheim".  Documented in the sheet's
            # Zeichenerklärung; not observed in the 2022 file.  If it ever appears the
            # row is not a partition any more and the caller must be told.
            raise ValueError(f"{where}: suppressed cell '.', which this file was not "
                             "expected to contain -- see Zeichenerklärung")
        if s.lstrip("-").isdigit():
            return int(s)
    raise ValueError(f"{where}: unrecognised count cell {v!r}")


def share(v):
    """A percentage cell -> float, or None.  '25,1%' and '(1,6%)' both occur.

    The parentheses mean "Aussagewert eingeschränkt" — the disclosure noise is large
    relative to the value — and are information about the count, not about the share.
    """
    if v is None:
        return None
    s = str(v).strip().strip("()").replace("%", "").replace(",", ".").strip()
    if not s or s == DASH:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def read():
    path = os.path.join(RAW, XLSX_NAME)
    if not os.path.exists(path):
        raise SystemExit(f"missing {path} -- run with --fetch first")
    _validate(path)

    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET]
    raw = list(ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True))
    wb.close()

    rows, flagged, dashes, trailer = [], 0, {}, 0
    for r in raw:
        level_de = r[COL_LEVEL]
        # Below the data the sheet carries its Zeichenerklärung and footnotes, as rows
        # with prose in the AGS column and nothing in Regionalebene.
        if level_de not in ("Bund", "Gemeinde"):
            trailer += 1
            continue
        level = "country" if level_de == "Bund" else "gemeinde"
        # The Regionalschlüssel is kept VERBATIM (spec §2.4); sources/de_geo.py derives
        # the 8-digit AGS from it for the boundary join.
        geo_id = str(r[COL_AGS]).strip()
        geo_name = str(r[COL_NAME]).strip()
        where = f"{geo_id} {geo_name}"

        pop = cell(r[COL_POP], where + " Einwohnerzahl")
        if pop is None:
            raise ValueError(f"{where}: population is the dash, which cannot happen -- "
                             "the sheet states the Einwohnerzahl is not perturbed")
        rows.append({
            "geo_id": geo_id, "geo_level": level, "geo_name": geo_name,
            "source_category": POP_CATEGORY, "count": pop,
            "basis": BASIS, "year": YEAR, "source_id": SOURCE_ID,
            "note": f"destatis Sonderauswertung {XLSX_NAME}; {TOTAL_NOTE}",
        })

        for col, pct_col, name in CATEGORIES:
            v = cell(r[col], f"{where} {name}")
            if v is None:
                dashes[name] = dashes.get(name, 0) + 1
                v = 0
            note = f"destatis Sonderauswertung {XLSX_NAME}"
            if isinstance(r[pct_col], str) and r[pct_col].strip().startswith("("):
                note += "; share parenthesised in source (Aussagewert eingeschränkt)"
                flagged += 1
            rows.append({
                "geo_id": geo_id, "geo_level": level, "geo_name": geo_name,
                "source_category": name, "count": v,
                "basis": BASIS, "year": YEAR, "source_id": SOURCE_ID,
                "note": note,
            })
    return rows, raw, flagged, dashes, trailer


def check(rows, raw, flagged, dashes, trailer):
    ok = True
    print(f"  skipped {trailer} trailer rows (Zeichenerklärung and footnotes)")

    gem = [r for r in rows if r["geo_level"] == "gemeinde"]
    nat = [r for r in rows if r["geo_level"] == "country"]

    units = {r["geo_id"] for r in gem}
    good = len(units) == N_GEMEINDEN
    ok &= good
    print(f"\n  {'OK ' if good else 'BAD'} {len(units):,} Gemeinden "
          f"(expected {N_GEMEINDEN:,})")

    # Every AGS is the 12-digit Regionalschlüssel.  Asserted rather than assumed, so a
    # reissue in a different format fails here instead of in the geo join (spec §12).
    bad_id = [u for u in units if not (len(u) == 12 and u.isdigit())]
    good = not bad_id
    ok &= good
    print(f"  {'OK ' if good else 'BAD'} every geo_id is 12 digits "
          f"({len(bad_id)} are not){'' if good else ': ' + str(bad_id[:5])}")

    # The published national row, against the figures this script was written for.
    nat_by_cat = {r["source_category"]: r["count"] for r in nat}
    for key, want in NATIONAL.items():
        cat = POP_CATEGORY if key == "pop" else key
        got = nat_by_cat.get(cat)
        good = got == want
        ok &= good
        print(f"  {'OK ' if good else 'BAD'} national {cat[:46]:<46} {got:>12,} "
              f"(expected {want:,})")

    # Gemeinde sums against the national row.  These are NOT expected to be equal: the
    # Deutsche im Ausland are in the national row and in no Gemeinde.
    def total(level, cat):
        return sum(r["count"] for r in rows
                   if r["geo_level"] == level and r["source_category"] == cat)

    gem_pop = total("gemeinde", POP_CATEGORY)
    gap = NATIONAL["pop"] - gem_pop
    good = gap == DEUTSCHE_IM_AUSLAND
    ok &= good
    print(f"\n  {'OK ' if good else 'BAD'} national pop − Gemeinde pop = {gap:,} "
          f"(expected {DEUTSCHE_IM_AUSLAND:,}, Deutsche im Ausland)")

    # Categories partition the population.  They cannot do so exactly: the Cell-Key
    # method perturbs each category cell but leaves the Einwohnerzahl untouched, and
    # the sheet says so ("Aus diesem Grund kann die Summe der Einzelergebnisse einer
    # Tabelle von der Einwohnerzahl abweichen").
    #
    # The band is computed from the method rather than chosen to pass (spec §12).  CKM
    # perturbations are per-cell, bounded, and approximately independent and unbiased,
    # so the sum over n cells grows like sqrt(n) and not like n.  A per-cell bound of 5
    # gives 5*sqrt(n); anything approaching 5n would mean the perturbations are not
    # independent, or that a column has been misread.
    n_cells = len(units) * len(CATEGORIES)
    band = 5 * math.sqrt(n_cells)
    cat_sum = sum(total("gemeinde", name) for _, _, name in CATEGORIES)
    diff = cat_sum - gem_pop
    good = abs(diff) <= band
    ok &= good
    print(f"  {'OK ' if good else 'BAD'} Gemeinde categories vs Gemeinde population: "
          f"{diff:+,} over {n_cells:,} cells (CKM band ±{band:,.0f})")

    nat_diff = sum(NATIONAL[n] for _, _, n in CATEGORIES) - NATIONAL["pop"]
    print(f"      the same check on the national row: {nat_diff:+,}")

    # THE COLUMN CHECK.  Every count column is followed by its percentage twin, so this
    # recomputes the share from the count and compares it with the published one.  If a
    # percentage column had been read as a count, the recomputed shares would be wrong
    # by tens of points in every large city and this fails on the first one; no other
    # check in this file would notice at all (spec §12, Croatia).
    #
    # It cannot be asserted as an equality, and a flat tolerance in percentage points
    # would be the wrong shape.  Two documented effects pull count/pop away from the
    # published share, and BOTH scale as people rather than as points:
    #
    #   * the Cell-Key method perturbs each category cell by a bounded number of PEOPLE
    #     while leaving the Einwohnerzahl untouched, so in a Gemeinde of 18 a shift of
    #     two people is eleven percentage points;
    #   * where the perturbed count would give an implausible share, destatis adjusts
    #     the SHARE and leaves the count — "Um solche offensichtlichen Unplausibilitäten
    #     zu verhindern, nimmt das Geheimhaltungsverfahren ... eine Anpassung des
    #     Anteils vor".  Ammeldingen an der Our has population 18 and 20 Catholics, and
    #     is published as 100.0%.
    #
    # So the discrepancy is converted BACK into people and bounded there, which is the
    # scale the method actually works on.  The published share carries one decimal, so
    # 0.05pp of rounding is allowed for before converting.  The implied deviation is
    # printed as a number, not just a verdict: if destatis changes the method, this
    # moves visibly instead of flipping a flag.
    worst_people, worst_where, n_cmp = 0.0, None, 0
    for r in raw:
        if r[COL_LEVEL] != "Gemeinde":
            continue
        pop = cell(r[COL_POP], "share check")
        if not pop:
            continue
        for col, pct_col, name in CATEGORIES:
            want = share(r[pct_col])
            v = cell(r[col], "share check")
            if want is None or v is None:
                continue
            n_cmp += 1
            people = max(0.0, abs(100.0 * v / pop - want) - 0.05) * pop / 100.0
            if people > worst_people:
                worst_people, worst_where = people, (str(r[1]), name, pop, v, want)
    good = worst_people <= CKM_MAX_DEVIATION
    ok &= good
    print(f"\n  {'OK ' if good else 'BAD'} count/population reproduces the published "
          f"share in {n_cmp:,} cells")
    print(f"      worst disagreement is {worst_people:.2f} people "
          f"(CKM per-cell bound {CKM_MAX_DEVIATION})")
    if worst_where:
        nm, cat, pop, v, want = worst_where
        print(f"      at {nm}: population {pop:,}, count {v:,}, published {want}% "
              f"— {cat[:40]}")

    print(f"\n  '{DASH}' true-zero cells: " + (str(dashes) if dashes else "none"))
    print(f"  cells whose share the source parenthesised as uncertain: {flagged:,} "
          f"({100.0 * flagged / n_cells:.1f}% of {n_cells:,})")

    print(f"\n  {len(rows):,} rows, {len(CATEGORIES)} categories + the universe total")
    print("\n  national totals:")
    for _, _, name in CATEGORIES:
        n = total("gemeinde", name)
        print(f"    {n:>12,}  {100.0 * n / gem_pop:5.1f}%  {name}")

    if not ok:
        raise SystemExit("reconciliation FAILED")


def main():
    if "--fetch" in sys.argv:
        fetch()
    rows, raw, flagged, dashes, trailer = read()
    check(rows, raw, flagged, dashes, trailer)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)
    print("\nwrote", OUT)


if __name__ == "__main__":
    main()
