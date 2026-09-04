"""Romania — INS, Recensământul Populaţiei şi Locuinţelor 2021, religion, down to UAT.

Reads (or fetches) data/raw/ro/ and writes data/normalized/ro.csv.

23 religions at the level of the **UAT** — municipiu, oraş, comună — 3,175 of them for
19.05M people, which is about 6,000 people a unit and comparable to a Polish gmina. That
is a good source, and it carries the only large Old Believer population in Europe that a
census names: `Crestina de Rit Vechi`, the Lipovans of the Danube delta, 28,362 people.

Three things make it harder than Poland, and all three are in the data rather than the
download:

  1. **`*` IS AN IN-BAND SUPPRESSION MARKER**, not a footnote. INS writes `*` where a cell
     is confidential and `-` where it is a true zero, both in the same numeric columns.
     Read naively with pandas the column becomes object dtype and every `*` silently
     becomes a string; summed as delivered it raises, and coerced with `errors="coerce"`
     it becomes NaN and the suppressed people vanish without a word. §3.2's shape, and the
     second time the project has met it after New Zealand's `-999`.
  2. **There are no SIRUTA codes.** Rows are identified by NAME only, so the join to
     boundaries is by (judeţ, name) and needs the Eurostat correspondence table. See
     `sources/ro_geo.md`.
  3. **The county rows are bare names, exactly like the communes.** `ALBA` is a county
     header and `Alba Iulia` is not, and nothing in the row says which. The county set
     comes from sheet 2.4.1, and a name is a county header the FIRST time it appears —
     which matters because Bucharest appears twice in a row, once as its own county and
     once as its own UAT.

THE DOMINANT CAVEAT: `Informatie nedisponibila` is **2,658,165 people, 13.95%** of the
country. INS collected the 2021 census largely from administrative registers, and religion
is not in them, so this is not a refusal like Poland's — it is a variable that could not be
established for one person in seven. It is excluded from the dots.

Usage:
    python sources/ro.py --fetch    download the workbook (0.5MB) if missing
    python sources/ro.py            normalise from data/raw/ro/
"""

import csv
import os
import re
import sys
import unicodedata

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
RAW = os.path.join(ROOT, "data", "raw", "ro")
OUT = os.path.join(ROOT, "data", "normalized", "ro.csv")

SOURCE_ID = "ro_rpl_2021"
YEAR = 2021
BASIS = "self_id"

XLSX_NAME = "Tabel-2.04.1-si-Tabel-2.04.2.xlsx"
XLSX_URL = ("https://www.recensamantromania.ro/wp-content/uploads/2023/06/"
            "Tabel-2.04.1-si-Tabel-2.04.2.xlsx")
MIN_BYTES = 300_000

COLUMNS = ["geo_id", "geo_level", "geo_name", "source_category", "count",
           "basis", "year", "source_id", "note"]

NATIONAL = 19_053_815        # INS, resident population, 1 December 2021

SHEET_JUDET = "Tab 2.4.1"
SHEET_UAT = "Tab 2.4.2"

TOTAL_COL = 1
FIRST_CAT, LAST_CAT = 2, 24      # inclusive; 23 religion columns

TOTAL_LABEL = "POPULATIA REZIDENTA TOTAL"
TOTAL_NOTE = "universe total, not a religion category"

SUPPRESSED = "*"
TRUE_ZERO = "-"

# The eight development regions and the four macroregions sit between ROMÂNIA and the
# counties on sheet 2.4.1 and are aggregates of them, so they are not counties and are
# not levels this file emits.
NOT_COUNTIES = {
    "romania", "macroregiunea 1", "macroregiunea 2", "macroregiunea 3", "macroregiunea 4",
    "nord-vest", "centru", "nord-est", "sud-est", "sud-muntenia", "bucuresti-ilfov",
    "sud-vest oltenia", "vest",
}


def fold(s):
    """Fold a Romanian name for comparison.

    ş/ș and ţ/ț are DIFFERENT codepoints — cedilla vs comma-below — and both are in use,
    sometimes inside one file. INS writes the census names with comma-below and Eurostat
    writes the same names with cedilla, so any name join that does not fold these matches
    almost nothing and looks like a vintage problem instead of an encoding one.
    """
    s = " ".join(str(s).split())
    s = (s.replace("ş", "s").replace("Ş", "S")
          .replace("ţ", "t").replace("Ţ", "T"))
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.casefold().strip()


def fetch():
    import requests
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    os.makedirs(RAW, exist_ok=True)
    dest = os.path.join(RAW, XLSX_NAME)
    if os.path.exists(dest) and os.path.getsize(dest) >= MIN_BYTES:
        print("already have", dest)
        return
    print("downloading", XLSX_URL)
    r = requests.get(XLSX_URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=300,
                     verify=False)
    r.raise_for_status()
    with open(dest, "wb") as fh:
        fh.write(r.content)
    size = os.path.getsize(dest)
    if size < MIN_BYTES:
        raise SystemExit(f"{dest} is {size:,} bytes, expected >= {MIN_BYTES:,}")
    print(f"  {size:,} bytes")


def _txt(v):
    return "" if v is None else " ".join(str(v).split())


def _value(cell):
    """Return (count, kind) where kind is 'n', 'suppressed' or 'zero'."""
    if cell is None:
        return None, "blank"
    if isinstance(cell, (int, float)):
        return int(cell), "n"
    s = str(cell).strip()
    if s == SUPPRESSED:
        return None, "suppressed"
    if s in (TRUE_ZERO, "–", "—"):
        return 0, "zero"
    if s == "":
        return None, "blank"
    # Anything else is a format change and must not be guessed at.
    raise SystemExit(f"unexpected cell value {cell!r} in a count column -- INS changed "
                     "the sheet, check for a new sentinel")


def read():
    import openpyxl

    path = os.path.join(RAW, XLSX_NAME)
    if not os.path.exists(path):
        raise SystemExit(f"missing {path} -- run with --fetch first")
    if os.path.getsize(path) < MIN_BYTES:
        raise SystemExit(f"{path} is truncated, re-run with --fetch")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for s in (SHEET_JUDET, SHEET_UAT):
        if s not in wb.sheetnames:
            raise SystemExit(f"missing sheet {s!r}; got {wb.sheetnames}")

    # ---- category names, from the header of the UAT sheet
    ws = wb[SHEET_UAT]
    header = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    cats = {c: _txt(header[c]) for c in range(FIRST_CAT, LAST_CAT + 1)}
    missing = [c for c, n in cats.items() if not n]
    if missing:
        raise SystemExit(f"blank category header in columns {missing}")

    # ---- the county set AND each county's own total, from the judet sheet.
    # The total is what makes the county headers findable on the UAT sheet; see below.
    counties = {}
    for r in wb[SHEET_JUDET].iter_rows(min_row=7, values_only=True):
        name = _txt(r[0])
        if not name or not isinstance(r[TOTAL_COL], (int, float)):
            continue
        if fold(name) in NOT_COUNTIES:
            continue
        counties[fold(name)] = (name, int(r[TOTAL_COL]))
    if len(counties) != 42:
        raise SystemExit(f"found {len(counties)} counties, expected 42: "
                         f"{sorted(counties)}")

    rows, stats = [], {"suppressed": 0, "zero": 0}

    def emit(gid, level, gname, cat, n, note):
        rows.append({"geo_id": gid, "geo_level": level, "geo_name": gname,
                     "source_category": cat, "count": n, "basis": BASIS,
                     "year": YEAR, "source_id": SOURCE_ID, "note": note})

    def emit_unit(gid, level, gname, r, note):
        tot, kind = _value(r[TOTAL_COL])
        if kind != "n":
            raise SystemExit(f"{gname}: total is not a number ({kind})")
        emit(gid, level, gname, TOTAL_LABEL, tot, note + "; " + TOTAL_NOTE)
        for c in range(FIRST_CAT, LAST_CAT + 1):
            n, kind = _value(r[c])
            stats[kind] = stats.get(kind, 0) + 1
            if kind == "n" and n > 0:
                emit(gid, level, gname, cats[c], n, note)

    # ---- national + counties, from the judet sheet
    for r in wb[SHEET_JUDET].iter_rows(min_row=7, values_only=True):
        name = _txt(r[0])
        if not name or not isinstance(r[TOTAL_COL], (int, float)):
            continue
        f = fold(name)
        if f == "romania":
            emit_unit("RO", "country", "România", r, f"INS RPL 2021 {SHEET_JUDET}")
        elif f in counties:
            emit_unit(name, "judet", name, r, f"INS RPL 2021 {SHEET_JUDET}")

    # ---- UATs, from the UAT sheet
    #
    # WHICH ROWS ARE COUNTY HEADERS. Counties and communes are both bare names and
    # nothing in the row says which it is, so the header has to be identified by its
    # NUMBER: a row is a county header when its name is a county's AND its total equals
    # that county's total on sheet 2.4.1.
    #
    # "First row with a county's name" is the obvious rule and it is WRONG, because two
    # county names are also commune names in other counties, and the communes sort first:
    #
    #   CĂLĂRAȘI    3,285 (a commune in Dolj) ... 283,458 (the county) ... two more communes
    #   SATU MARE   1,995 (a commune in Harghita) ... 330,668 (the county) ... one more
    #
    # Taking the commune as the header moves every following county boundary and produces
    # two "PRAHOVA|PĂULEȘTI" rows out of two different counties' Păulești — which the
    # reconciliation catches as categories exceeding the unit total, but only because the
    # two happened to collide. Six rows were misfiled and 600,861 people double-counted.
    #
    # Bucharest still needs the `seen` guard: it appears on two consecutive rows with
    # identical totals, once as its own county and once as the single UAT that fills it,
    # so both rows match the total test and only the first is the header.
    seen = set()
    current = None
    n_uat = 0
    for r in wb[SHEET_UAT].iter_rows(min_row=7, values_only=True):
        name = _txt(r[0])
        if not name or not isinstance(r[TOTAL_COL], (int, float)):
            continue
        f = fold(name)
        if f == "romania":
            continue
        if f in counties and f not in seen and int(r[TOTAL_COL]) == counties[f][1]:
            seen.add(f)
            current = counties[f][0]
            continue
        if current is None:
            raise SystemExit(f"UAT row {name!r} before any county header")
        n_uat += 1
        # geo_id is (county, name) because there is no code; ro_geo.py resolves it to a
        # SIRUTA code. The separator is one that cannot occur in a Romanian place name.
        emit_unit(f"{current}|{name}", "uat", name, r, f"INS RPL 2021 {SHEET_UAT}")

    if len(seen) != 42:
        raise SystemExit(f"only {len(seen)} of 42 counties found on {SHEET_UAT}")
    print(f"  parsed {n_uat:,} UAT rows in {len(seen)} counties")
    print(f"  cells: {stats.get('n', 0):,} numeric, "
          f"{stats['suppressed']:,} suppressed (*), {stats['zero']:,} true zero (-)")
    return rows


def check(rows):
    ok = True
    levels = {}
    totals, cats_sum = {}, {}
    for r in rows:
        lv, gid = r["geo_level"], r["geo_id"]
        levels.setdefault(lv, set()).add(gid)
        if r["source_category"] == TOTAL_LABEL:
            totals[(lv, gid)] = r["count"]
        else:
            cats_sum[(lv, gid)] = cats_sum.get((lv, gid), 0) + r["count"]

    expected = {"country": 1, "judet": 42, "uat": 3181}
    print("\n  units and population per level — levels are alternatives, never summed:")
    for lv in ("uat", "judet", "country"):
        got_units = len(levels.get(lv, ()))
        got_pop = sum(n for (l, _), n in totals.items() if l == lv)
        good = got_pop == NATIONAL and got_units == expected[lv]
        ok &= good
        print(f"    {'OK ' if good else 'BAD'} {lv:<8} {got_units:>6,} units  "
              f"{got_pop:>12,}  (expected {expected[lv]:,} units, {NATIONAL:,})")

    # The categories partition the total EXCEPT where INS suppressed a cell, so this is
    # a bound rather than an equality: they may never exceed it, and the shortfall is the
    # suppressed people. Reported rather than asserted.
    over = [(k, cats_sum[k], totals.get(k, 0)) for k in cats_sum
            if cats_sum[k] > totals.get(k, 0)]
    good = not over
    ok &= good
    print(f"\n  {'OK ' if good else 'BAD'} categories never exceed the unit total "
          f"({len(over)} violations)")
    for k, c, t in over[:5]:
        print(f"      {k}: categories {c:,} vs total {t:,}")

    for lv in ("uat", "judet", "country"):
        got = sum(n for (l, _), n in cats_sum.items() if l == lv)
        tot = sum(n for (l, _), n in totals.items() if l == lv)
        print(f"    {lv:<8} categories sum to {got:>12,} of {tot:,}  "
              f"({100.0 * got / tot:6.3f}%) — shortfall is INS suppression")

    by_cat = {}
    for r in rows:
        if r["geo_level"] == "country" and r["source_category"] != TOTAL_LABEL:
            by_cat[r["source_category"]] = r["count"]
    print(f"\n  {len(rows):,} rows, {len(by_cat)} categories:")
    for name, n in sorted(by_cat.items(), key=lambda x: -x[1]):
        print(f"    {n:>10,}  {name}")

    if not ok:
        raise SystemExit("reconciliation FAILED")


def main():
    if "--fetch" in sys.argv:
        fetch()
    rows = read()
    check(rows)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)
    print("\nwrote", OUT)


if __name__ == "__main__":
    main()
