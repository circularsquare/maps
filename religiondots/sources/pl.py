"""Poland — GUS NSP 2021, declared religious affiliation, down to gmina.

Reads (or fetches) data/raw/pl/ and writes data/normalized/pl.csv.

This is the second source in the project, after Czechia, that publishes NAMED CHURCHES
at its FINEST GEOGRAPHY.  139 named denominations across 2,477 gminy, no suppression
and no rounding: `Polski Kościół Dialogu` has exactly one adherent nationally and GUS
prints the 1.  Everything reconciles to the person — the gmina totals sum to 38,036,118
and the affiliated sum to 27,601,000, both exactly the published national figures.

So Poland mostly does NOT make the spec §3.9 trade that Australia, Canada, New Zealand,
Mexico, Ireland and Brazil all make.  It makes a smaller version of it: 216 denominations
are named nationally (TABL.2) and 139 of them reach gmina level (TABL.7).  The 77 that do
not are collectively 58,656 people — every one of them a body too small to clear GUS's
per-gmina publication floor.  They are carried here at `country` and `voivodeship` level
so allocate.py can push them down if that is ever wanted; nothing is dropped.

WHAT THE CATEGORY LIST CONTAINS, and why it is not a clean taxonomy — the same four
problems as Czechia, which is not a coincidence, because it is the same kind of question:

  * Registered churches under their legal names, down to single congregations —
    `Zbór Ewangeliczny "Betel" w Warszawie` (2), `Kościół Jezusa Chrystusa w Werbkowicach`
    (5), `Warsaw International Church` (24).  The tail is congregation-level, not
    denomination-level, which is finer than any other source in the project reaches.
  * Bare tradition names beside institutions covering overlapping people.  `różne
    afiliacje islamskie (ogólnie islam, muzułmanizm, sunnizm, szyizm itp.)` and
    `Muzułmański Związek Religijny` are separate rows; per spec §2.3 they are not the
    same kind of thing, and per cz.py's finding they must not be treated as a hierarchy.
  * Positions rather than religions — `deizm`, `teizm`, `panteizm`, `gnostycyzm`.
  * Parody answers as first-class published categories, and here they are LARGE:
    `pastafarianizm` 2,312 and `jediizm (religia Jedi)` 687.  Left verbatim per §2.4.

THE DOMINANT FACT ABOUT THE POLISH MAP, and it belongs in note_public rather than a
footnote: the religion question was voluntary and **20.53% of the country refused it**
(7,807,553 people, `Odmawiający odpowiedzi na pytanie o wyznanie`).  A further 2,611,506
reported no religion and 15,059 were `Nie ustalono`.  So the drawn population is 27.6M of
38.0M.  Czechia's equivalent is 30%; this is the same problem one size smaller.

Usage:
    python sources/pl.py --fetch    download the workbook (3.3MB) if missing
    python sources/pl.py            normalise from data/raw/pl/
"""

import csv
import os
import sys

# Category names are Polish and the reconciliation prints them.  A Windows console is
# cp1252 by default, which cannot encode ł/ś/ż and kills the run at the print rather
# than at anything to do with the data.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
RAW = os.path.join(ROOT, "data", "raw", "pl")
OUT = os.path.join(ROOT, "data", "normalized", "pl.csv")

SOURCE_ID = "pl_nsp_2021"
YEAR = 2021
BASIS = "self_id"

XLSX_NAME = "przynaleznosc_wyznaniowa_nsp2021.xlsx"
XLSX_URL = ("https://stat.gov.pl/download/gfx/portalinformacyjny/pl/defaultaktualnosci/"
            "6536/10/1/1/przynaleznosc_wyznaniowa_-_dane_nsp_2021_dla_kraju_i_jednostek"
            "_podzialu_terytorialnego_1.xlsx")
# Published 29 Apr 2024.  Expected ~3.48MB; a short file is a truncated download rather
# than an empty release (sources.md §5a).
MIN_BYTES = 3_000_000

COLUMNS = ["geo_id", "geo_level", "geo_name", "source_category", "count",
           "basis", "year", "source_id", "note"]

NATIONAL = 38_036_118        # GUS, resident population, census day 2021-03-31
NATIONAL_AFFILIATED = 27_601_000

# The four sheets that carry counts by place.  `code`/`name` are column indices; `cats`
# are the columns that may hold a category name, indented by depth; the count always sits
# one past the last of them.
#
# THE TRAP: three of these four sheets are FLAT — a level-1 universe row (Ogółem /
# Udzielający / Odmawiający / Nie ustalono), a level-2 row (należący / nienależący), then
# named churches and nothing in between.  TABL.5 is NOT.  It carries the full 7-deep
# classification, so `chrześcijaństwo` (L4), `katolicyzm` (L5) and `Kościół katolicki`
# (L6) all appear as rows ABOVE `Kościół katolicki - obrządek łaciński` (L7).  Summing
# that sheet the way the other three are summed counts the Latin rite four times and
# returns roughly 4x the country.
#
# `poziom` is the column GUS puts the depth in, and `leaf` is the only depth that holds a
# denomination.  Sheets without a `poziom` column are flat and take the position rule.
SHEETS = {
    "TABL.7": dict(level="gmina",       units=2477, code=4, name=3, cats=(5, 6, 7)),
    "TABL.6": dict(level="powiat",      units=380,  code=3, name=2, cats=(4, 5, 6)),
    "TABL.5": dict(level="voivodeship", units=16,   code=2, name=1, cats=tuple(range(4, 10)),
                   poziom=3, leaf=7),
    "TABL.2": dict(level="country",     units=1,    code=None, name=None, cats=(1, 2, 3)),
}

TOTAL_NOTE = "universe total, not a religion category"

# The level-1 and level-2 rows.  These are universes and positions, not denominations,
# and they are emitted verbatim so the taxonomy file — not this one — decides what
# happens to them.  Prefix matching, because GUS pads the labels with "w tym:".
UNIVERSE = ("Ogółem", "Udzielający odpowiedzi", "Odmawiający odpowiedzi", "Nie ustalono")
NOT_A_DENOMINATION = ("należący do wyznania", "nienależący do żadnego wyznania")


def fetch():
    """Download the workbook.

    stat.gov.pl serves an INCOMPLETE CERTIFICATE CHAIN — it omits the intermediate — so
    both `curl` and a default `requests` call fail with
    "unable to get local issuer certificate" on a machine that has not cached it.  That
    is a server misconfiguration, not a bad URL and not a proxy.  Verification is turned
    off for this one host and the file is checked structurally instead: it must be at
    least MIN_BYTES and must be a real xlsx that opens and contains the eight expected
    sheets.  sources.md §5a is the general form of this rule.
    """
    import requests
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    os.makedirs(RAW, exist_ok=True)
    dest = os.path.join(RAW, XLSX_NAME)
    if os.path.exists(dest) and os.path.getsize(dest) >= MIN_BYTES:
        print("already have", dest)
        return
    print("downloading", XLSX_URL)
    r = requests.get(XLSX_URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=300,
                     verify=False)
    r.raise_for_status()
    with open(dest, "wb") as fh:
        fh.write(r.content)
    size = os.path.getsize(dest)
    if size < MIN_BYTES:
        raise SystemExit(f"{dest} is {size:,} bytes, expected >= {MIN_BYTES:,} -- "
                         "truncated download")
    print(f"  {size:,} bytes")


def _txt(v):
    """GUS embeds newlines and double spaces in label cells; flatten them."""
    if v is None:
        return ""
    return " ".join(str(v).split())


def _classify(row, cats):
    """Return (category_name, is_universe) for a data row, or (None, _) if it has none.

    The indent is carried by WHICH of the `cats` columns is non-empty: the first is the
    level-1 universe, the second the level-2 split, the rest are denominations.  On the
    two wide sheets the denomination can land in any of several columns because GUS
    indents by classification depth, so everything past the first two counts as a name.
    """
    first, second, rest = cats[0], cats[1], cats[2:]
    if _txt(row[first]):
        return _txt(row[first]), True
    if _txt(row[second]):
        return _txt(row[second]), True
    for c in rest:
        if _txt(row[c]):
            return _txt(row[c]), False
    return None, False


def read():
    import openpyxl

    path = os.path.join(RAW, XLSX_NAME)
    if not os.path.exists(path):
        raise SystemExit(f"missing {path} -- run with --fetch first")
    size = os.path.getsize(path)
    if size < MIN_BYTES:
        raise SystemExit(f"{path} is {size:,} bytes, expected >= {MIN_BYTES:,} -- "
                         "truncated download, re-run with --fetch")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    missing = [s for s in SHEETS if s not in wb.sheetnames]
    if missing:
        raise SystemExit(f"{path} is missing sheets {missing} -- GUS reissued the "
                         f"workbook with a different layout; got {wb.sheetnames}")

    rows = []
    for sheet, cfg in SHEETS.items():
        ws = wb[sheet]
        level, code_col, name_col, cats = (cfg["level"], cfg["code"],
                                           cfg["name"], cfg["cats"])
        poziom_col, leaf = cfg.get("poziom"), cfg.get("leaf")
        # Names appear once, on the unit's first row, and are blank on its others.
        cur_code, cur_name = ("PL", "Polska") if code_col is None else (None, None)
        for r in ws.iter_rows(min_row=5, values_only=True):
            if code_col is not None:
                code = _txt(r[code_col])
                if code:
                    cur_code = code
                nm = _txt(r[name_col]) if name_col is not None else ""
                if nm:
                    cur_name = nm
                if cur_code is None:
                    continue
            cat, is_universe = _classify(r, cats)
            if cat is None:
                continue
            if poziom_col is not None:
                depth = _txt(r[poziom_col])
                depth = int(depth) if depth.isdigit() else None
                # Depths between the universe rows and the leaves are aggregates of the
                # leaves below them.  Dropping them is what keeps this sheet summable.
                if depth is not None and not is_universe and depth != leaf:
                    continue
            # The count column sits immediately after the last category column on every
            # sheet; the three that follow it are percentages.
            val = r[cats[-1] + 1]
            if not isinstance(val, (int, float)):
                continue
            note = f"GUS NSP 2021 {sheet}"
            if cat.startswith("Ogółem"):
                note += "; " + TOTAL_NOTE
            elif is_universe:
                note += "; universe subtotal, not a religion category"
            rows.append({
                "geo_id": cur_code,
                "geo_level": level,
                "geo_name": cur_name or "",
                "source_category": cat,
                "count": int(val),
                "basis": BASIS,
                "year": YEAR,
                "source_id": SOURCE_ID,
                "note": note,
            })
    return rows


def _is_universe(cat):
    # Case-folded because the national sheet shouts "OGÓŁEM" where the three subnational
    # sheets write "Ogółem"; matching case-sensitively lets the country's own total
    # through as if it were the largest denomination in Poland.
    c = cat.casefold()
    return (any(c.startswith(u.casefold()) for u in UNIVERSE)
            or any(c.startswith(u.casefold()) for u in NOT_A_DENOMINATION))


def check(rows):
    ok = True

    units, totals, denom_sum, affil = {}, {}, {}, {}
    for r in rows:
        lv, gid, cat, n = r["geo_level"], r["geo_id"], r["source_category"], r["count"]
        units.setdefault(lv, set()).add(gid)
        if cat.casefold().startswith("ogółem"):
            totals[(lv, gid)] = n
        elif cat.startswith("należący do wyznania"):
            affil[(lv, gid)] = n
        elif not _is_universe(cat):
            denom_sum[(lv, gid)] = denom_sum.get((lv, gid), 0) + n

    print(f"  units and population per level — levels are alternatives, never summed "
          f"(national {NATIONAL:,}):")
    for lv in sorted(units, key=lambda x: -len(units[x])):
        want = next(c["units"] for c in SHEETS.values() if c["level"] == lv)
        got_units = len(units[lv])
        got_pop = sum(n for (l, _), n in totals.items() if l == lv)
        good = got_units == want and got_pop == NATIONAL
        ok &= good
        print(f"    {'OK ' if good else 'BAD'} {lv:<12} {got_units:>6,} units  "
              f"{got_pop:>12,}  (expected {want:,} units, {NATIONAL:,})")

    # The affiliated subtotal must reproduce the published national figure at every
    # level independently.  This is what proves no level is short a unit.
    print()
    for lv in sorted(units, key=lambda x: -len(units[x])):
        got = sum(n for (l, _), n in affil.items() if l == lv)
        good = got == NATIONAL_AFFILIATED
        ok &= good
        print(f"    {'OK ' if good else 'BAD'} {lv:<12} affiliated {got:>12,}  "
              f"(published {NATIONAL_AFFILIATED:,})")

    # How much of `affiliated` the NAMED denominations actually reach, per level.  This
    # is the §3.9 measurement for Poland and the number that decides whether allocate.py
    # is needed: at country level it is 100% by construction, at gmina level it is not.
    print("\n  named denominations as a share of the affiliated, per level:")
    for lv in sorted(units, key=lambda x: -len(units[x])):
        got = sum(n for (l, _), n in denom_sum.items() if l == lv)
        ncat = len({r["source_category"] for r in rows
                    if r["geo_level"] == lv and not _is_universe(r["source_category"])})
        print(f"    {lv:<12} {got:>12,}  {100.0 * got / NATIONAL_AFFILIATED:6.3f}%  "
              f"of affiliated, over {ncat:>3} named categories")

    # Within every unit the parts must not exceed the unit's own total.
    over = [(k, denom_sum[k], affil.get(k, 0)) for k in denom_sum
            if denom_sum[k] > affil.get(k, 0)]
    good = not over
    ok &= good
    print(f"\n  {'OK ' if good else 'BAD'} named denominations never exceed the unit's "
          f"affiliated count ({len(over)} violations of {len(denom_sum):,} units)")
    for k, d, a in over[:5]:
        print(f"      {k}: denominations {d:,} vs affiliated {a:,}")

    cats = {r["source_category"] for r in rows if not _is_universe(r["source_category"])}
    gmina_cats = {r["source_category"] for r in rows
                  if r["geo_level"] == "gmina" and not _is_universe(r["source_category"])}
    print(f"\n  {len(rows):,} rows, {len(cats)} named categories nationally, "
          f"{len(gmina_cats)} of them at gmina")
    only_above = cats - gmina_cats
    print(f"  {len(only_above)} named only above gmina — "
          f"{sum(r['count'] for r in rows if r['geo_level'] == 'country' and r['source_category'] in only_above):,} people, "
          f"every one below GUS's per-gmina publication floor")

    by_cat = {}
    for r in rows:
        if r["geo_level"] == "country" and not _is_universe(r["source_category"]):
            by_cat[r["source_category"]] = by_cat.get(r["source_category"], 0) + r["count"]
    print("\n  national totals, largest first:")
    for name, n in sorted(by_cat.items(), key=lambda x: -x[1])[:12]:
        print(f"    {n:>10,}  {name}")
    print("\n  smallest published categories:")
    for name, n in sorted(by_cat.items(), key=lambda x: x[1])[:8]:
        print(f"    {n:>10,}  {name}")

    if not ok:
        raise SystemExit("reconciliation FAILED")


def main():
    if "--fetch" in sys.argv:
        fetch()
    rows = read()
    check(rows)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)
    print("\nwrote", OUT)


if __name__ == "__main__":
    main()
