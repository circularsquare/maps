"""Philippines — PSA 2020 Census of Population and Housing, religious affiliation.

Reads   data/raw/ph/3_Statistical_Table_for_Religious_Affiliation_RML_12082022_PMMJ_CRD_1.xlsx
        data/raw/ph/PSGC-1Q-2022-Publication-Datafile.xlsx
Writes  data/normalized/ph.csv

The census table (sheet "A") is a wide matrix: one row per geographic unit, one column per
religious affiliation, cells are persons in the *household* population. It carries no codes at
all, only names, so every row label is resolved against the PSGC publication datafile of
31 March 2022 — the exact edition the 2020 CPH technical notes say was used to disaggregate the
census. See sources/ph.md for why the vintage matters here.

Nothing is mapped to the taxonomy: source_category is the PSA's own column heading, verbatim
(spec.md 2.4, cross-source denomination matching is deferred).

Run:  python sources/ph.py
"""

import csv
import os
import re
import sys
import unicodedata

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
RAW = os.path.join(ROOT, "data", "raw", "ph")
OUT_DIR = os.path.join(ROOT, "data", "normalized")
OUT = os.path.join(OUT_DIR, "ph.csv")

CENSUS_XLSX = os.path.join(
    RAW, "3_Statistical_Table_for_Religious_Affiliation_RML_12082022_PMMJ_CRD_1.xlsx"
)
PSGC_XLSX = os.path.join(RAW, "PSGC-1Q-2022-Publication-Datafile.xlsx")

SOURCE_ID = "ph_cph_2020"
YEAR = 2020
BASIS = "self_id"

# The census table's own headline, from the press release, used as the reconciliation target.
PUBLISHED = {
    "household_population": 108667043,
    # headline: 78.8% of the household population, footnoted as excluding Catholic Charismatics
    "roman_catholic_excl_charismatic": 85645362,
    "catholic_charismatic": 74096,
    "islam": 6981710,
    "iglesia_ni_cristo": 2806524,
    "seventh_day_adventist": 862725,
    "aglipay": 818916,
    "iglesia_filipina_independiente": 640076,
    "bible_baptist_church": 540364,
    "uccp": 470792,
    "jehovahs_witness": 457245,
    "church_of_christ": 429921,
    "none": 43931,
    "not_reported": 15186,
    "other_religious_affiliations": 8954291,
    # PSGC 1Q 2022 national summary, "Nat'l Sum" sheet: total population, 2020 CPH.
    "total_population": 109035343,
}

# Row labels in the census table that do not match a PSGC name character for character.
# Left: the census label after normalisation. Right: the PSGC 10-digit code, chosen by hand.
LABEL_OVERRIDES = {
    # parenthetical old/alternate names the census table carries and the PSGC does not
    "samar (western samar)": "0806000000",
    "cotabato (north cotabato)": "1204700000",
    "davao de oro (compostela valley)": "1108200000",
    "city of general santos (dadiangas)": "1230800000",
    # the census table states the HUC treatment inside the label
    "basilan (excluding the city of isabela)": "1900700000",
    # Maguindanao here INCLUDES Cotabato City, which is exactly what PSGC province
    # 1903800000 contains as a code hierarchy (the PSGC's own population note excludes it).
    "maguindanao (including the city of cotabato)": "1903800000",
    # the 63 Cotabato barangays that voted into BARMM; PSGC calls this an "Interim Province Code"
    "interim province": "1999900000",
    "municipality of pateros": "1381701000",
    # the only region whose census label differs from its PSGC name
    "national capital region": "1300000000",
}

# Geographic level per PSGC "Geographic Level" value, plus the two special province-slot codes.
LEVEL_OF_PSGC_LEVEL = {
    "Reg": "region",
    "Prov": "province",
    "City": "city",
    "Mun": "municipality",
}

# Rows that are aggregates of other rows in the same file. Flagged so a consumer cannot sum the
# file blindly. The finest tier that partitions the country is province + city + municipality.
AGGREGATE_NOTE = {
    "country": "aggregate of the 17 region rows — do not add to them",
    "region": "aggregate of its province/city/municipality rows — do not add to them",
}

# Per-unit warnings, all of them about the 2020 CPH's geography rather than its numbers.
EXTRA_NOTES = {
    "1999900000": (
        "PSGC interim province code for the 63 Cotabato barangays that voted into BARMM "
        "(Special Geographic Area); no single polygon exists for it — PSGC 1Q 2022 splits it "
        "into 8 SGU clusters, 1999901000-1999908000"
    ),
    "1903800000": (
        "Maguindanao as tabulated here INCLUDES the City of Cotabato; undivided, i.e. before "
        "the 2022 split into Maguindanao del Norte and del Sur"
    ),
    "1204700000": (
        "Cotabato province excluding the 63 Special Geographic Area barangays, which are in "
        "the BARMM interim province row 1999900000"
    ),
    "1900700000": "Basilan excluding the City of Isabela, which is tabulated under Region IX",
    "0990101000": (
        "City of Isabela is geographically in Basilan but administratively part of Region IX; "
        "PSGC gives it the special province-slot code 0990100000"
    ),
}


def norm(s):
    """Fold a place name to a comparison key."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFC", str(s))
    s = s.replace("’", "'").replace("‘", "'")
    s = re.sub(r"\s+", " ", s).strip().lower()
    s = re.sub(r"\s*\(capital\)\s*$", "", s)
    return s.strip()


def load_psgc():
    """Return (by_key, by_code) over every PSGC entry above barangay level.

    by_key maps (2-digit region code, normalised name) -> row dict, which is enough to
    disambiguate the repeated city names (San Carlos, Naga, Talisay, San Fernando) without
    any fuzzy matching.
    """
    wb = openpyxl.load_workbook(PSGC_XLSX, read_only=True, data_only=True)
    ws = wb["PSGC"]
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header
    # Name collisions inside one region are real: Isabela province contains a municipality
    # called Quirino and Laguna one called Rizal, and both of those are also province names.
    # Rank by level so the province always wins its own name.
    rank = {"Reg": 0, "Prov": 1, "": 1, "None": 1, "SGU": 2, "City": 3, "Mun": 4, "SubMun": 5}
    by_key = {}
    by_code = {}
    for r in rows:
        code, name, level = r[0], r[1], r[3]
        if code is None or name is None:
            continue
        code = str(code).strip()
        if not re.fullmatch(r"\d{10}", code):
            continue
        level = "" if level is None else str(level).strip()
        if level == "Bgy":
            continue
        rec = {"code": code, "name": str(name).strip(), "level": level}
        by_code[code] = rec
        k = (code[:2], norm(name))
        prev = by_key.get(k)
        if prev is None or rank.get(level, 9) < rank.get(prev["level"], 9):
            by_key[k] = rec
    wb.close()
    return by_key, by_code


def load_census():
    """Return (categories, units).

    categories: list of the PSA's column headings, in file order.
    units: list of dicts {label, household_population, counts: [...]}, region headers included,
           in file order, so the region a unit sits under can be recovered by position.
    """
    wb = openpyxl.load_workbook(CENSUS_XLSX, read_only=True, data_only=True)
    ws = wb["A"]
    rows = [r for r in ws.iter_rows(values_only=True)]
    wb.close()

    header = rows[3]
    categories = []
    for j in range(2, len(header)):
        h = header[j]
        if h is None:
            raise SystemExit("blank category heading at column %d" % j)
        categories.append(re.sub(r"\s+", " ", str(h)).strip())

    units = []
    for r in rows[5:]:
        label = r[0]
        if label is None or str(label).strip() == "":
            continue
        label = re.sub(r"\s+", " ", str(label)).strip()
        if label.startswith("Note:") or label.startswith("Source:") or label.startswith("1 A "):
            continue
        hp = r[1]
        if hp is None:
            continue
        # "Interim Province 1" carries a trailing footnote marker; no PH place name ends in
        # a bare digit, so stripping one is safe and keeps geo_name clean.
        label = re.sub(r"\s+\d$", "", label)
        counts = []
        for j in range(2, 2 + len(categories)):
            v = r[j] if j < len(r) else None
            counts.append(0 if v is None else int(v))
        units.append({"label": label, "hp": int(hp), "counts": counts})
    return categories, units


def resolve(units, by_key, by_code):
    """Attach a PSGC code and a geo_level to every census row.

    Walks the file in order: "Philippines" is the country row, a row whose PSGC entry is a
    region opens a new region, and every row after it is looked up inside that region.
    """
    resolved = []
    region2 = None
    for u in units:
        key = norm(u["label"])
        if key == "philippines":
            resolved.append(dict(u, code="", level="country", psgc_name="Philippines"))
            continue

        code = LABEL_OVERRIDES.get(key)
        if code is None:
            hit = by_key.get((region2, key)) if region2 is not None else None
            if hit is None:
                # a region header: match across every region prefix, but only against Reg rows
                for pref in sorted({c[:2] for c in by_code}):
                    cand = by_key.get((pref, key))
                    if cand is not None and cand["level"] == "Reg":
                        hit = cand
                        break
            if hit is None:
                raise SystemExit("unresolved census row label: %r" % u["label"])
            code = hit["code"]

        rec = by_code[code]
        level = LEVEL_OF_PSGC_LEVEL.get(rec["level"])
        if level is None:
            # the two province-slot specials: "City of Isabela (Not a Province)" and the
            # interim BARMM province, both of which carry a blank Geographic Level
            level = "province"
        if code == "1999900000":
            level = "province"
        if rec["level"] == "Reg":
            region2 = code[:2]
        resolved.append(dict(u, code=code, level=level, psgc_name=rec["name"]))
    return resolved


def check(categories, resolved):
    """Reconcile against the PSA press release and against the file's own arithmetic."""
    ok = True
    idx = {c: i for i, c in enumerate(categories)}
    nat = [u for u in resolved if u["level"] == "country"][0]

    def cat(name):
        return nat["counts"][idx[name]]

    print("national reconciliation, PSA press release vs sheet A")
    checks = [
        ("household population", nat["hp"], PUBLISHED["household_population"]),
        # The press release's headline 85,645,362 (78.8%) carries the footnote "*Excluding
        # Catholic Charismatics numbering 74,096" — so it is this one column, not the pair.
        (
            "Roman Catholic (excl. Charismatic)",
            cat("Roman Catholic, excluding Catholic Charismatics"),
            PUBLISHED["roman_catholic_excl_charismatic"],
        ),
        ("Catholic Charismatic", cat("Catholic Charismatic"), PUBLISHED["catholic_charismatic"]),
        ("Islam", cat("Islam"), PUBLISHED["islam"]),
        ("Iglesia ni Cristo", cat("Iglesia ni Cristo"), PUBLISHED["iglesia_ni_cristo"]),
        (
            "Seventh Day Adventist",
            cat("Seventh Day Adventist"),
            PUBLISHED["seventh_day_adventist"],
        ),
        ("Aglipay", cat("Aglipay"), PUBLISHED["aglipay"]),
        (
            "Iglesia Filipina Independiente",
            cat("Iglesia Filipina Independiente"),
            PUBLISHED["iglesia_filipina_independiente"],
        ),
        ("Bible Baptist Church", cat("Bible Baptist Church"), PUBLISHED["bible_baptist_church"]),
        (
            "United Church of Christ in the Philippines",
            cat("United Church of Christ in the Philippines"),
            PUBLISHED["uccp"],
        ),
        ("Jehovah's Witness", cat("Jehovah's Witness"), PUBLISHED["jehovahs_witness"]),
        ("Church of Christ", cat("Church of Christ"), PUBLISHED["church_of_christ"]),
        ("None", cat("None"), PUBLISHED["none"]),
        ("Not reported", cat("Not reported"), PUBLISHED["not_reported"]),
    ]
    for name, got, want in checks:
        flag = "ok " if got == want else "XX "
        if got != want:
            ok = False
        print("  %s%-45s %12d  published %12d  diff %+d" % (flag, name, got, want, got - want))

    # the columns must sum to the household population, per unit
    worst = []
    for u in resolved:
        d = sum(u["counts"]) - u["hp"]
        if d:
            worst.append((abs(d), d, u["label"]))
    if worst:
        ok = False
        worst.sort(reverse=True)
        print("  XX %d units whose categories do not sum to household population" % len(worst))
        for _, d, lab in worst[:10]:
            print("       %-50s %+d" % (lab, d))
    else:
        print("  ok  every unit's categories sum exactly to its household population")

    # the fine tier must partition the country
    fine = [u for u in resolved if u["level"] in ("province", "city", "municipality")]
    regions = [u for u in resolved if u["level"] == "region"]
    tot_fine = sum(u["hp"] for u in fine)
    tot_reg = sum(u["hp"] for u in regions)
    print(
        "  %s fine tier %d units, %d people; regions %d units, %d people; national %d"
        % (
            "ok " if tot_fine == tot_reg == nat["hp"] else "XX ",
            len(fine),
            tot_fine,
            len(regions),
            tot_reg,
            nat["hp"],
        )
    )
    if not (tot_fine == tot_reg == nat["hp"]):
        ok = False

    print(
        "  note household population %d is %d short of the 2020 CPH total population %d (%.3f%%)"
        % (
            nat["hp"],
            PUBLISHED["total_population"] - nat["hp"],
            PUBLISHED["total_population"],
            100.0 * (PUBLISHED["total_population"] - nat["hp"]) / PUBLISHED["total_population"],
        )
    )
    return ok


def main():
    categories, units = load_census()
    by_key, by_code = load_psgc()
    resolved = resolve(units, by_key, by_code)

    print(
        "%d categories, %d geographic rows (%d country, %d region, %d province, %d city, %d municipality)"
        % (
            len(categories),
            len(resolved),
            sum(1 for u in resolved if u["level"] == "country"),
            sum(1 for u in resolved if u["level"] == "region"),
            sum(1 for u in resolved if u["level"] == "province"),
            sum(1 for u in resolved if u["level"] == "city"),
            sum(1 for u in resolved if u["level"] == "municipality"),
        )
    )
    ok = check(categories, resolved)

    if not os.path.isdir(OUT_DIR):
        os.makedirs(OUT_DIR)
    n = 0
    with open(OUT, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(
            [
                "geo_id",
                "geo_level",
                "geo_name",
                "source_category",
                "count",
                "basis",
                "year",
                "source_id",
                "note",
            ]
        )
        for u in resolved:
            # "PH" is not a PSGC code — the PSGC has none for the nation.
            geo_id = u["code"] if u["code"] else "PH"
            base_note = AGGREGATE_NOTE.get(u["level"], "")
            unit_note = EXTRA_NOTES.get(geo_id, "")
            rows = [
                (
                    "Household Population",
                    u["hp"],
                    "denominator, not a religious category; equals the sum of the other "
                    "rows for this geo_id",
                )
            ]
            for c, v in zip(categories, u["counts"]):
                rows.append((c, v, ""))
            for cat_name, val, extra in rows:
                note = "; ".join(x for x in (base_note, unit_note, extra) if x)
                w.writerow(
                    [
                        geo_id,
                        u["level"],
                        u["label"],
                        cat_name,
                        val,
                        BASIS,
                        YEAR,
                        SOURCE_ID,
                        note,
                    ]
                )
                n += 1
    print("wrote %s  (%d rows)" % (OUT, n))
    if not ok:
        sys.exit("reconciliation failed — see the XX lines above")


if __name__ == "__main__":
    main()
