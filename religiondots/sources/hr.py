"""Croatia — DZS, Popis 2021, religion, down to town/municipality.

Reads (or fetches) data/raw/hr/ and writes data/normalized/hr.csv.

12 categories over **556 towns and municipalities plus the 17 gradske četvrti of Zagreb**,
for 3.87M people. The city districts matter: Zagreb is 19.8% of Croatia and DZS publishes
religion for its 17 districts, so — like Czechia and Estonia and unlike Poland and Romania
— the capital does not have to be one polygon.

The category list is shallow (`Katolici`, `Pravoslavci`, `Protestanti`, `Muslimani`,
`Židovi`, `Istočne religije`…) and Croatia is 79% Catholic, so on its own this is a
two-colour map. The interest is entirely in the Serb Orthodox belt along the Bosnian and
Serbian borders and in Istria's unusually large irreligious share.

**THE REAL PRIZE IS SHEET 5 AND IT IS NOT INGESTED HERE — see `sources/hr.md` §4.** DZS
publishes a second table naming **54 individual churches** at the same geography, including
four Orthodox jurisdictions kept apart and eleven separate Jewish communities. It refines
two of sheet 2's residual categories rather than replacing the partition, so folding the two
together needs care that this file does not take; it is the largest available upgrade for
Croatia and it is deliberately deferred.

NO GEOGRAPHIC CODES. Like Romania, the workbook identifies rows by name only, so `geo_id`
is "COUNTY|NAME" and `sources/hr_geo.py` resolves it against the Eurostat LAU-NUTS
correspondence table. Same machinery, same reasons.

Usage:
    python sources/hr.py --fetch    download the workbook (18MB) if missing
    python sources/hr.py            normalise from data/raw/hr/
"""

import csv
import os
import sys

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
RAW = os.path.join(ROOT, "data", "raw", "hr")
OUT = os.path.join(ROOT, "data", "normalized", "hr.csv")

SOURCE_ID = "hr_popis_2021"
YEAR = 2021
BASIS = "self_id"

XLSX_NAME = "gradovi_opcine.xlsx"
XLSX_URL = ("https://podaci.dzs.hr/media/td3jvrbu/"
            "popis_2021-stanovnistvo_po_gradovima_opcinama.xlsx")
MIN_BYTES = 10_000_000

COLUMNS = ["geo_id", "geo_level", "geo_name", "source_category", "count",
           "basis", "year", "source_id", "note"]

NATIONAL = 3_871_833          # DZS, total population, census 2021
SHEET = "2."
HEADER_ROW = 8
FIRST_DATA_ROW = 9

COUNTY_COL, KIND_COL, NAME_COL, TOTAL_COL = 0, 1, 4, 5
NATIONAL_NAME = "Republika Hrvatska"

# Count columns; the odd column after each is the same figure as a percentage and is
# dropped. Taking the wrong one of a pair yields a map where every unit has ~100 people.
CAT_COLS = [7, 9, 11, 13, 15, 17, 19, 21, 23, 25, 27, 29]

TOTAL_LABEL = "Ukupno"
TOTAL_NOTE = "universe total, not a religion category"

KIND_LEVEL = {"Grad": "municipality", "Općina": "municipality",
              "Gradska četvrt": "city_district"}

TRUE_ZERO = ("-", "–", "—")


def fetch():
    import requests
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    os.makedirs(RAW, exist_ok=True)
    dest = os.path.join(RAW, XLSX_NAME)
    if os.path.exists(dest) and os.path.getsize(dest) >= MIN_BYTES:
        print("already have", dest)
        return
    print("downloading", XLSX_URL)
    r = requests.get(XLSX_URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=900,
                     verify=False)
    r.raise_for_status()
    with open(dest, "wb") as fh:
        fh.write(r.content)
    size = os.path.getsize(dest)
    if size < MIN_BYTES:
        raise SystemExit(f"{dest} is {size:,} bytes, expected >= {MIN_BYTES:,}")
    print(f"  {size:,} bytes")


def _txt(v):
    return "" if v is None else " ".join(str(v).split())


def _num(cell, where):
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return int(cell)
    s = str(cell).strip()
    if s in TRUE_ZERO:
        return 0
    if s == "":
        return None
    raise SystemExit(f"unexpected value {cell!r} in a count column at {where} -- "
                     "DZS changed the sheet, check for a new sentinel")


def _label(header, col):
    """DZS puts Croatian and English in one cell; keep the Croatian, which is the key."""
    raw = _txt(header[col])
    # e.g. 'Katolici Catholics', 'Ostali kršćani1) Other Christians1)'
    for marker in (" Other ", " Catholics", " Orthodox", " Protestants", " Muslims",
                   " Jews", " Oriental ", " Agnostics", " Not religious", " Not declared",
                   " Unknown"):
        i = raw.find(marker)
        if i > 0:
            return raw[:i].strip()
    return raw


def read():
    import openpyxl

    path = os.path.join(RAW, XLSX_NAME)
    if not os.path.exists(path):
        raise SystemExit(f"missing {path} -- run with --fetch first")
    if os.path.getsize(path) < MIN_BYTES:
        raise SystemExit(f"{path} is truncated, re-run with --fetch")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"no sheet {SHEET!r}; got {wb.sheetnames}")
    ws = wb[SHEET]

    header = list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW,
                               values_only=True))[0]
    cats = {c: _label(header, c) for c in CAT_COLS}
    if any(not v for v in cats.values()):
        raise SystemExit(f"blank category header: {cats}")

    rows = []
    current_county = None
    for ri, r in enumerate(ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True),
                           FIRST_DATA_ROW):
        county, kind, name = _txt(r[COUNTY_COL]), _txt(r[KIND_COL]), _txt(r[NAME_COL])
        total = _num(r[TOTAL_COL], f"row {ri}")
        if total is None:
            continue
        if county:
            current_county = county

        if county == NATIONAL_NAME:
            gid, level, gname = "HR", "country", "Republika Hrvatska"
        elif not kind:
            gid, level, gname = current_county, "county", current_county
        else:
            level = KIND_LEVEL.get(kind)
            if level is None:
                raise SystemExit(f"row {ri}: unknown unit kind {kind!r}")
            if not name:
                raise SystemExit(f"row {ri}: {kind} with no name")
            gid, gname = f"{current_county}|{name}", name

        note = f"DZS Popis 2021 sheet {SHEET}"
        rows.append({"geo_id": gid, "geo_level": level, "geo_name": gname,
                     "source_category": TOTAL_LABEL, "count": total, "basis": BASIS,
                     "year": YEAR, "source_id": SOURCE_ID,
                     "note": note + "; " + TOTAL_NOTE})
        for c in CAT_COLS:
            n = _num(r[c], f"row {ri} col {c}")
            if n:
                rows.append({"geo_id": gid, "geo_level": level, "geo_name": gname,
                             "source_category": cats[c], "count": n, "basis": BASIS,
                             "year": YEAR, "source_id": SOURCE_ID, "note": note})
    return rows


def check(rows):
    ok = True
    levels, totals, parts = {}, {}, {}
    for r in rows:
        k = (r["geo_level"], r["geo_id"])
        levels.setdefault(r["geo_level"], set()).add(r["geo_id"])
        if r["source_category"] == TOTAL_LABEL:
            totals[k] = r["count"]
        else:
            parts[k] = parts.get(k, 0) + r["count"]

    # GRAD ZAGREB IS NOT IN THE MUNICIPALITY LIST. DZS has already done for Croatia what
    # cz_geo.py and ee_geo.py have to do by hand: the capital appears ONLY as its 17
    # gradske četvrti, so `municipality` is 555 units rather than Croatia's 556, and the
    # complete cover is the two levels TOGETHER. Checking `municipality` against the
    # national total therefore fails by exactly Zagreb's 767,131, which is the correct
    # answer to the wrong question.
    expected = {"country": 1, "county": 21, "municipality": 555, "city_district": 17}
    print("  units and population per level — levels are alternatives, never summed:")
    for lv in ("municipality", "city_district", "county", "country"):
        n_units = len(levels.get(lv, ()))
        pop = sum(v for (l, _), v in totals.items() if l == lv)
        full = lv in ("county", "country")
        good = n_units == expected[lv] and (pop == NATIONAL if full else True)
        ok &= good
        tail = "" if full else "   <- partial; the two together are the cover"
        print(f"    {'OK ' if good else 'BAD'} {lv:<14} {n_units:>5,} units  "
              f"{pop:>10,}  (expected {expected[lv]:,} units){tail}")

    drawn_units = len(levels.get("municipality", ())) + len(levels.get("city_district", ()))
    drawn_pop = sum(v for (l, _), v in totals.items()
                    if l in ("municipality", "city_district"))
    good = drawn_units == 572 and drawn_pop == NATIONAL
    ok &= good
    print(f"    {'OK ' if good else 'BAD'} {'DRAWN COVER':<14} {drawn_units:>5,} units  "
          f"{drawn_pop:>10,}  (555 municipalities + 17 Zagreb districts = 572)")

    # DZS neither rounds nor suppresses, so the categories must partition every unit exactly.
    bad = {k: (totals.get(k, 0), parts.get(k, 0)) for k in totals
           if parts.get(k, 0) != totals[k]}
    ok &= not bad
    print(f"\n  {'OK ' if not bad else 'BAD'} categories partition the total in all "
          f"{len(totals):,} units ({len(bad)} mismatched)")
    for k, (t, p) in list(bad.items())[:5]:
        print(f"      {k}: total {t:,} vs categories {p:,}")

    by_cat = {}
    for r in rows:
        if r["geo_level"] == "country" and r["source_category"] != TOTAL_LABEL:
            by_cat[r["source_category"]] = r["count"]
    print(f"\n  {len(rows):,} rows, {len(by_cat)} categories:")
    for name, n in sorted(by_cat.items(), key=lambda x: -x[1]):
        print(f"    {n:>10,}  {100.0 * n / NATIONAL:5.2f}%  {name}")

    if not ok:
        raise SystemExit("reconciliation FAILED")


def main():
    if "--fetch" in sys.argv:
        fetch()
    rows = read()
    check(rows)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)
    print("\nwrote", OUT)


if __name__ == "__main__":
    main()
